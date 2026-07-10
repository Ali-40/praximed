"""
build_50call_command_center.py

Builds the upgraded PraxisMed outreach command center for Ali's Monday calling system.

Outputs (all written to docs/sales/outreach/START_HERE_FOR_ALI/):
  OUTREACH_COMMAND_CENTER.xlsx  — 10-sheet master workbook
  1_MONDAY_CALLS.xlsx           — standalone 50-call list
  2_SUNDAY_EMAILS.xlsx          — standalone 50-email candidate list
  0_READ_ME_FIRST.md            — updated daily guide
  4_CALL_SCRIPT_PRINTABLE.md    — updated call scripts with Alex wording
  5_EMAIL_TEMPLATES_PRINTABLE.md — both email versions

Specialty mix for Monday Calls (50 total):
  child_adolescent_psychiatry   7
  adult_psychiatry              8
  dermatology                   10
  gynecology                    10
  orthopedics                   7
  internal_medicine             5
  ent + neurology + urology     3

Safety constraints (never remove):
  No auto-email. No auto-call. No mass-send. No private data. No patient data. No PHI.
  Sender identity: "Alex, PraxisMed" is a communication alias for Ali Abdeltawab.
  NOT a fake person. Ali can explain this to anyone who asks.
"""

from __future__ import annotations

import csv
import os
import sys
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.dirname(os.path.dirname(_HERE))
_LEADS_CSV = os.path.join(_REPO, "docs", "sales", "outreach", "praxisplan_all_high_potential_leads.csv")
_OUT_DIR = os.path.join(_REPO, "docs", "sales", "outreach", "START_HERE_FOR_ALI")

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------

NAVY     = "1F4E79"
TEAL     = "117A65"
ORANGE   = "CA6F1E"
PURPLE   = "5B2C6F"
GRAY_LT  = "F2F3F4"
GRAY_MD  = "D5D8DC"
WHITE    = "FFFFFF"
YELLOW   = "FFF3CD"
RED_LT   = "FADBD8"

# ---------------------------------------------------------------------------
# Specialty mix for Monday Calls
# ---------------------------------------------------------------------------

SPECIALTY_MIX: list[tuple[str, int]] = [
    ("child_adolescent_psychiatry", 7),
    ("adult_psychiatry", 8),
    ("dermatology", 10),
    ("gynecology", 10),
    ("orthopedics", 7),
    ("internal_medicine", 5),
    ("ent", 1),
    ("neurology", 1),
    ("urology", 1),
]

# ---------------------------------------------------------------------------
# German pain points per specialty
# ---------------------------------------------------------------------------

PAIN_POINTS: dict[str, str] = {
    "child_adolescent_psychiatry": "Terminanfragen für Kinder- und Jugendpsychiatrie sind komplex — Eltern rufen häufig an und warten auf Rückrufe.",
    "adult_psychiatry": "Psychotherapiepraxen erhalten viele Anfragen. Rückrufe gehen unter, Patienten wenden sich woanders hin.",
    "dermatology": "Dermatologische Privatpraxen haben viele Terminanfragen. Rückrufwünsche gehen ohne System leicht verloren.",
    "gynecology": "Gynäkologische Praxen bekommen viele Anrufe. Eine klare Rückrufliste spart Zeit für das Praxisteam.",
    "orthopedics": "Orthopädische Praxen haben hohe Anrufvolumen, besonders für Erstkonsultationen.",
    "internal_medicine": "Internistische Privatpraxen erhalten viele unstrukturierte Anfragen, die ein einfaches System klären kann.",
    "ent": "HNO-Praxen haben oft lange Wartelisten. Rückrufanfragen sauber aufzunehmen spart Zeit.",
    "neurology": "Neurologische Praxen erhalten komplexe Terminanfragen. Klare Rückruflisten helfen dem Team.",
    "urology": "Urologische Privatpraxen profitieren von einer übersichtlichen Rückruf- und Anfrageverwaltung.",
    "pediatrics": "Kinderarztpraxen erhalten viele Anrufe von besorgten Eltern. Klare Rückruflisten helfen.",
    "aesthetic_medicine": "Ästhetische Praxen haben Anfragen für Beratungstermine, die sauber organisiert werden müssen.",
    "ophthalmology": "Augenarztpraxen haben lange Wartelisten. Rückruf-Organisation spart Telefonzeit.",
    "private_dental": "Privatzahnarztpraxen profitieren von einer klaren Anfrageverwaltung.",
    "plastic_surgery": "Plastisch-chirurgische Praxen erhalten viele Beratungsanfragen — strukturiert erfassen spart Zeit.",
    "private_group_practices": "Gruppenpraxen haben hohe Anrufvolumen und profitieren von einer zentralen Rückrufliste.",
}

SUBJECTS: dict[str, str] = {
    "child_adolescent_psychiatry": "Rückruf-Anfragen für Ihre kinder- und jugendpsychiatrische Praxis",
    "adult_psychiatry": "Terminanfragen in Ihrer Psychotherapiepraxis besser organisieren",
    "dermatology": "Weniger verpasste Rückrufe für Ihre Dermatologie-Praxis",
    "gynecology": "Rückruf-Anfragen für Ihre gynäkologische Praxis übersichtlich verwalten",
    "orthopedics": "Terminanfragen in Ihrer orthopädischen Praxis sauber aufnehmen",
    "internal_medicine": "Rückruf-Organisation für Ihre internistische Privatpraxis",
    "ent": "Anfragen für Ihre HNO-Praxis klar organisieren",
    "neurology": "Rückruf-Anfragen für Ihre neurologische Praxis",
    "urology": "Anfragen für Ihre urologische Privatpraxis übersichtlich erfassen",
    "pediatrics": "Eltern-Rückrufe für Ihre Kinderarztpraxis besser verwalten",
    "aesthetic_medicine": "Beratungsanfragen für Ihre Praxis sauber aufnehmen",
    "ophthalmology": "Rückruf-Anfragen für Ihre augenärztliche Praxis",
    "private_dental": "Anfragen für Ihre Privatzahnarztpraxis organisieren",
    "plastic_surgery": "Beratungsanfragen für Ihre chirurgische Praxis erfassen",
    "private_group_practices": "Rückruf-Organisation für Ihre Gruppenpraxis",
}

# ---------------------------------------------------------------------------
# Lead loading and selection
# ---------------------------------------------------------------------------

def _load_leads() -> list[dict[str, str]]:
    leads = []
    with open(_LEADS_CSV, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            leads.append(row)
    return leads


def _eligible(lead: dict[str, str]) -> bool:
    status = (lead.get("Outreach Status") or "").lower().strip()
    if status in ("do not contact", "not interested", "demo booked"):
        return False
    if not (lead.get("Phone") or "").strip():
        return False
    if (lead.get("City") or "").strip().lower() != "wien":
        return False
    return True


def _score(lead: dict[str, str]) -> int:
    return int(lead.get("Priority Score") or 0)


def _select_call_leads(leads: list[dict[str, str]]) -> list[dict[str, str]]:
    eligible = [l for l in leads if _eligible(l)]
    selected: list[dict[str, str]] = []
    used_ids: set[str] = set()

    for specialty_key, count in SPECIALTY_MIX:
        pool = sorted(
            [l for l in eligible if l.get("Specialty Key") == specialty_key and l.get("Lead ID") not in used_ids],
            key=_score,
            reverse=True,
        )
        picked = pool[:count]
        selected.extend(picked)
        used_ids.update(l["Lead ID"] for l in picked)

    return selected


def _select_email_leads(leads: list[dict[str, str]], call_ids: set[str]) -> list[dict[str, str]]:
    eligible = [
        l for l in leads
        if _eligible(l)
        and l.get("Lead ID") not in call_ids
        and (l.get("Praxisplan Profile URL") or "").startswith("http")
    ]
    return sorted(eligible, key=_score, reverse=True)[:50]


# ---------------------------------------------------------------------------
# Email draft generator
# ---------------------------------------------------------------------------

def _email_draft_alex(lead: dict[str, str]) -> str:
    specialty_key = lead.get("Specialty Key", "")
    pain = PAIN_POINTS.get(specialty_key, "Telefonanfragen und Rückrufe werden schnell unübersichtlich.")
    return (
        f"Guten Tag,\n\n"
        f"mein Name ist Alex von PraxisMed.\n\n"
        f"{pain}\n\n"
        f"PraxisMed ist eine einfache KI-Rezeption für Wiener Privatpraxen. "
        f"Es ersetzt kein bestehendes Praxissystem wie LATIDO oder Ihren Kalender. "
        f"Es hilft davor: Anrufe und Rückruf-Anfragen werden sauber aufgenommen "
        f"und für Ihr Praxisteam übersichtlich sortiert.\n\n"
        f"Die KI bestätigt keine Termine, stellt keine Diagnosen und gibt keine "
        f"medizinische Beratung. Ihr Team bleibt immer in Kontrolle.\n\n"
        f"Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?\n\n"
        f"Mit freundlichen Grüßen\n"
        f"Alex\n"
        f"PraxisMed"
    )


def _subject_for(lead: dict[str, str]) -> str:
    sk = lead.get("Specialty Key", "")
    return SUBJECTS.get(sk, "Rückruf-Anfragen in Ihrer Praxis einfacher organisieren")


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(wrap=False, h="left", v="center") -> Alignment:
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)


def _dv(formula: str, sqref: str) -> DataValidation:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = sqref
    return dv


def _header_row(ws, headers: list[str], color: str = NAVY, row: int = 1) -> None:
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = _fill(color)
        cell.font = _font(bold=True, color=WHITE)
        cell.alignment = _align(h="center")


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _row_fill(ws, row: int, ncols: int, color: str) -> None:
    for ci in range(1, ncols + 1):
        ws.cell(row=row, column=ci).fill = _fill(color)


# ---------------------------------------------------------------------------
# Build Monday Calls sheet
# ---------------------------------------------------------------------------

MON_HEADERS = [
    "Call Order", "Lead ID", "Doctor Name", "Specialty", "Phone",
    "District", "Priority Score", "Likely Objection", "Opening Line",
    "Outreach Status", "Call Result", "Contact Person", "Follow-up Date",
    "Demo Offered", "Demo Booked", "Notes",
]

MON_WIDTHS = {
    1: 10, 2: 12, 3: 28, 4: 28, 5: 20,
    6: 22, 7: 14, 8: 35, 9: 50,
    10: 18, 11: 18, 12: 20, 13: 16,
    14: 14, 15: 14, 16: 35,
}

STATUS_DV      = '"Not contacted,Called – no answer,Called – busy,Left voicemail,Spoke to reception,Interested,Not interested,Do not contact,Follow-up needed"'
CALL_RESULT_DV = '"—,No answer,Busy,Voicemail,Reception took message,Spoke to doctor,Interested – demo,Not interested,Call later,Wrong number"'
YESNO_DV       = '"No,Yes"'

OBJECTIONS: dict[str, str] = {
    "child_adolescent_psychiatry": "Wir haben kein System",
    "adult_psychiatry": "Wir nutzen LATIDO",
    "dermatology": "Wir sind ausgebucht",
    "gynecology": "Kein Budget",
    "orthopedics": "Wir nutzen LATIDO",
    "internal_medicine": "Kein Bedarf",
    "ent": "Wir haben eine Rezeption",
    "neurology": "Zu beschäftigt",
    "urology": "Kein Budget",
}

OPENING_LINE_TMPL = (
    "Guten Tag, mein Name ist Ali Abdeltawab, ich baue PraxisMed. "
    "Darf ich kurz fragen — wie organisieren Sie aktuell Rückrufwünsche?"
)


def _build_monday_calls_sheet(ws, call_leads: list[dict[str, str]]) -> None:
    ws.cell(row=1, column=1).value = "MONDAY TARGET: 50 calls. Update status immediately after every call."
    ws.cell(row=1, column=1).fill = _fill(ORANGE)
    ws.cell(row=1, column=1).font = _font(bold=True, color=WHITE, size=12)
    ws.cell(row=1, column=1).alignment = _align(h="center")
    ws.merge_cells(f"A1:{get_column_letter(len(MON_HEADERS))}1")

    _header_row(ws, MON_HEADERS, color=NAVY, row=2)

    for i, lead in enumerate(call_leads, 1):
        row = i + 2
        sk = lead.get("Specialty Key", "")
        data = [
            i,
            lead.get("Lead ID", ""),
            lead.get("Doctor Name", ""),
            lead.get("Specialty Label DE") or lead.get("Specialty", ""),
            lead.get("Phone", ""),
            lead.get("District", ""),
            int(lead.get("Priority Score") or 0),
            OBJECTIONS.get(sk, "Kein Bedarf"),
            OPENING_LINE_TMPL,
            "Not contacted",
            "—",
            "",
            "",
            "No",
            "No",
            "",
        ]
        fill_color = GRAY_LT if i % 2 == 0 else WHITE
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = _fill(fill_color)
            cell.alignment = _align(wrap=(ci == 9))

    # Dropdowns (rows 3 to 52)
    ws.add_data_validation(_dv(STATUS_DV, f"J3:J{len(call_leads)+2}"))
    ws.add_data_validation(_dv(CALL_RESULT_DV, f"K3:K{len(call_leads)+2}"))
    ws.add_data_validation(_dv(YESNO_DV, f"N3:N{len(call_leads)+2}"))
    ws.add_data_validation(_dv(YESNO_DV, f"O3:O{len(call_leads)+2}"))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(MON_HEADERS))}{len(call_leads)+2}"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    _set_col_widths(ws, MON_WIDTHS)


# ---------------------------------------------------------------------------
# Build Sunday Emails sheet
# ---------------------------------------------------------------------------

SUN_HEADERS = [
    "Email Order", "Lead ID", "Doctor Name", "Specialty",
    "Praxisplan Profile URL", "Subject", "Email Draft (Alex / PraxisMed Team)",
    "Manual Send Status", "Reply Status", "Follow-up Date", "Notes",
]

SUN_WIDTHS = {
    1: 12, 2: 12, 3: 28, 4: 28,
    5: 45, 6: 50, 7: 80,
    8: 20, 9: 18, 10: 16, 11: 30,
}

SEND_STATUS_DV  = '"Not sent,Sent – contact form,Sent – email,Sent – LinkedIn,Skipped – no contact"'
REPLY_STATUS_DV = '"—,No reply,Replied – interested,Replied – not interested,Replied – asked question,Replied – Do not contact"'


def _build_sunday_emails_sheet(ws, email_leads: list[dict[str, str]]) -> None:
    ws.cell(row=1, column=1).value = (
        "MANUAL REVIEW ONLY — Copy one email at a time. No mass-send. No auto-send. "
        "Use the Praxisplan Profile URL to find the contact form or email address."
    )
    ws.cell(row=1, column=1).fill = _fill(RED_LT)
    ws.cell(row=1, column=1).font = _font(bold=True, color="7B241C", size=11)
    ws.cell(row=1, column=1).alignment = _align(wrap=True, h="center")
    ws.merge_cells(f"A1:{get_column_letter(len(SUN_HEADERS))}1")
    ws.row_dimensions[1].height = 30

    _header_row(ws, SUN_HEADERS, color=TEAL, row=2)

    for i, lead in enumerate(email_leads, 1):
        row = i + 2
        draft = _email_draft_alex(lead)
        subject = _subject_for(lead)
        fill_color = GRAY_LT if i % 2 == 0 else WHITE
        data = [
            i,
            lead.get("Lead ID", ""),
            lead.get("Doctor Name", ""),
            lead.get("Specialty Label DE") or lead.get("Specialty", ""),
            lead.get("Praxisplan Profile URL", ""),
            subject,
            draft,
            "Not sent",
            "—",
            "",
            "",
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = _fill(fill_color)
            cell.alignment = _align(wrap=(ci in (6, 7)), v="top")
            if ci == 7:
                ws.row_dimensions[row].height = 100

    ws.add_data_validation(_dv(SEND_STATUS_DV, f"H3:H{len(email_leads)+2}"))
    ws.add_data_validation(_dv(REPLY_STATUS_DV, f"I3:I{len(email_leads)+2}"))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(SUN_HEADERS))}{len(email_leads)+2}"
    ws.row_dimensions[2].height = 20
    _set_col_widths(ws, SUN_WIDTHS)


# ---------------------------------------------------------------------------
# Build Sender Identity sheet
# ---------------------------------------------------------------------------

def _build_sender_identity_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False

    def _section(row: int, title: str, color: str) -> None:
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = _fill(color)
        cell.font = _font(bold=True, color=WHITE, size=12)
        cell.alignment = _align(h="left")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 22

    def _row(row: int, label: str, value: str, value_bold=False) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True, size=11)
        lc.alignment = _align()
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _font(bold=value_bold, size=11)
        vc.alignment = _align(wrap=True)
        ws.merge_cells(f"B{row}:D{row}")
        ws.row_dimensions[row].height = 20

    def _note(row: int, text: str, color=YELLOW) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = _fill(color)
        cell.font = Font(italic=True, size=10, name="Calibri", color="7D6608")
        cell.alignment = _align(wrap=True, h="left")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 36

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    _section(1, "SENDER IDENTITY — PraxisMed Outreach", NAVY)

    _note(2, (
        "Ali Abdeltawab is the founder of PraxisMed. He also uses the name 'Alex' "
        "for clinic outreach because it is easier to pronounce. This is NOT a fake person. "
        "If anyone asks, Ali can explain: 'My full name is Ali Abdeltawab — I also go by Alex "
        "for easier communication.' Legal identity is never hidden."
    ))

    _section(4, "Primary Brand Identity", TEAL)
    _row(5, "Sender Name",       "PraxisMed Team")
    _row(6, "Email Display",     "Alex, PraxisMed")
    _row(7, "Sign-off (email)",  "Alex\nPraxisMed")
    _row(8, "Phone intro",       "Guten Tag, mein Name ist Alex von PraxisMed.")

    _section(10, "Transparent Identity (use when asked)", PURPLE)
    _row(11, "Full legal name",  "Ali Abdeltawab")
    _row(12, "Explanation",      "Mein vollständiger Name ist Ali Abdeltawab — ich verwende für Praxiskontakte auch Alex.")
    _row(13, "Sign-off (email)", "Ali Abdeltawab (Alex)\nPraxisMed")

    _section(15, "How to set up in Gmail", ORANGE)
    _note(16, (
        "Gmail → Settings (gear icon) → See all settings → Accounts and Import → "
        "Send mail as → Add another email address → Enter 'Alex, PraxisMed' and your Gmail address."
    ))

    _section(18, "What is and is not allowed", "C0392B")
    _note(19, (
        "ALLOWED: Use 'Alex, PraxisMed' as sender display name.  "
        "ALLOWED: Sign emails as 'Alex, PraxisMed'.  "
        "NOT ALLOWED: Claim to be a different person if directly asked.  "
        "NOT ALLOWED: Use a fake surname or fabricated identity.  "
        "ALWAYS: Respond honestly if a clinic asks for your full name."
    ), "FADBD8")


# ---------------------------------------------------------------------------
# Build Weekly Plan sheet
# ---------------------------------------------------------------------------

WEEKLY_PLAN = [
    ("Samstag",   "11.07.2026", "Vorbereitung",
     "• Call-Liste prüfen (Monday Calls Tab)\n• Email-Liste prüfen (Sunday Emails Tab)\n• Heute keine Anrufe – Praxen geschlossen",
     "Bereit für Sonntag"),
    ("Sonntag",   "12.07.2026", "Emails senden",
     "• Sunday Emails Tab öffnen\n• 50 Emails manuell senden – einzeln kopieren und einfügen\n• 'Manual Send Status' nach jedem Send aktualisieren\n• Kein Massen-Send. Kein Auto-Send.",
     "50 Emails gesendet"),
    ("Montag",    "13.07.2026", "Anrufe starten (ab 08:00)",
     "• Monday Calls Tab öffnen\n• Call Order 1, 2, 3 … folgen\n• Status nach jedem Anruf sofort eintragen\n• Ziel: 50 Anrufversuche",
     "50 Anrufversuche"),
    ("Dienstag",  "14.07.2026", "Nachfassen – Anrufe",
     "• Follow Ups Tab prüfen\n• Praxen zurückrufen, die 'Call later' gesagt haben\n• Neue Einträge im Master Leads Tab ergänzen",
     "25 Nachfass-Anrufe"),
    ("Mittwoch",  "15.07.2026", "Nachfassen – Emails",
     "• Email-Antworten prüfen\n• Interessierten Praxen Demo anbieten\n• 'Do not contact' sofort eintragen",
     "10 Antworten bearbeitet"),
    ("Donnerstag","16.07.2026", "Zweiter Anlauf",
     "• 'No answer' und 'Voicemail' Praxen erneut anrufen\n• Neue Praxen aus Master Leads für nächste Woche vorbereiten",
     "20 zweite Anrufversuche"),
]

WP_HEADERS = ["Tag", "Datum", "Aktivität", "Details", "Ziel"]


def _build_weekly_plan_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 26

    ws.cell(row=1, column=1, value="WOCHENPLAN — PraxisMed Outreach Sprint")
    ws.cell(row=1, column=1).fill = _fill(NAVY)
    ws.cell(row=1, column=1).font = _font(bold=True, color=WHITE, size=13)
    ws.cell(row=1, column=1).alignment = _align(h="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 28

    _header_row(ws, WP_HEADERS, color=TEAL, row=2)
    ws.row_dimensions[2].height = 20

    row_colors = [WHITE, GRAY_LT]
    call_colors = {
        "Samstag": GRAY_MD,
        "Sonntag": "D6EAF8",
        "Montag": "D5F5E3",
        "Dienstag": "FEF9E7",
        "Mittwoch": "FEF9E7",
        "Donnerstag": "FEF9E7",
    }

    for i, (tag, datum, activity, details, ziel) in enumerate(WEEKLY_PLAN):
        row = i + 3
        color = call_colors.get(tag, row_colors[i % 2])
        data = [tag, datum, activity, details, ziel]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = _fill(color)
            cell.font = _font(bold=(ci == 1), size=11)
            cell.alignment = _align(wrap=True, v="top")
        ws.row_dimensions[row].height = 72


# ---------------------------------------------------------------------------
# Build Follow Ups sheet
# ---------------------------------------------------------------------------

FOLLOWUP_HEADERS = [
    "Lead ID", "Doctor Name", "Specialty", "Phone", "Status",
    "Last Call Date", "Last Result", "Follow-up Date", "Notes",
]


def _build_followup_sheet(ws) -> None:
    _header_row(ws, FOLLOWUP_HEADERS, color=ORANGE)
    ws.cell(row=1, column=1).font = _font(bold=True, color=WHITE)
    ws.freeze_panes = "A2"
    widths = {1: 12, 2: 28, 3: 28, 4: 20, 5: 20, 6: 16, 7: 22, 8: 16, 9: 35}
    _set_col_widths(ws, widths)
    note = ws.cell(row=2, column=1,
                   value="This sheet is for manual tracking of leads that need a follow-up call or email.")
    note.fill = _fill(YELLOW)
    note.font = Font(italic=True, size=10, name="Calibri", color="7D6608")
    ws.merge_cells(f"A2:{get_column_letter(len(FOLLOWUP_HEADERS))}2")


# ---------------------------------------------------------------------------
# Build Master Leads sheet
# ---------------------------------------------------------------------------

MASTER_KEEP = [
    "Lead ID", "Specialty Key", "Specialty Label DE", "Doctor Name",
    "Practice Name", "Specialty", "Address", "Postal Code", "City", "District",
    "Phone", "Email", "Website", "Praxisplan Profile URL",
    "Priority Score", "Priority Reason",
    "Outreach Status", "Call Attempt 1 Date", "Call Attempt 1 Result",
    "Call Attempt 2 Date", "Call Attempt 2 Result",
    "Email Sent Date", "Email Result",
    "Follow-up Date", "Demo Offered", "Demo Booked", "Notes",
]


def _build_master_leads_sheet(ws, leads: list[dict[str, str]]) -> None:
    _header_row(ws, MASTER_KEEP, color=NAVY)
    for i, lead in enumerate(leads, 1):
        row = i + 1
        fill_color = GRAY_LT if i % 2 == 0 else WHITE
        for ci, key in enumerate(MASTER_KEEP, 1):
            cell = ws.cell(row=row, column=ci, value=lead.get(key, ""))
            cell.fill = _fill(fill_color)
            cell.alignment = _align()
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_KEEP))}1"
    for ci in range(1, len(MASTER_KEEP) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20


# ---------------------------------------------------------------------------
# Build Call Script sheet
# ---------------------------------------------------------------------------

CALL_SCRIPT_CONTENT = [
    ("Eröffnung", NAVY,
     "Guten Tag, mein Name ist Ali Abdeltawab — ich baue PraxisMed, "
     "eine einfache KI-Rezeption für Wiener Privatpraxen.\n\n"
     "Es ersetzt nicht Ihre Rezeption und auch nicht Ihr bestehendes Praxissystem. "
     "PraxisMed hilft, Terminanfragen und Rückrufe sauber zu organisieren.\n\n"
     "Darf ich Ihnen in 5 Minuten zeigen, wie aus einem Anruf eine Rückruf-Anfrage im Dashboard wird?"),

    ("Wenn: LATIDO / bestehendes System", TEAL,
     "Perfekt, dann müssen wir LATIDO nicht ersetzen.\n"
     "PraxisMed organisiert nur den Telefon- und Rückruf-Teil davor.\n"
     "Ihr Team trägt bestätigte Termine wie gewohnt in LATIDO oder Ihrem Praxissystem ein."),

    ("Wenn: Schicken Sie uns eine E-Mail", TEAL,
     "Gerne. An welche E-Mail-Adresse darf ich die kurze Info schicken?\n"
     "Ich halte es kurz — es geht um weniger verpasste Terminanfragen\n"
     "und eine klare Rückruf-Liste für Ihr Praxisteam.\n\n"
     "(E-Mail-Adresse notieren → Manuell senden. Kein Auto-Send.)"),

    ("Wenn: Kein Interesse", ORANGE,
     "Verstehe, danke für Ihre Zeit.\n"
     "Ich notiere, dass wir Sie nicht weiter kontaktieren.\n\n"
     "(→ Status: Do not contact. Sofort im Tracker eintragen.)"),

    ("Wenn: Was kostet das?", TEAL,
     "Für den Pilot starten wir einfach:\n"
     "30 Tage Testphase, Setup ab 390 Euro.\n"
     "Danach je nach Umfang etwa 290 bis 490 Euro monatlich für kleine Privatpraxen."),

    ("Wenn: Kalender / Buchung?", TEAL,
     "Im Pilot starten wir zuerst mit der Rückruf- und Anfrage-Übersicht.\n"
     "Der nächste Schritt ist ein einfacher Kalender-Workflow:\n"
     "Anfrage prüfen, Termin vorschlagen und vom Praxisteam bestätigen."),

    ("Wenn: Gibt KI medizinische Beratung / Diagnosen?", "C0392B",
     "Nein. PraxisMed gibt keine medizinische Beratung,\n"
     "stellt keine Diagnosen und bestätigt keine Termine automatisch.\n"
     "Ihr Team bleibt immer in Kontrolle."),
]


def _build_call_script_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 75

    ws.cell(row=1, column=1, value="TELEFON-SKRIPT — PraxisMed")
    ws.cell(row=1, column=1).fill = _fill(NAVY)
    ws.cell(row=1, column=1).font = _font(bold=True, color=WHITE, size=13)
    ws.cell(row=1, column=1).alignment = _align(h="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    row = 2
    for title, color, script in CALL_SCRIPT_CONTENT:
        tc = ws.cell(row=row, column=1, value=title)
        tc.fill = _fill(color)
        tc.font = _font(bold=True, color=WHITE, size=11)
        tc.alignment = _align(wrap=True, v="center")

        sc = ws.cell(row=row, column=2, value=script)
        sc.fill = _fill(GRAY_LT)
        sc.font = Font(size=11, name="Calibri")
        sc.alignment = _align(wrap=True, v="top")

        lines = script.count("\n") + 1
        ws.row_dimensions[row].height = max(60, lines * 16)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Nach jedem Anruf sofort eintragen:").fill = _fill(ORANGE)
    ws.cell(row=row, column=1).font = _font(bold=True, color=WHITE)
    ws.cell(row=row, column=2,
            value="Outreach Status | Call Result | Contact Person | Follow-up Date | Demo Offered | Demo Booked | Notes").fill = _fill(YELLOW)
    ws.row_dimensions[row].height = 22


# ---------------------------------------------------------------------------
# Build Email Templates sheet
# ---------------------------------------------------------------------------

EMAIL_TEMPLATE_CONTENT = """=== BETREFF-OPTIONEN ===

1. Rückruf-Anfragen in Ihrer Praxis einfacher organisieren
2. Kurze Frage zu Terminanfragen in Ihrer Praxis
3. Weniger verpasste Anfragen für Ihre Praxis

---

=== VERSION A — Alex / PraxisMed Team ===
(Empfohlen für Erstkontakt)

Guten Tag,

mein Name ist Alex von PraxisMed.

PraxisMed ist eine einfache KI-Rezeption für Wiener Privatpraxen.
Es ersetzt kein bestehendes Praxissystem wie LATIDO oder Ihren Kalender.
Es hilft davor: Anrufe und Rückruf-Anfragen werden sauber aufgenommen
und für Ihr Praxisteam übersichtlich sortiert.

Die KI bestätigt keine Termine, stellt keine Diagnosen und gibt keine
medizinische Beratung. Ihr Team bleibt immer in Kontrolle.

Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?

Mit freundlichen Grüßen
Alex
PraxisMed

---

=== VERSION B — Transparente Identität (Ali "Alex" Abdeltawab) ===
(Verwenden, wenn volle Transparenz gewünscht oder wenn direkt nach vollem Namen gefragt wird)

Guten Tag,

mein Name ist Ali Abdeltawab — ich verwende für Praxiskontakte auch
den Namen Alex, falls das einfacher ist.

Ich baue PraxisMed, eine einfache KI-Rezeption für Wiener Privatpraxen.
PraxisMed ersetzt kein bestehendes Praxissystem wie LATIDO oder Ihren Kalender.
Es hilft davor: Anrufe und Rückruf-Anfragen werden sauber aufgenommen
und für Ihr Praxisteam übersichtlich sortiert.

Die KI bestätigt keine Termine, stellt keine Diagnosen und gibt keine
medizinische Beratung. Ihr Team bleibt immer in Kontrolle.

Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?

Mit freundlichen Grüßen
Ali Abdeltawab (Alex)
PraxisMed

---

=== KURZVERSION FÜR KONTAKTFORMULARE ===

Guten Tag, ich bin Alex von PraxisMed, einer einfachen KI-Rezeption
für Wiener Privatpraxen. PraxisMed ersetzt LATIDO oder bestehende Systeme nicht,
sondern organisiert Anrufe und Rückruf-Anfragen für Ihr Praxisteam.
Die KI bestätigt keine Termine und gibt keine medizinische Beratung.
Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?

---

=== LATIDO / BESTEHENDES SYSTEM HINWEIS ===

Falls Sie bereits LATIDO oder ein anderes Praxissystem nutzen —
PraxisMed ersetzt das nicht. Es hilft davor: Rückrufwünsche und
Telefonanfragen werden sauber aufgenommen, Ihr Team trägt bestätigte
Termine wie gewohnt ins System ein.

---

=== REGELN ===

- Kein Massen-Send (No mass emailing)
- Kein Auto-Send (No automated sending)
- Manual review before every send
- Keine Diagnosen-Behauptungen (No diagnosis claims)
- Keine medizinische Beratung (No medical advice claims)
- Keine automatische Terminbestätigung (No auto-confirm)
- Team bleibt immer in Kontrolle (Staff stays in control)
- Do not contact: sofort markieren und Stop
- No patient data. No PHI.
"""


def _build_email_templates_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    ws.cell(row=1, column=1, value="EMAIL TEMPLATES — MANUAL REVIEW ONLY — No mass-send. No auto-send.")
    ws.cell(row=1, column=1).fill = _fill(RED_LT)
    ws.cell(row=1, column=1).font = _font(bold=True, color="7B241C", size=12)
    ws.cell(row=1, column=1).alignment = _align(h="center")
    ws.row_dimensions[1].height = 24

    cell = ws.cell(row=2, column=1, value=EMAIL_TEMPLATE_CONTENT)
    cell.font = Font(size=11, name="Courier New")
    cell.alignment = _align(wrap=True, v="top")
    ws.row_dimensions[2].height = 800


# ---------------------------------------------------------------------------
# Build Stats sheet
# ---------------------------------------------------------------------------

def _build_stats_sheet(ws, n_calls: int, n_emails: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    ws.cell(row=1, column=1, value="OUTREACH STATS — Week of 11.07.2026")
    ws.cell(row=1, column=1).fill = _fill(NAVY)
    ws.cell(row=1, column=1).font = _font(bold=True, color=WHITE, size=13)
    ws.cell(row=1, column=1).alignment = _align(h="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    def _stat_row(row, label, formula_or_val, target=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True, size=11)
        lc.alignment = _align()
        vc = ws.cell(row=row, column=2, value=formula_or_val)
        vc.font = _font(size=11)
        vc.alignment = _align(h="center")
        if target is not None:
            tc = ws.cell(row=row, column=3, value=target)
            tc.font = _font(size=11, color="117A65")
            tc.alignment = _align(h="center")
        ws.row_dimensions[row].height = 20

    ws.cell(row=2, column=1, value="Metric").fill = _fill(TEAL)
    ws.cell(row=2, column=1).font = _font(bold=True, color=WHITE)
    ws.cell(row=2, column=2, value="Actual").fill = _fill(TEAL)
    ws.cell(row=2, column=2).font = _font(bold=True, color=WHITE)
    ws.cell(row=2, column=2).alignment = _align(h="center")
    ws.cell(row=2, column=3, value="Target").fill = _fill(TEAL)
    ws.cell(row=2, column=3).font = _font(bold=True, color=WHITE)
    ws.cell(row=2, column=3).alignment = _align(h="center")

    _stat_row(3, "Total Call Targets (Monday)",    n_calls,   50)
    _stat_row(4, "Total Email Candidates (Sunday)", n_emails,  50)
    _stat_row(5, "Calls Attempted",
              '=COUNTIF(\'Monday Calls\'!J3:J52,"<>Not contacted")', 50)
    _stat_row(6, "Calls – No Answer",
              '=COUNTIF(\'Monday Calls\'!K3:K52,"No answer")', "—")
    _stat_row(7, "Calls – Interested",
              '=COUNTIF(\'Monday Calls\'!J3:J52,"Interested")', "—")
    _stat_row(8, "Calls – Not Interested",
              '=COUNTIF(\'Monday Calls\'!J3:J52,"Not interested")', "—")
    _stat_row(9, "Calls – Do Not Contact",
              '=COUNTIF(\'Monday Calls\'!J3:J52,"Do not contact")', 0)
    _stat_row(10, "Demo Offered",
              '=COUNTIF(\'Monday Calls\'!N3:N52,"Yes")', "—")
    _stat_row(11, "Demo Booked",
              '=COUNTIF(\'Monday Calls\'!O3:O52,"Yes")', ">0")
    _stat_row(12, "Emails Sent (Sunday)",
              '=COUNTIF(\'Sunday Emails\'!H3:H52,"Sent – email")+COUNTIF(\'Sunday Emails\'!H3:H52,"Sent – contact form")', 50)
    _stat_row(13, "Email Replies – Interested",
              '=COUNTIF(\'Sunday Emails\'!I3:I52,"Replied – interested")', "—")


# ---------------------------------------------------------------------------
# Build START HERE sheet
# ---------------------------------------------------------------------------

def _build_start_here_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 60

    lines = [
        ("PraxisMed — Outreach Command Center", NAVY, 16, True),
        ("Week of 11.07.2026 — 50 Calls · 50 Emails", TEAL, 13, False),
        ("", WHITE, 11, False),
        ("SATURDAY 11.07 — PREPARE ONLY", ORANGE, 12, True),
        ("• Open 'Monday Calls' tab — review your 50 call targets", WHITE, 11, False),
        ("• Open 'Sunday Emails' tab — review your 50 email candidates", WHITE, 11, False),
        ("• Print 4_CALL_SCRIPT_PRINTABLE.md if needed", WHITE, 11, False),
        ("• No calls today. No emails today. Clinics are closed.", GRAY_LT, 11, False),
        ("", WHITE, 11, False),
        ("SUNDAY 12.07 — SEND EMAILS MANUALLY", TEAL, 12, True),
        ("• Open 'Sunday Emails' tab", WHITE, 11, False),
        ("• Pick leads with a Praxisplan Profile URL", WHITE, 11, False),
        ("• Visit the URL → find contact form or email", WHITE, 11, False),
        ("• Copy the email draft → paste → adapt → send ONE AT A TIME", WHITE, 11, False),
        ("• Update 'Manual Send Status' after each send", WHITE, 11, False),
        ("• NO mass-send. NO auto-send.", RED_LT, 11, True),
        ("", WHITE, 11, False),
        ("MONDAY 13.07 — START CALLING (from 08:00)", "D5F5E3", 12, True),
        ("• Open 'Monday Calls' tab", WHITE, 11, False),
        ("• Follow Call Order 1 → 50", WHITE, 11, False),
        ("• Update Outreach Status and Call Result after every call", WHITE, 11, False),
        ("• Target: 50 call attempts", WHITE, 11, False),
        ("• Check 'Stats' tab for live progress", WHITE, 11, False),
        ("", WHITE, 11, False),
        ("SHEETS IN THIS WORKBOOK", NAVY, 12, True),
        ("Monday Calls     — 50 call targets with scripts and status tracking", GRAY_LT, 11, False),
        ("Sunday Emails    — 50 email candidates with drafts (Alex / PraxisMed Team)", GRAY_LT, 11, False),
        ("Sender Identity  — Who is Alex? How to set up sender name in Gmail", GRAY_LT, 11, False),
        ("Weekly Plan      — Sat/Sun/Mon/Tue/Wed/Thu schedule", GRAY_LT, 11, False),
        ("Follow Ups       — Manual follow-up tracker", GRAY_LT, 11, False),
        ("Master Leads     — Full 1,245-lead database", GRAY_LT, 11, False),
        ("Call Script      — German phone scripts for all situations", GRAY_LT, 11, False),
        ("Email Templates  — Both email versions + rules", GRAY_LT, 11, False),
        ("Stats            — Live counts (target: 50)", GRAY_LT, 11, False),
    ]

    for row, (text, color, size, bold) in enumerate(lines, 1):
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = _fill(color)
        cell.font = Font(bold=bold, size=size, name="Calibri",
                         color="FFFFFF" if color not in (WHITE, GRAY_LT, RED_LT, "D5F5E3") else "1A1A1A")
        cell.alignment = _align(h="left" if row > 2 else "center", wrap=True)
        ws.row_dimensions[row].height = 20 if size <= 11 else 28


# ---------------------------------------------------------------------------
# Build standalone Monday Calls xlsx
# ---------------------------------------------------------------------------

def _build_standalone_monday(call_leads: list[dict[str, str]], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monday Calls"
    _build_monday_calls_sheet(ws, call_leads)
    wb.save(path)


# ---------------------------------------------------------------------------
# Build standalone Sunday Emails xlsx
# ---------------------------------------------------------------------------

def _build_standalone_sunday(email_leads: list[dict[str, str]], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sunday Emails"
    _build_sunday_emails_sheet(ws, email_leads)
    wb.save(path)


# ---------------------------------------------------------------------------
# Build full command center
# ---------------------------------------------------------------------------

def build_command_center(call_leads: list[dict[str, str]], email_leads: list[dict[str, str]],
                         all_leads: list[dict[str, str]], path: str) -> None:
    wb = openpyxl.Workbook()
    sheets = [
        "START HERE", "Monday Calls", "Sunday Emails", "Sender Identity",
        "Weekly Plan", "Follow Ups", "Master Leads", "Call Script", "Email Templates", "Stats",
    ]

    # Create sheets in order
    wb.active.title = sheets[0]
    for name in sheets[1:]:
        wb.create_sheet(name)

    _build_start_here_sheet(wb["START HERE"])
    _build_monday_calls_sheet(wb["Monday Calls"], call_leads)
    _build_sunday_emails_sheet(wb["Sunday Emails"], email_leads)
    _build_sender_identity_sheet(wb["Sender Identity"])
    _build_weekly_plan_sheet(wb["Weekly Plan"])
    _build_followup_sheet(wb["Follow Ups"])
    _build_master_leads_sheet(wb["Master Leads"], all_leads)
    _build_call_script_sheet(wb["Call Script"])
    _build_email_templates_sheet(wb["Email Templates"])
    _build_stats_sheet(wb["Stats"], len(call_leads), len(email_leads))

    wb.save(path)


# ---------------------------------------------------------------------------
# Write markdown files
# ---------------------------------------------------------------------------

README_MD = """\
# PraxisMed Outreach — Read Me First

---

## Saturday 11.07.2026

**Prepare only.**

- Review `OUTREACH_COMMAND_CENTER.xlsx` → Monday Calls sheet
- Review `OUTREACH_COMMAND_CENTER.xlsx` → Sunday Emails sheet
- Nothing to send today. No calls today. Clinics are closed.

---

## Sunday 12.07.2026

**Prepare and manually send selected emails and contact forms.**

- Open `2_SUNDAY_EMAILS.xlsx`
- Pick the top 10 leads that have a website or contact form
- Copy the email draft for each one individually
- Paste into your email app or the clinic's contact form
- Send manually — one at a time
- Update the **Manual Send Status** column after each send
- **Do not mass-send. Do not use any automation.**

---

## Monday 13.07.2026

**Start phone calls. Clinics open from 08:00.**

- Open `1_MONDAY_CALLS.xlsx`
- Call from top to bottom (Call Order 1, 2, 3…)
- Use `4_CALL_SCRIPT_PRINTABLE.md` if you need the script on paper
- After every call, update the Status and Call Result columns
- Update `3_MASTER_TRACKER.xlsx` with full details after each call

---

## Files in this folder

| File | When to use |
|---|---|
| `OUTREACH_COMMAND_CENTER.xlsx` | Main workbook — all sheets in one place |
| `1_MONDAY_CALLS.xlsx` | Standalone call list for Monday |
| `2_SUNDAY_EMAILS.xlsx` | Standalone email drafts for Sunday |
| `3_MASTER_TRACKER.xlsx` | Full 1,245-lead database |
| `4_CALL_SCRIPT_PRINTABLE.md` | Print-ready phone scripts |
| `5_EMAIL_TEMPLATES_PRINTABLE.md` | Print-ready email templates |

---

## Do not use raw CSV files unless needed

The files in `docs/sales/outreach/daily_plans/` and the other outreach subfolders
are technical builder outputs. They are not meant for daily use.
Everything you need is in this folder.

---

## Rules (short version)

- Public practice contacts only
- Manual calls only — no auto-dial
- Manual emails only — no bulk send
- Mark **Do not contact** immediately if requested
- No patient data here, ever
"""

CALL_SCRIPT_MD = """\
# Call Script — Printable Version

**PraxisMed — Ali Abdeltawab**
Print this and keep it next to the phone.

---

## Opening / Haupteröffnung

Guten Tag, mein Name ist Ali Abdeltawab.
Ich baue PraxisMed, eine einfache KI-Rezeption für Wiener Privatpraxen.

Es ersetzt nicht Ihre Rezeption und auch nicht Ihr bestehendes Praxissystem.
PraxisMed hilft, Terminanfragen und Rückrufe sauber zu organisieren.

Darf ich Ihnen in 5 Minuten zeigen, wie aus einem Anruf eine Rückruf-Anfrage im Dashboard wird?

---

## Wenn: Sie nutzen bereits LATIDO / ein bestehendes System

Perfekt, dann müssen wir LATIDO nicht ersetzen.
PraxisMed organisiert nur den Telefon- und Rückruf-Teil davor.
Ihr Team trägt bestätigte Termine wie gewohnt in LATIDO oder Ihrem Praxissystem ein.

---

## Wenn: Schicken Sie uns eine E-Mail

Gerne. An welche E-Mail-Adresse darf ich die kurze Info schicken?
Ich halte es kurz — es geht um weniger verpasste Terminanfragen
und eine klare Rückruf-Liste für Ihr Praxisteam.

*(E-Mail-Adresse notieren → Manuell senden. Kein Auto-Send.)*

---

## Wenn: Kein Interesse

Verstehe, danke für Ihre Zeit.
Ich notiere, dass wir Sie nicht weiter kontaktieren.

*(→ Status: Do not contact. Sofort im Tracker eintragen.)*

---

## Wenn: Was kostet das?

Für den Pilot starten wir einfach:
30 Tage Testphase, Setup ab 390 Euro.
Danach je nach Umfang etwa 290 bis 490 Euro monatlich für kleine Privatpraxen.

---

## Wenn: Kalender / Buchung?

Im Pilot starten wir zuerst mit der Rückruf- und Anfrage-Übersicht.
Der nächste Schritt ist ein einfacher Kalender-Workflow:
Anfrage prüfen, Termin vorschlagen und vom Praxisteam bestätigen.

---

## Wenn: Gibt KI medizinische Beratung / Diagnosen?

Nein. PraxisMed gibt keine medizinische Beratung,
stellt keine Diagnosen und bestätigt keine Termine automatisch.
Ihr Team bleibt immer in Kontrolle.

---

## Wenn: Wer ist Alex?

Mein Name ist Ali Abdeltawab — ich verwende für Praxiskontakte auch den Namen Alex.
Falls Sie eine E-Mail von Alex, PraxisMed erhalten haben, war das ich.

---

## Nach jedem Anruf sofort eintragen

- Outreach Status
- Call Attempt 1 Date + Result
- Contact Person (Name Rezeption)
- Follow-up Date (falls vereinbart)
- Demo Offered / Demo Booked
- Objection
- Notes

---

*No auto-calling. No auto-email. No patient data. No PHI.*
"""

EMAIL_TEMPLATES_MD = """\
# Email Templates — Printable Version

**Manual review only. Copy one email at a time. Do not mass-send.**

Every email must be personally reviewed and adapted before sending.
Use your email client or the clinic's contact form.
No automated sending. No bulk sending.

---

## Subject line options

1. Rückruf-Anfragen in Ihrer Praxis einfacher organisieren
2. Kurze Frage zu Terminanfragen in Ihrer Praxis
3. Weniger verpasste Anfragen für Ihre Praxis

---

## Version A — Alex / PraxisMed Team (Empfohlen für Erstkontakt)

```
Guten Tag,

mein Name ist Alex von PraxisMed.

PraxisMed ist eine einfache KI-Rezeption für Wiener Privatpraxen.
Es ersetzt kein bestehendes Praxissystem wie LATIDO oder Ihren Kalender.
Es hilft davor: Anrufe und Rückruf-Anfragen werden sauber aufgenommen
und für Ihr Praxisteam übersichtlich sortiert.

Die KI bestätigt keine Termine, stellt keine Diagnosen und gibt keine
medizinische Beratung. Ihr Team bleibt immer in Kontrolle.

Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?

Mit freundlichen Grüßen
Alex
PraxisMed
```

---

## Version B — Transparente Identität (Ali "Alex" Abdeltawab)

Use this version when full transparency is preferred, or when asked for your full name.

```
Guten Tag,

mein Name ist Ali Abdeltawab — ich verwende für Praxiskontakte auch
den Namen Alex, falls das einfacher ist.

Ich baue PraxisMed, eine einfache KI-Rezeption für Wiener Privatpraxen.
PraxisMed ersetzt kein bestehendes Praxissystem wie LATIDO oder Ihren Kalender.
Es hilft davor: Anrufe und Rückruf-Anfragen werden sauber aufgenommen
und für Ihr Praxisteam übersichtlich sortiert.

Die KI bestätigt keine Termine, stellt keine Diagnosen und gibt keine
medizinische Beratung. Ihr Team bleibt immer in Kontrolle.

Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?

Mit freundlichen Grüßen
Ali Abdeltawab (Alex)
PraxisMed
```

---

## Short version (for contact forms)

```
Guten Tag, ich bin Alex von PraxisMed, einer einfachen KI-Rezeption
für Wiener Privatpraxen. PraxisMed ersetzt LATIDO oder bestehende Systeme nicht,
sondern organisiert Anrufe und Rückruf-Anfragen für Ihr Praxisteam.
Die KI bestätigt keine Termine und gibt keine medizinische Beratung.
Darf ich Ihnen eine kurze 5-Minuten-Demo zeigen?
```

---

## LATIDO / existing system note (add when relevant)

```
Falls Sie bereits LATIDO oder ein anderes Praxissystem nutzen —
PraxisMed ersetzt das nicht. Es hilft davor: Rückrufwünsche und
Telefonanfragen werden sauber aufgenommen, Ihr Team trägt bestätigte
Termine wie gewohnt ins System ein.
```

---

## Who is Alex? (if asked)

Alex is the outreach name Ali Abdeltawab uses for easier communication.
If a clinic asks, Ali explains:
"Mein vollständiger Name ist Ali Abdeltawab — ich verwende für Praxiskontakte auch Alex."
This is not a fake identity. Legal name is never hidden.

---

## Rules

- No mass emailing
- No automated sending
- No diagnosis claims
- No medical advice claims
- No appointment auto-confirmation claims
- Staff stays in control — always state this
- Respect opt-out immediately: mark Do not contact and stop
- No patient data, no PHI
"""


def _write_markdown_files(out_dir: str) -> None:
    files = {
        "0_READ_ME_FIRST.md": README_MD,
        "4_CALL_SCRIPT_PRINTABLE.md": CALL_SCRIPT_MD,
        "5_EMAIL_TEMPLATES_PRINTABLE.md": EMAIL_TEMPLATES_MD,
    }
    for name, content in files.items():
        with open(os.path.join(out_dir, name), "w", encoding="utf-8") as f:
            f.write(content)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Loading leads...")
    all_leads = _load_leads()
    print(f"  Total leads: {len(all_leads)}")

    print("Selecting 50 call targets...")
    call_leads = _select_call_leads(all_leads)
    print(f"  Selected: {len(call_leads)} call leads")

    call_ids = {l["Lead ID"] for l in call_leads}
    print("Selecting 50 email candidates...")
    email_leads = _select_email_leads(all_leads, call_ids)
    print(f"  Selected: {len(email_leads)} email leads")

    os.makedirs(_OUT_DIR, exist_ok=True)

    print("Building OUTREACH_COMMAND_CENTER.xlsx (10 sheets)...")
    cmd_path = os.path.join(_OUT_DIR, "OUTREACH_COMMAND_CENTER.xlsx")
    build_command_center(call_leads, email_leads, all_leads, cmd_path)
    size = os.path.getsize(cmd_path)
    print(f"  Written: {cmd_path} ({size:,} bytes)")

    print("Building 1_MONDAY_CALLS.xlsx...")
    mon_path = os.path.join(_OUT_DIR, "1_MONDAY_CALLS.xlsx")
    _build_standalone_monday(call_leads, mon_path)
    print(f"  Written: {mon_path} ({os.path.getsize(mon_path):,} bytes)")

    print("Building 2_SUNDAY_EMAILS.xlsx...")
    sun_path = os.path.join(_OUT_DIR, "2_SUNDAY_EMAILS.xlsx")
    _build_standalone_sunday(email_leads, sun_path)
    print(f"  Written: {sun_path} ({os.path.getsize(sun_path):,} bytes)")

    print("Writing markdown files...")
    _write_markdown_files(_OUT_DIR)
    print("  Written: 0_READ_ME_FIRST.md, 4_CALL_SCRIPT_PRINTABLE.md, 5_EMAIL_TEMPLATES_PRINTABLE.md")

    print("\nDone. Specialty breakdown for Monday Calls:")
    from collections import Counter
    sp_counts = Counter(l.get("Specialty Key") for l in call_leads)
    for sp, cnt in sorted(sp_counts.items()):
        print(f"  {sp}: {cnt}")


if __name__ == "__main__":
    main()
