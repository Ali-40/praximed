#!/usr/bin/env python3
"""
build_praxisplan_multi_specialty_leads.py

Multi-specialty Praxisplan.at lead database builder.

Reads a JSON config of specialties, scrapes each one that has a source_url,
and produces per-specialty XLSX/CSV files plus one combined master workbook.

RESPONSIBLE USE:
- Collects only publicly visible practice contact details from Praxisplan.at.
- No private data, no patient data, no PHI.
- No automated mass-emailing or auto-calling.
- Rate-limited: 1.5–2.5s between requests.
- Manual outreach tracker only.

Usage:
  # Build all specialties (scrape live where URL is configured):
  python scripts/sales/build_praxisplan_multi_specialty_leads.py --all

  # Build one specialty only:
  python scripts/sales/build_praxisplan_multi_specialty_leads.py --specialty dermatology

  # Generate templates only (no network):
  python scripts/sales/build_praxisplan_multi_specialty_leads.py --templates-only

  # Use custom config:
  python scripts/sales/build_praxisplan_multi_specialty_leads.py --all \\
      --config docs/sales/outreach/praxisplan_specialty_sources.json
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import date
from typing import Optional

try:
    import requests
    from bs4 import BeautifulSoup
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_URL = "https://www.praxisplan.at"
DEFAULT_CONFIG_PATH = "docs/sales/outreach/praxisplan_specialty_sources.json"
DEFAULT_OUTPUT_DIR = "docs/sales/outreach"
MASTER_XLSX_NAME = "praxisplan_all_high_potential_leads.xlsx"
MASTER_CSV_NAME = "praxisplan_all_high_potential_leads.csv"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}
REQUEST_DELAY_MIN = 1.5
REQUEST_DELAY_MAX = 2.5

# ---------------------------------------------------------------------------
# Column definition (extended with specialty metadata columns)
# ---------------------------------------------------------------------------

COLUMNS = [
    "Lead ID",
    "Specialty Tier",
    "Specialty Key",
    "Specialty Label DE",
    "Specialty Label EN",
    "Doctor Name",
    "Title",
    "Practice Name",
    "Specialty",
    "Sub-specialty / Notes from Listing",
    "Address",
    "Postal Code",
    "City",
    "District",
    "Phone",
    "Email",
    "Website",
    "Praxisplan Profile URL",
    "Source URL",
    "Source",
    "Existing System Mentioned",
    "Likely LATIDO / Online Booking",
    "Priority Score",
    "Priority Reason",
    "Outreach Status",
    "Call Attempt 1 Date",
    "Call Attempt 1 Result",
    "Call Attempt 2 Date",
    "Call Attempt 2 Result",
    "Email Sent Date",
    "Email Result",
    "Walk-in Date",
    "Walk-in Result",
    "Contact Person",
    "Best Time to Call",
    "Follow-up Date",
    "Demo Offered",
    "Demo Booked",
    "Demo Date",
    "Pilot Interest",
    "Objection",
    "Next Action",
    "Notes",
    "Last Updated",
]

# ---------------------------------------------------------------------------
# Dropdown / validation values
# ---------------------------------------------------------------------------

OUTREACH_STATUS_VALUES = [
    "Not contacted", "Called", "No answer", "Reception reached",
    "Asked to send email", "Email sent", "Follow-up needed",
    "Demo offered", "Demo booked", "Interested", "Not interested",
    "Wrong target", "Do not contact",
]
CALL_RESULT_VALUES = [
    "No answer", "Busy", "Reception reached", "Doctor unavailable",
    "Asked to send email", "Interested", "Not interested", "Call later", "Demo booked",
]
EMAIL_RESULT_VALUES = [
    "Not sent", "Sent", "Replied interested", "Replied not interested", "No reply", "Bounced",
]
WALK_IN_RESULT_VALUES = [
    "Not visited", "Left flyer", "Reception conversation", "Doctor conversation",
    "Asked to email", "Demo booked", "Not interested",
]
DEMO_YN_VALUES = ["Yes", "No"]
PILOT_INTEREST_VALUES = ["Unknown", "Low", "Medium", "High", "Not interested"]
PRIORITY_SCORE_VALUES = ["1", "2", "3", "4", "5"]

VIENNA_DISTRICTS: dict[str, str] = {
    "1010": "1. Bezirk – Innere Stadt", "1020": "2. Bezirk – Leopoldstadt",
    "1030": "3. Bezirk – Landstraße", "1040": "4. Bezirk – Wieden",
    "1050": "5. Bezirk – Margareten", "1060": "6. Bezirk – Mariahilf",
    "1070": "7. Bezirk – Neubau", "1080": "8. Bezirk – Josefstadt",
    "1090": "9. Bezirk – Alsergrund", "1100": "10. Bezirk – Favoriten",
    "1110": "11. Bezirk – Simmering", "1120": "12. Bezirk – Meidling",
    "1130": "13. Bezirk – Hietzing", "1140": "14. Bezirk – Penzing",
    "1150": "15. Bezirk – Rudolfsheim-Fünfhaus", "1160": "16. Bezirk – Ottakring",
    "1170": "17. Bezirk – Hernals", "1180": "18. Bezirk – Währing",
    "1190": "19. Bezirk – Döbling", "1200": "20. Bezirk – Brigittenau",
    "1210": "21. Bezirk – Floridsdorf", "1220": "22. Bezirk – Donaustadt",
    "1230": "23. Bezirk – Liesing",
}

# Fake template rows — one per specialty_key for templates-only mode
FAKE_TEMPLATE_ROWS: dict[str, dict] = {
    "child_adolescent_psychiatry": {
        "Doctor Name": "Demo Kinderpsychiatrie Wien",
        "Title": "Dr.", "Specialty": "Kinder- u. Jugendpsychiatrie u. Psychotherapeutische Medizin",
        "Address": "Musterstraße 1, 1010 Wien", "Postal Code": "1010",
        "City": "Wien", "District": "1. Bezirk – Innere Stadt", "Phone": "+43 000 000000",
    },
    "dermatology": {
        "Doctor Name": "Demo Dermatologie Wien",
        "Title": "Dr.", "Specialty": "Haut- u. Geschlechtskrankheiten",
        "Address": "Beispielgasse 5, 1080 Wien", "Postal Code": "1080",
        "City": "Wien", "District": "8. Bezirk – Josefstadt", "Phone": "+43 000 000000",
    },
    "gynecology": {
        "Doctor Name": "Demo Frauenarzt Praxis",
        "Title": "Dr.", "Specialty": "Frauenheilkunde u. Geburtshilfe",
        "Address": "Teststraße 12, 1090 Wien", "Postal Code": "1090",
        "City": "Wien", "District": "9. Bezirk – Alsergrund", "Phone": "+43 000 000000",
    },
    "orthopedics": {
        "Doctor Name": "Demo Orthopädie Zentrum",
        "Title": "Dr.", "Specialty": "Orthopädie und Traumatologie",
        "Address": "Hauptplatz 3, 1060 Wien", "Postal Code": "1060",
        "City": "Wien", "District": "6. Bezirk – Mariahilf", "Phone": "+43 000 000000",
    },
    "internal_medicine": {
        "Doctor Name": "Demo Innere Medizin Praxis",
        "Title": "Univ.Prof. Dr.", "Specialty": "Innere Medizin",
        "Address": "Ringstraße 7, 1010 Wien", "Postal Code": "1010",
        "City": "Wien", "District": "1. Bezirk – Innere Stadt", "Phone": "+43 000 000000",
    },
    "ent": {
        "Doctor Name": "Demo HNO Ordination",
        "Title": "Dr.", "Specialty": "Hals-, Nasen- u. Ohrenheilkunde",
        "Address": "Kaiserstraße 22, 1070 Wien", "Postal Code": "1070",
        "City": "Wien", "District": "7. Bezirk – Neubau", "Phone": "+43 000 000000",
    },
    "urology": {
        "Doctor Name": "Demo Urologie Praxis",
        "Title": "Dr.", "Specialty": "Urologie",
        "Address": "Landstraße 44, 1030 Wien", "Postal Code": "1030",
        "City": "Wien", "District": "3. Bezirk – Landstraße", "Phone": "+43 000 000000",
    },
    "neurology": {
        "Doctor Name": "Demo Neurologie Praxis",
        "Title": "Dr.", "Specialty": "Neurologie",
        "Address": "Währinger Str. 12, 1090 Wien", "Postal Code": "1090",
        "City": "Wien", "District": "9. Bezirk – Alsergrund", "Phone": "+43 000 000000",
    },
    "ophthalmology": {
        "Doctor Name": "Demo Augenarzt Praxis",
        "Title": "Dr.", "Specialty": "Augenheilkunde u. Optometrie",
        "Address": "Tuchlauben 10, 1010 Wien", "Postal Code": "1010",
        "City": "Wien", "District": "1. Bezirk – Innere Stadt", "Phone": "+43 000 000000",
    },
    "pediatrics": {
        "Doctor Name": "Demo Kinderarzt Ordination",
        "Title": "Dr.", "Specialty": "Kinder- u. Jugendheilkunde",
        "Address": "Favoriten Str. 30, 1100 Wien", "Postal Code": "1100",
        "City": "Wien", "District": "10. Bezirk – Favoriten", "Phone": "+43 000 000000",
    },
    "private_dental": {
        "Doctor Name": "Demo Privat Zahnarzt",
        "Title": "Dr.", "Specialty": "Zahnmedizin (privat)",
        "Address": "Graben 1, 1010 Wien", "Postal Code": "1010",
        "City": "Wien", "District": "1. Bezirk – Innere Stadt", "Phone": "+43 000 000000",
    },
    "aesthetic_medicine": {
        "Doctor Name": "Demo Ästhetische Medizin",
        "Title": "Dr.", "Specialty": "Plastische, Rekonstruktive u. Ästhetische Chirurgie",
        "Address": "Kärntner Ring 5, 1010 Wien", "Postal Code": "1010",
        "City": "Wien", "District": "1. Bezirk – Innere Stadt", "Phone": "+43 000 000000",
    },
    "plastic_surgery": {
        "Doctor Name": "Demo Plastische Chirurgie",
        "Title": "Dr.", "Specialty": "Plastische u. Rekonstruktive Chirurgie",
        "Address": "Operngasse 3, 1040 Wien", "Postal Code": "1040",
        "City": "Wien", "District": "4. Bezirk – Wieden", "Phone": "+43 000 000000",
    },
    "adult_psychiatry": {
        "Doctor Name": "Demo Psychiatrie Praxis",
        "Title": "Dr.", "Specialty": "Psychiatrie u. Psychotherapeutische Medizin",
        "Address": "Alserstraße 18, 1090 Wien", "Postal Code": "1090",
        "City": "Wien", "District": "9. Bezirk – Alsergrund", "Phone": "+43 000 000000",
    },
    "private_group_practices": {
        "Doctor Name": "Demo Gruppenpraxis Wien",
        "Title": "", "Specialty": "Gruppenpraxis / Mehrere Fachgebiete",
        "Address": "Mariahilfer Str. 100, 1060 Wien", "Postal Code": "1060",
        "City": "Wien", "District": "6. Bezirk – Mariahilf", "Phone": "+43 000 000000",
    },
}

# ---------------------------------------------------------------------------
# Parsing helpers (standalone — does not import from single-specialty script)
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
POSTAL_RE = re.compile(r"\b(\d{4})\b")

TITLE_PREFIXES = (
    "Univ.Prof. Dr.", "Univ.Doz. Dr.", "Prof. Dr.", "Priv.Doz. Dr.",
    "Prim. Dr.", "OA Dr.", "Univ.Prof.", "Prof.", "Prim.", "Dr.",
    "Mag.", "DI", "MSc", "Bakk.", "BSc",
)


def _extract_title(name: str) -> tuple[str, str]:
    for prefix in TITLE_PREFIXES:
        if name.startswith(prefix + " "):
            return prefix, name[len(prefix):].strip()
    return "", name


def _parse_address(raw: str) -> tuple[str, str, str, str]:
    postal_match = POSTAL_RE.search(raw)
    postal = postal_match.group(1) if postal_match else ""
    city = ""
    district = VIENNA_DISTRICTS.get(postal, "")
    if postal:
        after = raw[postal_match.end():].strip(" ,")
        city = after.split(",")[0].strip() if after else ""
        street = raw[:postal_match.start()].strip(" ,")
    else:
        street = raw
    return street, postal, city, district


def _clean_specialty(raw: str) -> str:
    for prefix in ("Ordination: ", "Angestellt: ", "Angestellte*r: "):
        if raw.startswith(prefix):
            return raw[len(prefix):]
    return raw


def _score_priority(rec: dict, tier: int) -> tuple[int, str]:
    score = 0
    reasons: list[str] = []
    if rec.get("Phone"):
        score += 1
        reasons.append("public phone")
    if rec.get("Email") or rec.get("Website"):
        score += 1
        reasons.append("email/website available" if rec.get("Email") else "website available")
    spec = (rec.get("Specialty", "") + rec.get("Doctor Name", "")).lower()
    if "wahl" in spec or "privat" in spec:
        score += 1
        reasons.append("Wahlarzt/private")
    if rec.get("City", "").lower() in ("wien", "vienna"):
        score += 1
        reasons.append("Vienna")
    if tier == 1:
        score += 1
        reasons.append("Tier 1 high-priority specialty")
    elif tier == 2 and rec.get("Phone"):
        score += 1
        reasons.append("Tier 2 phone-available")
    score = min(score, 5)
    reason_str = ", ".join(reasons).capitalize() if reasons else "Needs manual review"
    return score, reason_str


# ---------------------------------------------------------------------------
# Network
# ---------------------------------------------------------------------------

def _fetch(url: str, session: "requests.Session") -> Optional["BeautifulSoup"]:
    try:
        r = session.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        return BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        print(f"  [WARN] Could not fetch {url}: {exc}", file=sys.stderr)
        return None


def _parse_listing_page(soup: "BeautifulSoup", source_url: str) -> list[dict]:
    records = []
    for block in soup.select("div.search-result"):
        name_el = block.select_one("p.label.mb-0")
        link_el = block.select_one("a[href]")
        if not name_el or not link_el:
            continue

        full_name = name_el.get_text(strip=True)
        title, _ = _extract_title(full_name)
        profile_href = link_el["href"]
        profile_url = (BASE_URL + profile_href) if profile_href.startswith("/") else profile_href

        link_text = link_el.get_text(separator="\n", strip=True)
        lines = [ln.strip() for ln in link_text.split("\n") if ln.strip()]

        specialty_raw = lines[0] if lines else ""
        specialty = _clean_specialty(specialty_raw)

        address_raw, phone = "", ""
        for line in lines[1:]:
            if re.match(r"^\+43", line) or re.match(r"^0[1-9]", line):
                phone = line
            elif re.search(r"\d{4}", line):
                address_raw = line
            elif not address_raw and len(line) > 5:
                address_raw = line

        address, postal, city, district = _parse_address(address_raw)
        records.append({
            "Doctor Name": full_name, "Title": title, "Practice Name": "",
            "Specialty": specialty, "Sub-specialty / Notes from Listing": "",
            "Address": address, "Postal Code": postal, "City": city, "District": district,
            "Phone": phone, "Email": "", "Website": "",
            "Praxisplan Profile URL": profile_url, "Source URL": source_url,
        })
    return records


def _scrape_specialty(source_url: str) -> list[dict]:
    if not _HAS_REQUESTS:
        print("  [WARN] requests/bs4 not available — skipping live scrape", file=sys.stderr)
        return []

    session = requests.Session()
    all_records: list[dict] = []
    seen: set[tuple] = set()
    page = 1

    while True:
        paged_url = source_url + (f"&page={page}" if page > 1 else "")
        print(f"    Fetching page {page}...")
        soup = _fetch(paged_url, session)
        if not soup:
            break

        records = _parse_listing_page(soup, source_url)
        if not records:
            break

        for rec in records:
            key = (rec["Doctor Name"].lower(), rec["Phone"], rec.get("Praxisplan Profile URL", ""))
            if key in seen:
                continue
            seen.add(key)
            all_records.append(rec)

        print(f"    → page {page}: {len(records)} entries ({len(all_records)} total)")

        if len(records) < 15:
            break

        page += 1
        time.sleep(REQUEST_DELAY_MIN + (REQUEST_DELAY_MAX - REQUEST_DELAY_MIN) * 0.5)

    return all_records


def _scrape_from_html(html_path: str, source_url: str) -> list[dict]:
    with open(html_path, encoding="utf-8", errors="replace") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")
    return _parse_listing_page(soup, source_url)


# ---------------------------------------------------------------------------
# Row building
# ---------------------------------------------------------------------------

def _blank_row(lead_id: str) -> dict:
    row: dict = {col: "" for col in COLUMNS}
    row["Lead ID"] = lead_id
    row["Source"] = "Praxisplan"
    row["Outreach Status"] = "Not contacted"
    row["Demo Offered"] = "No"
    row["Demo Booked"] = "No"
    row["Pilot Interest"] = "Unknown"
    row["Likely LATIDO / Online Booking"] = "Unknown"
    row["Last Updated"] = str(date.today())
    return row


def _build_rows(
    records: list[dict],
    spec: dict,
    lead_id_start: int,
) -> list[dict]:
    tier = spec.get("tier", 3)
    rows = []
    for i, rec in enumerate(records):
        row = _blank_row(f"LEAD-{lead_id_start + i:05d}")
        row["Specialty Tier"] = str(tier)
        row["Specialty Key"] = spec.get("specialty_key", "")
        row["Specialty Label DE"] = spec.get("specialty_label_de", "")
        row["Specialty Label EN"] = spec.get("specialty_label_en", "")
        row["Source URL"] = spec.get("source_url", "")
        for col in ("Doctor Name", "Title", "Practice Name", "Specialty",
                    "Sub-specialty / Notes from Listing", "Address", "Postal Code",
                    "City", "District", "Phone", "Email", "Website",
                    "Praxisplan Profile URL"):
            row[col] = rec.get(col, "")
        score, reason = _score_priority(rec, tier)
        row["Priority Score"] = str(score)
        row["Priority Reason"] = reason
        rows.append(row)
    return rows


def _build_template_row(spec: dict, lead_id: str) -> dict:
    key = spec.get("specialty_key", "")
    fake = FAKE_TEMPLATE_ROWS.get(key, {"Doctor Name": f"Demo {spec.get('specialty_label_de', key)}", "Phone": "+43 000 000000"})
    row = _blank_row(lead_id)
    row["Specialty Tier"] = str(spec.get("tier", 3))
    row["Specialty Key"] = key
    row["Specialty Label DE"] = spec.get("specialty_label_de", "")
    row["Specialty Label EN"] = spec.get("specialty_label_en", "")
    row["Source URL"] = spec.get("source_url", "")
    for k, v in fake.items():
        row[k] = v
    row["Priority Score"] = "3"
    row["Priority Reason"] = "Needs manual review"
    return row


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _col_letter(name: str) -> str:
    return get_column_letter(COLUMNS.index(name) + 1)


def _dv(formula: str, sqref: str) -> "DataValidation":
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = sqref
    return dv


def _quoted(values: list[str]) -> str:
    return '"' + ",".join(values) + '"'


COL_WIDTHS = {
    "Lead ID": 12, "Specialty Tier": 13, "Specialty Key": 22,
    "Specialty Label DE": 36, "Specialty Label EN": 32,
    "Doctor Name": 28, "Title": 14, "Practice Name": 24,
    "Specialty": 34, "Sub-specialty / Notes from Listing": 28,
    "Address": 32, "Postal Code": 12, "City": 12, "District": 28,
    "Phone": 20, "Email": 28, "Website": 30, "Praxisplan Profile URL": 38,
    "Source URL": 20, "Source": 12, "Existing System Mentioned": 22,
    "Likely LATIDO / Online Booking": 22, "Priority Score": 13,
    "Priority Reason": 30, "Outreach Status": 22,
    "Call Attempt 1 Date": 18, "Call Attempt 1 Result": 22,
    "Call Attempt 2 Date": 18, "Call Attempt 2 Result": 22,
    "Email Sent Date": 16, "Email Result": 20,
    "Walk-in Date": 16, "Walk-in Result": 22,
    "Contact Person": 22, "Best Time to Call": 18, "Follow-up Date": 16,
    "Demo Offered": 13, "Demo Booked": 13, "Demo Date": 14,
    "Pilot Interest": 16, "Objection": 28, "Next Action": 28,
    "Notes": 40, "Last Updated": 14,
}

HEADER_FILL = None  # lazily initialized in functions
HEADER_FONT = None


def _write_ws(ws: "openpyxl.worksheet.worksheet.Worksheet", rows: list[dict],
              sheet_title: str = "Leads", header_color: str = "1F4E79") -> None:
    hfill = PatternFill("solid", fgColor=header_color)
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(COLUMNS, 1):
            ws.cell(row=ri, column=ci, value=row.get(col, ""))
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[_col_letter(col)].width = width


def _add_dvs(ws: "openpyxl.worksheet.worksheet.Worksheet", max_row: int = 5000) -> None:
    def r(col_name: str) -> str:
        return f"{_col_letter(col_name)}2:{_col_letter(col_name)}{max_row}"

    dvs = [
        _dv(_quoted(OUTREACH_STATUS_VALUES), r("Outreach Status")),
        _dv(_quoted(CALL_RESULT_VALUES), r("Call Attempt 1 Result")),
        _dv(_quoted(CALL_RESULT_VALUES), r("Call Attempt 2 Result")),
        _dv(_quoted(EMAIL_RESULT_VALUES), r("Email Result")),
        _dv(_quoted(WALK_IN_RESULT_VALUES), r("Walk-in Result")),
        _dv(_quoted(DEMO_YN_VALUES), r("Demo Offered")),
        _dv(_quoted(DEMO_YN_VALUES), r("Demo Booked")),
        _dv(_quoted(PILOT_INTEREST_VALUES), r("Pilot Interest")),
        _dv(_quoted(PRIORITY_SCORE_VALUES), r("Priority Score")),
    ]
    for dv in dvs:
        ws.add_data_validation(dv)


def _write_summary_ws(ws: "openpyxl.worksheet.worksheet.Worksheet",
                      leads_ws_name: str, max_row: int = 5000) -> None:
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14

    def col(name: str) -> str:
        return f"'{leads_ws_name}'!{_col_letter(name)}2:{_col_letter(name)}{max_row}"

    rows = [
        ("Metric", "Count"),
        ("Total leads", f"=COUNTA('{leads_ws_name}'!{_col_letter('Lead ID')}2:{_col_letter('Lead ID')}{max_row})"),
        ("Phone available", f"=COUNTIF({col('Phone')},\"?*\")"),
        ("Email available", f"=COUNTIF({col('Email')},\"?*\")"),
        ("Website available", f"=COUNTIF({col('Website')},\"?*\")"),
        ("Not contacted", f"=COUNTIF({col('Outreach Status')},\"Not contacted\")"),
        ("Called", f"=COUNTIF({col('Outreach Status')},\"Called\")"),
        ("Demo booked", f"=COUNTIF({col('Demo Booked')},\"Yes\")"),
        ("Interested", f"=COUNTIF({col('Outreach Status')},\"Interested\")"),
        ("Follow-up needed", f"=COUNTIF({col('Outreach Status')},\"Follow-up needed\")"),
    ]
    for ri, (label, value) in enumerate(rows, 1):
        ws.cell(row=ri, column=1, value=label).font = Font(bold=(ri == 1))
        ws.cell(row=ri, column=2, value=value)


def save_specialty_xlsx(rows: list[dict], output_path: str, spec: dict) -> None:
    if not _HAS_OPENPYXL:
        print("  [WARN] openpyxl not available — skipping XLSX", file=sys.stderr)
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    _write_ws(ws, rows)
    _add_dvs(ws)
    ws2 = wb.create_sheet("Summary")
    _write_summary_ws(ws2, "Leads")
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"    Saved: {output_path}")


def save_specialty_csv(rows: list[dict], output_path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"    Saved: {output_path}")


TIER_COLORS = {"1": "1F4E79", "2": "2E5C8A", "3": "3A6B9E"}


def save_master_workbook(
    all_rows: list[dict],
    rows_by_spec: dict[str, list[dict]],
    output_path: str,
    specs: list[dict],
) -> None:
    if not _HAS_OPENPYXL:
        print("  [WARN] openpyxl not available — skipping master XLSX", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Leads"
    _write_ws(ws_all, all_rows, header_color="1F2D3D")
    _add_dvs(ws_all, max_row=10000)

    ws_summary = wb.create_sheet("Summary")
    _write_summary_ws(ws_summary, "All Leads", max_row=10000)

    for spec in specs:
        key = spec.get("specialty_key", "")
        label = (spec.get("specialty_label_en", key) or key)[:28]
        safe_title = re.sub(r"[\\/:*?\"<>|]", "", label)[:31]
        rows = rows_by_spec.get(key, [])
        color = TIER_COLORS.get(str(spec.get("tier", 3)), "1F4E79")
        ws = wb.create_sheet(safe_title)
        _write_ws(ws, rows, header_color=color)
        _add_dvs(ws, max_row=5000)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"  Master XLSX saved: {output_path}")


def save_master_csv(all_rows: list[dict], output_path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(all_rows)
    print(f"  Master CSV  saved: {output_path}")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def load_config(config_path: str) -> list[dict]:
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def run(
    specs: list[dict],
    output_dir: str,
    templates_only: bool = False,
    no_enrich: bool = True,
) -> None:
    all_rows: list[dict] = []
    rows_by_spec: dict[str, list[dict]] = {}
    global_lead_counter = 1
    global_dedup: set[tuple] = set()
    total_skipped_dups = 0

    for spec in specs:
        key = spec.get("specialty_key", "unknown")
        slug = spec.get("output_slug", key)
        source_url = spec.get("source_url", "")
        tier = spec.get("tier", 3)

        print(f"\n[{key}] (Tier {tier})")

        if templates_only or not source_url:
            if not source_url:
                print(f"  No source_url — generating template row.")
            else:
                print(f"  Templates-only mode.")
            row = _build_template_row(spec, f"TMPL-{key.upper()[:10]}-001")
            rows_by_spec[key] = [row]
        else:
            print(f"  Scraping: {source_url}")
            records = _scrape_specialty(source_url)
            print(f"  → {len(records)} records scraped")
            rows = _build_rows(records, spec, global_lead_counter)
            global_lead_counter += len(rows)
            rows_by_spec[key] = rows

        # Per-specialty output
        specialty_rows = rows_by_spec[key]
        xlsx_path = os.path.join(output_dir, f"{slug}_leads.xlsx")
        csv_path = os.path.join(output_dir, f"{slug}_leads.csv")
        save_specialty_xlsx(specialty_rows, xlsx_path, spec)
        save_specialty_csv(specialty_rows, csv_path)

        # Dedup before adding to master
        for row in specialty_rows:
            dedup_key = (
                row.get("Doctor Name", "").lower(),
                row.get("Phone", ""),
                row.get("Praxisplan Profile URL", ""),
            )
            if dedup_key in global_dedup:
                total_skipped_dups += 1
                continue
            global_dedup.add(dedup_key)
            all_rows.append(row)

        time.sleep(0.5)

    # Master workbook
    print(f"\nBuilding master workbook ({len(all_rows)} unique leads)...")
    master_xlsx = os.path.join(output_dir, MASTER_XLSX_NAME)
    master_csv = os.path.join(output_dir, MASTER_CSV_NAME)
    save_master_workbook(all_rows, rows_by_spec, master_xlsx, specs)
    save_master_csv(all_rows, master_csv)

    # Summary
    phone_count = sum(1 for r in all_rows if r.get("Phone"))
    email_count = sum(1 for r in all_rows if r.get("Email"))
    website_count = sum(1 for r in all_rows if r.get("Website"))
    print(f"\n{'=' * 60}")
    print(f"  Total unique leads:  {len(all_rows)}")
    print(f"  Duplicates removed:  {total_skipped_dups}")
    print(f"  With phone:          {phone_count}")
    print(f"  With email:          {email_count}")
    print(f"  With website:        {website_count}")
    print(f"  Specialties:         {len(specs)}")
    print(f"{'=' * 60}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Multi-specialty Praxisplan lead database builder"
    )
    parser.add_argument(
        "--all", action="store_true",
        help="Build all specialties from config"
    )
    parser.add_argument(
        "--specialty", metavar="KEY",
        help="Build one specialty by specialty_key (e.g. dermatology)"
    )
    parser.add_argument(
        "--templates-only", action="store_true",
        help="Generate template files only (no network)"
    )
    parser.add_argument(
        "--config", default=DEFAULT_CONFIG_PATH,
        help=f"Path to specialty config JSON (default: {DEFAULT_CONFIG_PATH})"
    )
    parser.add_argument(
        "--output-dir", default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory (default: {DEFAULT_OUTPUT_DIR})"
    )
    args = parser.parse_args()

    if not args.all and not args.specialty and not args.templates_only:
        parser.print_help()
        sys.exit(1)

    config_path = args.config
    if not os.path.isfile(config_path):
        print(f"[ERROR] Config file not found: {config_path}", file=sys.stderr)
        sys.exit(1)

    specs = load_config(config_path)

    if args.specialty:
        specs = [s for s in specs if s.get("specialty_key") == args.specialty]
        if not specs:
            print(f"[ERROR] specialty_key '{args.specialty}' not found in config", file=sys.stderr)
            sys.exit(1)

    run(
        specs=specs,
        output_dir=args.output_dir,
        templates_only=args.templates_only,
    )


if __name__ == "__main__":
    main()
