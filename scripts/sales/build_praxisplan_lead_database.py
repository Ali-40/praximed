#!/usr/bin/env python3
"""
build_praxisplan_lead_database.py

Builds a sales lead database from public Praxisplan.at listings.

RESPONSIBLE USE:
- Collects only publicly visible practice contact details.
- No private data, no patient data, no PHI.
- No automated mass-emailing or auto-calling.
- Rate-limited: 1.5–3s between requests.
- This is a manual outreach tracker only.

Usage:
  # Mode A — fetch live from URL
  python scripts/sales/build_praxisplan_lead_database.py \\
      --url "https://www.praxisplan.at/suche?specialization=71" \\
      --output docs/sales/outreach/praxisplan_child_psychiatry_leads.xlsx

  # Mode B — parse locally saved HTML
  python scripts/sales/build_praxisplan_lead_database.py \\
      --input-html docs/sales/outreach/praxisplan_saved_page.html \\
      --output docs/sales/outreach/praxisplan_child_psychiatry_leads.xlsx

  # Template only (no network access)
  python scripts/sales/build_praxisplan_lead_database.py --template-only
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from datetime import date
from pathlib import Path
from typing import Optional

try:
    import requests
    from bs4 import BeautifulSoup
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BASE_URL = "https://www.praxisplan.at"
DEFAULT_SOURCE_URL = (
    "https://www.praxisplan.at/suche?name=&specialization=71&zip=&gender="
    "&diploma=&specialty=&insurance=&vaccine=&locale="
)
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}
REQUEST_DELAY_MIN = 1.5
REQUEST_DELAY_MAX = 2.5

COLUMNS = [
    "Lead ID",
    "Doctor Name",
    "Title",
    "Practice Name",
    "Specialty",
    "Sub-specialty / Notes from Listing",
    "Address",
    "Postal Code",
    "City",
    "District",
    "Phone",
    "Email",
    "Website",
    "Praxisplan Profile URL",
    "Source URL",
    "Source",
    "Existing System Mentioned",
    "Likely LATIDO / Online Booking",
    "Priority Score",
    "Priority Reason",
    "Outreach Status",
    "Call Attempt 1 Date",
    "Call Attempt 1 Result",
    "Call Attempt 2 Date",
    "Call Attempt 2 Result",
    "Email Sent Date",
    "Email Result",
    "Walk-in Date",
    "Walk-in Result",
    "Contact Person",
    "Best Time to Call",
    "Follow-up Date",
    "Demo Offered",
    "Demo Booked",
    "Demo Date",
    "Pilot Interest",
    "Objection",
    "Next Action",
    "Notes",
    "Last Updated",
]

OUTREACH_STATUS_VALUES = [
    "Not contacted",
    "Called",
    "No answer",
    "Reception reached",
    "Asked to send email",
    "Email sent",
    "Follow-up needed",
    "Demo offered",
    "Demo booked",
    "Interested",
    "Not interested",
    "Wrong target",
    "Do not contact",
]

CALL_RESULT_VALUES = [
    "No answer",
    "Busy",
    "Reception reached",
    "Doctor unavailable",
    "Asked to send email",
    "Interested",
    "Not interested",
    "Call later",
    "Demo booked",
]

EMAIL_RESULT_VALUES = [
    "Not sent",
    "Sent",
    "Replied interested",
    "Replied not interested",
    "No reply",
    "Bounced",
]

WALK_IN_RESULT_VALUES = [
    "Not visited",
    "Left flyer",
    "Reception conversation",
    "Doctor conversation",
    "Asked to email",
    "Demo booked",
    "Not interested",
]

DEMO_YN_VALUES = ["Yes", "No"]

PILOT_INTEREST_VALUES = ["Unknown", "Low", "Medium", "High", "Not interested"]

PRIORITY_SCORE_VALUES = ["1", "2", "3", "4", "5"]

VIENNA_DISTRICTS: dict[str, str] = {
    "1010": "1. Bezirk – Innere Stadt",
    "1020": "2. Bezirk – Leopoldstadt",
    "1030": "3. Bezirk – Landstraße",
    "1040": "4. Bezirk – Wieden",
    "1050": "5. Bezirk – Margareten",
    "1060": "6. Bezirk – Mariahilf",
    "1070": "7. Bezirk – Neubau",
    "1080": "8. Bezirk – Josefstadt",
    "1090": "9. Bezirk – Alsergrund",
    "1100": "10. Bezirk – Favoriten",
    "1110": "11. Bezirk – Simmering",
    "1120": "12. Bezirk – Meidling",
    "1130": "13. Bezirk – Hietzing",
    "1140": "14. Bezirk – Penzing",
    "1150": "15. Bezirk – Rudolfsheim-Fünfhaus",
    "1160": "16. Bezirk – Ottakring",
    "1170": "17. Bezirk – Hernals",
    "1180": "18. Bezirk – Währing",
    "1190": "19. Bezirk – Döbling",
    "1200": "20. Bezirk – Brigittenau",
    "1210": "21. Bezirk – Floridsdorf",
    "1220": "22. Bezirk – Donaustadt",
    "1230": "23. Bezirk – Liesing",
}

FAKE_EXAMPLE_ROWS = [
    {
        "Doctor Name": "Demo Kinderpsychiatrie Wien",
        "Title": "Dr.",
        "Practice Name": "Demo Kinderpsychiatrie Wien",
        "Specialty": "Kinder- u. Jugendpsychiatrie",
        "Address": "Musterstraße 1, 1010 Wien",
        "Postal Code": "1010",
        "City": "Wien",
        "District": "1. Bezirk – Innere Stadt",
        "Phone": "+43 1 123 456 78",
        "Email": "",
        "Website": "",
        "Praxisplan Profile URL": "https://www.praxisplan.at/EXAMPLE",
        "Priority Score": "4",
        "Priority Reason": "Public phone, Vienna, high callback specialty",
    },
    {
        "Doctor Name": "Demo Psychotherapie Praxis",
        "Title": "Mag.",
        "Practice Name": "Demo Psychotherapie Praxis",
        "Specialty": "Psychotherapeutische Medizin",
        "Address": "Beispielgasse 5/3, 1060 Wien",
        "Postal Code": "1060",
        "City": "Wien",
        "District": "6. Bezirk – Mariahilf",
        "Phone": "+43 664 987 65 43",
        "Email": "praxis@example.at",
        "Website": "https://www.example-praxis.at",
        "Praxisplan Profile URL": "https://www.praxisplan.at/EXAMPLE",
        "Priority Score": "5",
        "Priority Reason": "Email available, strong outreach target",
    },
    {
        "Doctor Name": "Demo Privatpraxis 1060",
        "Title": "Dr.",
        "Practice Name": "Demo Privatpraxis 1060",
        "Specialty": "Kinder- u. Jugendpsychiatrie",
        "Address": "Testgasse 12, 1060 Wien",
        "Postal Code": "1060",
        "City": "Wien",
        "District": "6. Bezirk – Mariahilf",
        "Phone": "",
        "Email": "",
        "Website": "https://www.demo-privatpraxis.at",
        "Praxisplan Profile URL": "https://www.praxisplan.at/EXAMPLE",
        "Priority Score": "2",
        "Priority Reason": "Website missing, phone only",
    },
    {
        "Doctor Name": "Demo Familienpraxis Wien",
        "Title": "Dr.",
        "Practice Name": "Demo Familienpraxis Wien",
        "Specialty": "Allgemeinmedizin",
        "Address": "Hauptstraße 22/4, 1090 Wien",
        "Postal Code": "1090",
        "City": "Wien",
        "District": "9. Bezirk – Alsergrund",
        "Phone": "+43 1 999 88 77",
        "Email": "",
        "Website": "https://www.demo-familie.at",
        "Praxisplan Profile URL": "https://www.praxisplan.at/EXAMPLE",
        "Priority Score": "3",
        "Priority Reason": "Needs manual review",
    },
    {
        "Doctor Name": "Demo Ordination Innere Stadt",
        "Title": "Univ.Prof. Dr.",
        "Practice Name": "Demo Ordination Innere Stadt",
        "Specialty": "Kinder- u. Jugendpsychiatrie u. Psychotherapeutische Medizin",
        "Address": "Ringstraße 4, 1010 Wien",
        "Postal Code": "1010",
        "City": "Wien",
        "District": "1. Bezirk – Innere Stadt",
        "Phone": "+43 1 555 00 11",
        "Email": "ordination@example.at",
        "Website": "https://www.demo-ordination.at",
        "Praxisplan Profile URL": "https://www.praxisplan.at/EXAMPLE",
        "Priority Score": "5",
        "Priority Reason": "Public phone, Vienna, high callback specialty",
    },
]

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"(\+43[\d\s/\-]{6,}|\b0[\d\s/\-]{8,})")
POSTAL_RE = re.compile(r"\b(\d{4})\b")


def _parse_address(raw: str) -> tuple[str, str, str, str]:
    """Return (address, postal_code, city, district) from a raw address string."""
    postal_match = POSTAL_RE.search(raw)
    postal = postal_match.group(1) if postal_match else ""
    city = ""
    district = VIENNA_DISTRICTS.get(postal, "")

    if postal:
        # Everything after postal code is city
        after = raw[postal_match.end():].strip(" ,")
        city = after.split(",")[0].strip() if after else ""
        # Street = everything before postal
        street = raw[:postal_match.start()].strip(" ,")
    else:
        street = raw

    return street, postal, city, district


def _extract_title(name: str) -> tuple[str, str]:
    """Split academic title prefix from doctor name."""
    title_prefixes = (
        "Univ.Prof. Dr.", "Univ.Doz. Dr.", "Prof. Dr.", "Priv.Doz. Dr.",
        "Prim. Dr.", "OA Dr.", "Univ.Prof.", "Prof.", "Prim.", "Dr.",
        "Mag.", "DI", "MSc", "Bakk.", "BSc"
    )
    for prefix in title_prefixes:
        if name.startswith(prefix + " "):
            return prefix, name[len(prefix):].strip()
    return "", name


def _clean_specialty(raw: str) -> str:
    for prefix in ("Ordination: ", "Angestellt: ", "Angestellte*r: "):
        if raw.startswith(prefix):
            return raw[len(prefix):]
    return raw


def _score_priority(record: dict) -> tuple[int, str]:
    score = 0
    reasons = []

    if record.get("Phone"):
        score += 1
        reasons.append("public phone")
    if record.get("Email") or record.get("Website"):
        score += 1
        reasons.append("email/website available" if record.get("Email") else "website available")
    # Wahlarzt / private indicator
    spec = record.get("Specialty", "").lower()
    name = record.get("Doctor Name", "").lower()
    if "wahl" in spec or "privat" in spec:
        score += 1
        reasons.append("Wahlarzt/private")
    if record.get("City", "").lower() in ("wien", "vienna"):
        score += 1
        reasons.append("Vienna")
    # High callback specialty
    if "jugend" in spec or "kind" in spec or "psych" in spec:
        score += 1
        reasons.append("high callback specialty")

    score = min(score, 5)
    reason_str = ", ".join(reasons).capitalize() if reasons else "Needs manual review"
    return score, reason_str


def _make_blank_row(lead_id: int, source_url: str) -> dict:
    row: dict = {col: "" for col in COLUMNS}
    row["Lead ID"] = f"LEAD-{lead_id:04d}"
    row["Source"] = "Praxisplan"
    row["Source URL"] = source_url
    row["Outreach Status"] = "Not contacted"
    row["Demo Offered"] = "No"
    row["Demo Booked"] = "No"
    row["Pilot Interest"] = "Unknown"
    row["Likely LATIDO / Online Booking"] = "Unknown"
    row["Existing System Mentioned"] = ""
    row["Last Updated"] = str(date.today())
    return row


# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

def _fetch(url: str, session: "requests.Session") -> Optional["BeautifulSoup"]:
    try:
        r = session.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        return BeautifulSoup(r.text, "html.parser")
    except Exception as exc:
        print(f"  [WARN] Could not fetch {url}: {exc}", file=sys.stderr)
        return None


def _parse_listing_page(soup: "BeautifulSoup", source_url: str) -> list[dict]:
    records = []
    for block in soup.select("div.search-result"):
        name_el = block.select_one("p.label.mb-0")
        link_el = block.select_one("a[href]")
        if not name_el or not link_el:
            continue

        full_name = name_el.get_text(strip=True)
        title, doctor_name = _extract_title(full_name)
        profile_href = link_el["href"]
        profile_url = (BASE_URL + profile_href) if profile_href.startswith("/") else profile_href

        # Parse address, phone, specialty from link text
        link_text = link_el.get_text(separator="\n", strip=True)
        lines = [ln.strip() for ln in link_text.split("\n") if ln.strip()]

        specialty_raw = lines[0] if lines else ""
        specialty = _clean_specialty(specialty_raw)

        address_raw = ""
        phone = ""
        for line in lines[1:]:
            if re.match(r"^\+43", line) or re.match(r"^0[1-9]", line):
                phone = line
            elif re.search(r"\d{4}", line):
                address_raw = line
            elif not address_raw and len(line) > 5:
                address_raw = line

        address, postal, city, district = _parse_address(address_raw)

        rec = {
            "Doctor Name": full_name,
            "Title": title,
            "Practice Name": "",
            "Specialty": specialty,
            "Sub-specialty / Notes from Listing": "",
            "Address": address,
            "Postal Code": postal,
            "City": city,
            "District": district,
            "Phone": phone,
            "Email": "",
            "Website": "",
            "Praxisplan Profile URL": profile_url,
            "Source URL": source_url,
        }
        records.append(rec)
    return records


def _enrich_from_profile(rec: dict, session: "requests.Session") -> None:
    """Visit the Praxisplan profile page for website and email (public only)."""
    profile_url = rec.get("Praxisplan Profile URL", "")
    if not profile_url or "EXAMPLE" in profile_url:
        return
    soup = _fetch(profile_url, session)
    if not soup:
        return

    for a in soup.find_all("a", href=True):
        href = str(a["href"])
        if href.startswith("mailto:"):
            email = href[7:].split("?")[0].strip()
            if EMAIL_RE.match(email):
                rec["Email"] = email
        elif (
            href.startswith("http")
            and "praxisplan.at" not in href
            and "aekwien.at" not in href
            and "google.com" not in href
            and "openstreetmap.org" not in href
        ):
            if not rec.get("Website"):
                rec["Website"] = href

    # Fallback: emails in page text
    if not rec.get("Email"):
        emails = EMAIL_RE.findall(soup.get_text())
        if emails:
            rec["Email"] = emails[0]


def scrape(source_url: str, enrich_profiles: bool = True) -> list[dict]:
    if not _HAS_REQUESTS:
        print("[ERROR] requests / beautifulsoup4 not installed. Use --template-only or install dependencies.", file=sys.stderr)
        return []

    session = requests.Session()
    all_records: list[dict] = []
    seen: set[str] = set()
    errors = 0

    page = 1
    while True:
        paged_url = source_url + (f"&page={page}" if page > 1 else "")
        print(f"  Fetching page {page}: {paged_url}")
        soup = _fetch(paged_url, session)
        if not soup:
            errors += 1
            break

        records = _parse_listing_page(soup, source_url)
        if not records:
            break

        for rec in records:
            dedup_key = (rec["Doctor Name"].lower(), rec["Phone"], rec.get("Praxisplan Profile URL", ""))
            if dedup_key in seen:
                continue
            seen.add(dedup_key)
            all_records.append(rec)

        print(f"    → {len(records)} entries, {len(all_records)} total so far")

        # Check for next page
        next_links = soup.select(".pagination a[href]")
        has_next = any(f"page={page + 1}" in (a.get("href", "") or "") for a in next_links)
        if not has_next and len(records) < 15:
            break
        if not has_next:
            # Try next page anyway in case pagination is not shown
            page += 1
            delay = REQUEST_DELAY_MIN + (REQUEST_DELAY_MAX - REQUEST_DELAY_MIN) * 0.5
            time.sleep(delay)
            # Probe next page
            probe = _fetch(source_url + f"&page={page}", session)
            if probe and _parse_listing_page(probe, source_url):
                records2 = _parse_listing_page(probe, source_url)
                for rec in records2:
                    dedup_key = (rec["Doctor Name"].lower(), rec["Phone"], rec.get("Praxisplan Profile URL", ""))
                    if dedup_key in seen:
                        continue
                    seen.add(dedup_key)
                    all_records.append(rec)
                print(f"    → page {page}: {len(records2)} entries, {len(all_records)} total")
                page += 1
                continue
            else:
                break
        else:
            page += 1

        time.sleep(REQUEST_DELAY_MIN + (REQUEST_DELAY_MAX - REQUEST_DELAY_MIN) * 0.5)

    if enrich_profiles:
        print(f"\n  Enriching {len(all_records)} profiles (this takes a while)...")
        for i, rec in enumerate(all_records):
            print(f"    [{i + 1}/{len(all_records)}] {rec['Doctor Name']}")
            _enrich_from_profile(rec, session)
            time.sleep(REQUEST_DELAY_MIN + (REQUEST_DELAY_MAX - REQUEST_DELAY_MIN) * 0.5)

    return all_records


def parse_local_html(html_path: str, source_url: str) -> list[dict]:
    with open(html_path, encoding="utf-8", errors="replace") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")
    return _parse_listing_page(soup, source_url)


# ---------------------------------------------------------------------------
# Build rows
# ---------------------------------------------------------------------------

def build_rows(records: list[dict], source_url: str) -> list[dict]:
    rows = []
    for i, rec in enumerate(records, 1):
        row = _make_blank_row(i, source_url)
        for col in ("Doctor Name", "Title", "Practice Name", "Specialty",
                    "Sub-specialty / Notes from Listing", "Address", "Postal Code",
                    "City", "District", "Phone", "Email", "Website",
                    "Praxisplan Profile URL"):
            row[col] = rec.get(col, "")
        score, reason = _score_priority(rec)
        row["Priority Score"] = str(score)
        row["Priority Reason"] = reason
        rows.append(row)
    return rows


def build_example_rows(source_url: str) -> list[dict]:
    rows = []
    for i, ex in enumerate(FAKE_EXAMPLE_ROWS, 1):
        row = _make_blank_row(i, source_url)
        for k, v in ex.items():
            row[k] = v
        if not row.get("Priority Score"):
            row["Priority Score"] = "3"
            row["Priority Reason"] = "Needs manual review"
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _dv(formula: str, sqref: str) -> "DataValidation":
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = sqref
    return dv


def _col_idx(name: str) -> int:
    return COLUMNS.index(name) + 1


def _col_letter(name: str) -> str:
    return get_column_letter(_col_idx(name))


def save_xlsx(rows: list[dict], output_path: str) -> None:
    if not _HAS_OPENPYXL:
        print("[ERROR] openpyxl not installed. Cannot create XLSX.", file=sys.stderr)
        return

    wb = openpyxl.Workbook()

    # --- Leads sheet ---
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Add data rows
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(COLUMNS, 1):
            ws.cell(row=ri, column=ci, value=row.get(col, ""))

    # Dropdowns — apply to rows 2:2000
    max_row = 2000

    def dv_range(col_name: str) -> str:
        return f"{_col_letter(col_name)}2:{_col_letter(col_name)}{max_row}"

    def quoted(values: list[str]) -> str:
        return '"' + ",".join(values) + '"'

    dvs = [
        _dv(quoted(OUTREACH_STATUS_VALUES), dv_range("Outreach Status")),
        _dv(quoted(CALL_RESULT_VALUES), dv_range("Call Attempt 1 Result")),
        _dv(quoted(CALL_RESULT_VALUES), dv_range("Call Attempt 2 Result")),
        _dv(quoted(EMAIL_RESULT_VALUES), dv_range("Email Result")),
        _dv(quoted(WALK_IN_RESULT_VALUES), dv_range("Walk-in Result")),
        _dv(quoted(DEMO_YN_VALUES), dv_range("Demo Offered")),
        _dv(quoted(DEMO_YN_VALUES), dv_range("Demo Booked")),
        _dv(quoted(PILOT_INTEREST_VALUES), dv_range("Pilot Interest")),
        _dv(quoted(PRIORITY_SCORE_VALUES), dv_range("Priority Score")),
    ]
    for dv in dvs:
        ws.add_data_validation(dv)

    # Column widths
    col_widths = {
        "Lead ID": 12, "Doctor Name": 28, "Title": 14, "Practice Name": 24,
        "Specialty": 34, "Sub-specialty / Notes from Listing": 30,
        "Address": 32, "Postal Code": 12, "City": 14, "District": 28,
        "Phone": 20, "Email": 28, "Website": 30, "Praxisplan Profile URL": 38,
        "Source URL": 20, "Source": 12, "Existing System Mentioned": 22,
        "Likely LATIDO / Online Booking": 22, "Priority Score": 14,
        "Priority Reason": 30, "Outreach Status": 22,
        "Call Attempt 1 Date": 18, "Call Attempt 1 Result": 22,
        "Call Attempt 2 Date": 18, "Call Attempt 2 Result": 22,
        "Email Sent Date": 16, "Email Result": 20,
        "Walk-in Date": 16, "Walk-in Result": 22,
        "Contact Person": 22, "Best Time to Call": 18,
        "Follow-up Date": 16, "Demo Offered": 14, "Demo Booked": 14,
        "Demo Date": 14, "Pilot Interest": 16, "Objection": 30,
        "Next Action": 30, "Notes": 40, "Last Updated": 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[_col_letter(col)].width = width

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 16

    leads_col = _col_letter("Lead ID")
    phone_col = _col_letter("Phone")
    email_col = _col_letter("Email")
    website_col = _col_letter("Website")
    status_col = _col_letter("Outreach Status")
    demo_col = _col_letter("Demo Booked")

    summary_data = [
        ("Metric", "Count"),
        ("Total leads", f"=COUNTA(Leads!{leads_col}2:{leads_col}{max_row})"),
        ("Phone available", f"=COUNTIF(Leads!{phone_col}2:{phone_col}{max_row},\"?*\")"),
        ("Email available", f"=COUNTIF(Leads!{email_col}2:{email_col}{max_row},\"?*\")"),
        ("Website available", f"=COUNTIF(Leads!{website_col}2:{website_col}{max_row},\"?*\")"),
        ("Not contacted", f"=COUNTIF(Leads!{status_col}2:{status_col}{max_row},\"Not contacted\")"),
        ("Called", f"=COUNTIF(Leads!{status_col}2:{status_col}{max_row},\"Called\")"),
        ("Demo booked", f"=COUNTIF(Leads!{demo_col}2:{demo_col}{max_row},\"Yes\")"),
        ("Interested", f"=COUNTIF(Leads!{status_col}2:{status_col}{max_row},\"Interested\")"),
        ("Follow-up needed", f"=COUNTIF(Leads!{status_col}2:{status_col}{max_row},\"Follow-up needed\")"),
    ]

    for ri, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=ri, column=1, value=label).font = Font(bold=(ri == 1))
        ws2.cell(row=ri, column=2, value=value)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"  Saved XLSX: {output_path}")


def save_csv(rows: list[dict], output_path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Saved CSV:  {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Build Praxisplan lead database")
    parser.add_argument("--url", default=DEFAULT_SOURCE_URL, help="Praxisplan search URL")
    parser.add_argument("--input-html", help="Path to locally saved HTML file (Mode B)")
    parser.add_argument("--output", default="docs/sales/outreach/praxisplan_child_psychiatry_leads.xlsx",
                        help="Output XLSX path")
    parser.add_argument("--no-enrich", action="store_true",
                        help="Skip visiting individual profile pages")
    parser.add_argument("--template-only", action="store_true",
                        help="Create template with fake example rows only (no network)")
    args = parser.parse_args()

    source_url = args.url
    xlsx_path = args.output
    csv_path = xlsx_path.replace(".xlsx", ".csv")

    if args.template_only:
        print("Creating template with example rows only (no network access)...")
        rows = build_example_rows(source_url)
    elif args.input_html:
        print(f"Parsing local HTML: {args.input_html}")
        records = parse_local_html(args.input_html, source_url)
        print(f"  Parsed {len(records)} records from local file")
        rows = build_rows(records, source_url)
    else:
        print(f"Fetching from: {source_url}")
        records = scrape(source_url, enrich_profiles=not args.no_enrich)
        print(f"\n  Total scraped: {len(records)}")
        rows = build_rows(records, source_url)

    print(f"\nBuilding output files...")
    save_csv(rows, csv_path)
    save_xlsx(rows, xlsx_path)

    # Summary
    phone_count = sum(1 for r in rows if r.get("Phone"))
    email_count = sum(1 for r in rows if r.get("Email"))
    website_count = sum(1 for r in rows if r.get("Website"))
    print(f"\n{'=' * 50}")
    print(f"  Total rows:      {len(rows)}")
    print(f"  With phone:      {phone_count}")
    print(f"  With email:      {email_count}")
    print(f"  With website:    {website_count}")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    main()
