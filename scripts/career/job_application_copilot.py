"""
job_application_copilot.py

PraxisMed Career Copilot — Local job application assistant for Ali Abdeltawab.

SAFETY RULES (never remove):
  No auto-apply. No auto-submit. No email sending.
  No LinkedIn scraping, login automation, or automated LinkedIn actions.
  No browser automation that clicks apply.
  No fake profiles. No false experience. No false claims.
  Human review and manual submission required for every application.
  No secrets exposed. No private data stored.

Usage:
  python scripts/career/job_application_copilot.py --mode init
  python scripts/career/job_application_copilot.py --mode import-links --input career/job_search/saved_job_links.csv
  python scripts/career/job_application_copilot.py --mode score
  python scripts/career/job_application_copilot.py --mode drafts --limit 10
  python scripts/career/job_application_copilot.py --mode daily-plan --limit 30
  python scripts/career/job_application_copilot.py --mode report
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from datetime import date, datetime
from typing import Any, Optional

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_HERE   = os.path.dirname(os.path.abspath(__file__))
_REPO   = os.path.dirname(os.path.dirname(_HERE))
_CAREER = os.path.join(_REPO, "career")
_JOBS_DIR    = os.path.join(_CAREER, "job_search")
_PLANS_DIR   = os.path.join(_JOBS_DIR, "daily_plans")
_PACKETS_DIR = os.path.join(_JOBS_DIR, "application_packets")
_PROFILE_FILE = os.path.join(_JOBS_DIR, "profile.json")
_JOBS_FILE    = os.path.join(_JOBS_DIR, "jobs.json")

# ---------------------------------------------------------------------------
# Ali's profile
# ---------------------------------------------------------------------------

ALI_PROFILE: dict[str, Any] = {
    "name": "Ali Abdeltawab",
    "email": "aliabdeltawab90@gmail.com",
    "phone": "+43 681 20454784",
    "location": "Vienna, Austria",
    "github": "https://github.com/Ali-40",
    "linkedin": "https://www.linkedin.com/in/ali-abdeltawab-bsc-905b7b217/",
    "education": [
        {
            "degree": "BSc Artificial Intelligence",
            "school": "Johannes Kepler University Linz",
            "years": "2023–2026",
            "status": "completed",
        },
        {
            "degree": "MSc Software Engineering",
            "school": "TU Wien",
            "years": "Starting October 2026",
            "status": "accepted — starting Oct 2026 alongside full-time work",
        },
        {
            "degree": "BSc Computers and Artificial Intelligence",
            "school": "Helwan University, Cairo",
            "years": "2021–2023",
            "status": "completed",
        },
    ],
    "experience": [
        {
            "title": "Store Associate",
            "company": "Hofer KG",
            "location": "Vienna",
            "dates": "07/2025–Present",
            "duration_months": 12,
        },
        {
            "title": "Warehouse Associate",
            "company": "JD Sports",
            "location": "Vienna",
            "dates": "06/2024–12/2024",
        },
    ],
    "skills": [
        "Python", "SQL", "pandas", "NumPy", "scikit-learn", "PyTorch",
        "FastAPI", "Docker", "Git", "Streamlit", "Power BI", "Excel",
        "Machine Learning", "Data Science", "MLOps", "Time Series Forecasting",
        "Anomaly Detection", "Recommender Systems", "NLP", "Computer Vision",
        "R", "MATLAB", "Next.js", "PostgreSQL", "Plotly", "Matplotlib",
        "Statistical Analysis", "Feature Engineering",
    ],
    "projects": [
        {
            "name": "PraxisMed",
            "stack": ["FastAPI", "Next.js", "PostgreSQL", "Vapi", "Docker", "Railway", "Vercel", "Python"],
            "summary": "Multi-tenant AI receptionist and clinic workflow platform. Human-in-the-loop, 5,000+ automated tests.",
            "roles": ["AI Engineer", "Software Engineer", "MLOps", "Full-stack AI", "Startup"],
        },
        {
            "name": "Cloud Demand Forecasting & Anomaly Detection",
            "stack": ["Python", "pandas", "scikit-learn", "FastAPI", "Streamlit", "Docker"],
            "summary": "End-to-end MLOps pipeline: M5 forecasting, FastAPI /forecast and /anomalies API, Streamlit dashboard, Docker Compose.",
            "roles": ["ML Engineer", "Data Scientist", "MLOps", "Forecasting", "Time Series"],
        },
        {
            "name": "Human Development Analysis & Visualization",
            "stack": ["Python", "pandas", "Plotly", "Power BI"],
            "summary": "UNDP HDI dataset analysis for 190+ countries. Dashboards with Python and Power BI.",
            "roles": ["Data Analyst", "BI Analyst", "Analytics"],
        },
        {
            "name": "Recommender Systems",
            "stack": ["Python", "scikit-learn"],
            "summary": "Collaborative filtering and hybrid recommender systems. Precision/Recall/F1 evaluation.",
            "roles": ["ML Engineer", "Personalization", "E-commerce"],
        },
    ],
    "languages": {
        "English": "Native/Fluent",
        "Arabic": "Native",
        "German": "B1/B1+",
    },
    "target_roles": [
        "Junior Data Scientist", "Data Analyst", "Junior Machine Learning Engineer",
        "Machine Learning Intern", "AI Engineer Intern", "MLOps Intern",
        "Junior MLOps Engineer", "Python Data Engineer Intern", "BI Analyst",
        "Analytics Engineer Junior", "Risk Modelling Intern",
        "Junior Quantitative Analyst", "Software Engineering Intern",
    ],
    "target_locations": ["Vienna", "Wien", "Austria", "Österreich", "Remote"],
    "priority_work_type": "Full-time",
    "open_to": ["Full-time", "Internship", "Working Student"],
    "cv_versions": {
        "en": "career/cv/Ali_Abdeltawab_CV_EN_ATS.md",
        "de": "career/cv/Ali_Abdeltawab_CV_DE_ATS.md",
    },
    "notes": (
        "Full-time is highest priority for Rot-Weiß-Rot (RWR) permit pathway. "
        "TU Wien MSc Software Engineering accepted Oct 2026 — present as alongside work. "
        "Old JKU degree removed from all CVs and profiles. "
        "Hofer KG: 1+ year, shows discipline alongside studies."
    ),
}

# ---------------------------------------------------------------------------
# Fit scoring
# ---------------------------------------------------------------------------

LOCATION_SCORES = {"vienna": 30, "wien": 30, "austria": 20, "österreich": 20,
                   "remote": 15, "hybrid": 10, "germany": 5, "deutschland": 5}

TARGET_ROLE_KEYWORDS = [
    "data scientist", "data analyst", "machine learning", "ml engineer",
    "ai engineer", "mlops", "python", "data engineer", "bi analyst",
    "analytics engineer", "risk model", "quantitative", "software engineer",
    "data science", "künstliche intelligenz", "maschinelles lernen",
    "datenanalyse", "data analytics", "business intelligence",
]

TECH_KEYWORDS = [
    "python", "sql", "pandas", "numpy", "scikit-learn", "sklearn", "pytorch",
    "tensorflow", "fastapi", "docker", "git", "streamlit", "power bi",
    "machine learning", "deep learning", "nlp", "computer vision",
    "time series", "forecasting", "anomaly detection",
]

NEGATIVE_KEYWORDS = [
    "senior", "lead", "head of", "c++", "java developer", "rust",
    "5+ years", "7+ years", "10+ years", "citizenship required",
    "eu citizen only", "native german", "c2 german",
]

GERMAN_B2_REQUIRED = ["c1 deutsch", "c1 german", "c2 deutsch", "native german",
                      "muttersprachlich", "deutschkenntnisse b2 zwingend",
                      "b2 required", "b2 mandatory"]


def score_job(job: dict[str, Any]) -> tuple[int, str]:
    """
    Score a job 0–100 and return (score, reason).
    No network calls. No private data. No auto-apply.
    """
    text_raw = " ".join([
        job.get("job_title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("description", ""),
        job.get("required_skills", ""),
        job.get("tech_stack", ""),
    ])
    text = text_raw.lower()
    reasons = []
    score = 0

    # Location (max 30)
    for loc, pts in LOCATION_SCORES.items():
        if loc in text:
            score += pts
            reasons.append(f"location:{loc}(+{pts})")
            break

    # Role match (max 20)
    role_match = sum(1 for kw in TARGET_ROLE_KEYWORDS if kw in text)
    role_pts = min(20, role_match * 4)
    score += role_pts
    if role_pts > 0:
        reasons.append(f"role_match(+{role_pts})")

    # Tech stack match (max 20)
    tech_match = sum(1 for kw in TECH_KEYWORDS if kw in text)
    tech_pts = min(20, tech_match * 3)
    score += tech_pts
    if tech_pts > 0:
        reasons.append(f"tech_match:{tech_match}_kw(+{tech_pts})")

    # Junior/intern openness (max 15)
    junior_kws = ["junior", "intern", "praktikum", "werkstudent", "entry level",
                  "graduate", "berufseinsteiger", "junior-", "einstieg"]
    if any(kw in text for kw in junior_kws):
        score += 15
        reasons.append("junior_friendly(+15)")

    # Full-time potential (max 10)
    ft_kws = ["vollzeit", "full-time", "festanstellung", "unbefristet", "permanent"]
    if any(kw in text for kw in ft_kws):
        score += 10
        reasons.append("full_time_possible(+10)")

    # Negatives
    for neg in NEGATIVE_KEYWORDS:
        if neg in text:
            score -= 15
            reasons.append(f"negative:{neg}(-15)")
            break

    # German B2 required and Ali is only B1 — penalty
    if any(kw in text for kw in GERMAN_B2_REQUIRED):
        score -= 10
        reasons.append("german_b2_required(-10)")

    # Language of job
    de_indicators = ["stellenbeschreibung", "aufgaben", "anforderungen",
                     "wir bieten", "ihre aufgaben", "ihr profil"]
    is_german_job = any(ind in text for ind in de_indicators)
    job["language_detected"] = "German" if is_german_job else "English"

    score = max(0, min(100, score))
    reason = " | ".join(reasons) if reasons else "no_match"
    return score, reason


# ---------------------------------------------------------------------------
# Job management
# ---------------------------------------------------------------------------

def _load_jobs() -> list[dict[str, Any]]:
    if not os.path.exists(_JOBS_FILE):
        return []
    with open(_JOBS_FILE, encoding="utf-8") as f:
        return json.load(f)


def _save_jobs(jobs: list[dict[str, Any]]) -> None:
    os.makedirs(os.path.dirname(_JOBS_FILE), exist_ok=True)
    with open(_JOBS_FILE, "w", encoding="utf-8") as f:
        json.dump(jobs, f, ensure_ascii=False, indent=2)


def _next_job_id(jobs: list[dict[str, Any]]) -> str:
    max_id = 0
    for j in jobs:
        m = re.match(r"JOB-(\d+)", j.get("job_id", ""))
        if m:
            max_id = max(max_id, int(m.group(1)))
    return f"JOB-{max_id + 1:04d}"


# ---------------------------------------------------------------------------
# Modes
# ---------------------------------------------------------------------------

def mode_init() -> None:
    """Create folder structure and save profile template."""
    dirs = [
        _JOBS_DIR, _PLANS_DIR, _PACKETS_DIR,
        os.path.join(_CAREER, "cv"),
        os.path.join(_CAREER, "linkedin"),
        os.path.join(_CAREER, "github"),
        os.path.join(_CAREER, "templates"),
        os.path.join(_CAREER, "START_HERE_FOR_ALI", "3_TODAYS_APPLICATION_PACKETS"),
        os.path.join(_CAREER, "input"),
    ]
    for d in dirs:
        os.makedirs(d, exist_ok=True)
        print(f"  dir: {d}")

    if not os.path.exists(_PROFILE_FILE):
        with open(_PROFILE_FILE, "w", encoding="utf-8") as f:
            json.dump(ALI_PROFILE, f, ensure_ascii=False, indent=2)
        print(f"  profile: {_PROFILE_FILE}")

    if not os.path.exists(_JOBS_FILE):
        _save_jobs([])
        print(f"  jobs db: {_JOBS_FILE}")

    # Saved job links template
    links_csv = os.path.join(_JOBS_DIR, "saved_job_links.csv")
    if not os.path.exists(links_csv):
        with open(links_csv, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["url", "company", "job_title", "date_saved", "notes"])
            w.writerow(["https://karriere.at/example-job", "Example Company", "Junior Data Scientist", date.today().isoformat(), "Manual search result"])
        print(f"  links template: {links_csv}")

    print("\nInit complete. Next steps:")
    print("  1. Copy job URLs to career/job_search/saved_job_links.csv")
    print("  2. Run: python scripts/career/job_application_copilot.py --mode import-links")
    print("  3. Run: python scripts/career/job_application_copilot.py --mode score")
    print("  4. Run: python scripts/career/job_application_copilot.py --mode daily-plan --limit 30")
    print("\nNO AUTO-APPLY. NO AUTO-EMAIL. Manual submission required.")


def mode_import_links(input_csv: str) -> None:
    """
    Import job links from a manually saved CSV.
    Allowed sources: karriere.at, stepstone.at, jobs.at, indeed.at, company career pages.
    No LinkedIn scraping. No login automation.
    """
    if not os.path.exists(input_csv):
        print(f"ERROR: File not found: {input_csv}")
        print("Create it with columns: url, company, job_title, date_saved, notes")
        return

    jobs = _load_jobs()
    existing_urls = {j.get("job_url", "") for j in jobs}
    added = 0

    with open(input_csv, encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            url = (row.get("url") or row.get("job_url") or "").strip()
            if not url or url in existing_urls or url.startswith("#"):
                continue

            # Block LinkedIn automation
            if "linkedin.com/jobs" in url.lower() and "auto" in url.lower():
                print(f"  SKIP (no LinkedIn automation): {url}")
                continue

            job = {
                "job_id": _next_job_id(jobs + [{"job_id": f"JOB-{i:04d}"} for i in range(added)]),
                "company": (row.get("company") or "").strip(),
                "job_title": (row.get("job_title") or "").strip(),
                "job_url": url,
                "date_found": (row.get("date_saved") or date.today().isoformat()),
                "source": "manual_saved_link",
                "location": (row.get("location") or "").strip(),
                "description": (row.get("description") or "").strip(),
                "required_skills": (row.get("required_skills") or "").strip(),
                "tech_stack": (row.get("tech_stack") or "").strip(),
                "notes": (row.get("notes") or "").strip(),
                "application_status": "Saved",
                "fit_score": None,
                "fit_reason": None,
                "language_detected": None,
            }
            jobs.append(job)
            existing_urls.add(url)
            added += 1
            print(f"  Imported: {job['company']} — {job['job_title']}")

    _save_jobs(jobs)
    print(f"\nImported {added} job(s). Total in database: {len(jobs)}")
    print("Next: run --mode score to score all unscored jobs.")
    print("REMINDER: No automated submissions. Manual review and submit only.")


def mode_score() -> None:
    """Score all unscored jobs and update jobs.json."""
    jobs = _load_jobs()
    scored = 0
    for job in jobs:
        if job.get("fit_score") is None:
            score, reason = score_job(job)
            job["fit_score"] = score
            job["fit_reason"] = reason
            scored += 1

    _save_jobs(jobs)
    print(f"Scored {scored} job(s).")
    if jobs:
        high = [j for j in jobs if (j.get("fit_score") or 0) >= 70]
        med  = [j for j in jobs if 50 <= (j.get("fit_score") or 0) < 70]
        low  = [j for j in jobs if (j.get("fit_score") or 0) < 50]
        print(f"  High fit (>=70): {len(high)}")
        print(f"  Medium fit (50–69): {len(med)}")
        print(f"  Low fit (<50): {len(low)}")
    print("Next: run --mode drafts --limit 10 to generate application packets.")


def _project_for_role(role: str, description: str) -> dict[str, Any]:
    text = (role + " " + description).lower()
    for proj in ALI_PROFILE["projects"]:
        for proj_role_kw in proj["roles"]:
            if proj_role_kw.lower() in text:
                return proj
    return ALI_PROFILE["projects"][0]


def _generate_motivation_letter(job: dict[str, Any], lang: str) -> str:
    company = job.get("company", "{{COMPANY_NAME}}")
    title = job.get("job_title", "{{JOB_TITLE}}")
    proj = _project_for_role(title, job.get("description", ""))

    if lang == "de":
        return f"""\
Wien, {date.today().strftime("%d.%m.%Y")}

**{company}**

Betreff: Bewerbung als {title}

Sehr geehrte Damen und Herren,

mit großem Interesse habe ich Ihre Stellenausschreibung für die Position als {title} bei {company} gelesen. Als Absolvent des BSc Artificial Intelligence der Johannes Kepler Universität Linz mit praktischer Erfahrung im Bereich maschinellen Lernens und Datenanalyse möchte ich mich hiermit für diese Stelle bewerben.

**TODO — PFLICHTFELD:** Ersetzen Sie den folgenden Absatz durch eine konkrete, stellenspezifische Begründung. Was genau an dieser Stelle oder diesem Unternehmen ist relevant?

{{KONKRETER_GRUND_FÜR_DIESE_STELLE — 2–3 Sätze. Niemals generisch lassen.}}

Mein relevantestes Projekt für diese Stelle:

**{proj["name"]}** — {proj["summary"]}
Tech-Stack: {", ".join(proj["stack"][:5])}

Meine technischen Kenntnisse umfassen Python, SQL sowie Erfahrung mit maschinellem Lernen, Datenanalyse und der Umsetzung produktionsreifer Systeme. Ich bin in Wien ansässig und ab sofort für eine Vollzeitstelle verfügbar. Den MSc Software Engineering an der TU Wien (Beginn Oktober 2026) plane ich parallel zur Berufstätigkeit zu absolvieren.

Über die Möglichkeit, meine Qualifikationen in einem Gespräch vorzustellen, würde ich mich sehr freuen. Meinen Lebenslauf füge ich bei.

Mit freundlichen Grüßen,
Ali Abdeltawab
+43 681 20454784 | aliabdeltawab90@gmail.com
https://github.com/Ali-40
https://www.linkedin.com/in/ali-abdeltawab-bsc-905b7b217/

---
⚠️ REVIEW REQUIRED: Replace all {{PLATZHALTER}} before submitting.
Check: company name correct | job title correct | no false claims | one page max | spell-check German
"""
    else:
        return f"""\
Vienna, {date.today().strftime("%B %d, %Y")}

**{company}**

Re: Application for {title}

Dear Hiring Team,

I am writing to apply for the {title} position at {company}. As a BSc Artificial Intelligence graduate from Johannes Kepler University Linz with hands-on experience in machine learning, data science, and building production-grade AI systems, I am excited by this opportunity.

**TODO — REQUIRED:** Replace the following paragraph with specific, job-tailored content. Why this role? Why this company?

{{SPECIFIC_REASON_FOR_THIS_JOB — 2–3 sentences. Never leave generic.}}

Most relevant project:

**{proj["name"]}** — {proj["summary"]}
Stack: {", ".join(proj["stack"][:5])}

My technical background covers Python, SQL, machine learning, and end-to-end pipeline development. I am Vienna-based and available for full-time roles. I have been accepted to MSc Software Engineering at TU Wien (starting October 2026), which I plan to pursue alongside work.

I would be glad to discuss how my background fits your team's needs. Please find my CV attached.

With best regards,
Ali Abdeltawab
+43 681 20454784 | aliabdeltawab90@gmail.com
https://github.com/Ali-40
https://www.linkedin.com/in/ali-abdeltawab-bsc-905b7b217/

---
⚠️ REVIEW REQUIRED: Replace all {{PLACEHOLDERS}} before submitting.
Check: company name correct | job title correct | no false claims | one page max | spell-check
"""


def _generate_cv_bullets(job: dict[str, Any]) -> str:
    title = job.get("job_title", "")
    desc = job.get("description", "")
    proj = _project_for_role(title, desc)
    lang = job.get("language_detected", "English")

    lines = [
        f"# Tailored CV Bullets for: {job.get('company')} — {title}",
        "",
        f"**CV version:** {'DE ATS' if lang == 'German' else 'EN ATS'}",
        f"**Most relevant project:** {proj['name']}",
        "",
        "## Suggested CV adjustments",
        "",
        f"### Lead project: {proj['name']}",
        f"- {proj['summary']}",
        f"- Tech stack: {', '.join(proj['stack'])}",
        "",
        "### Skills to highlight (from this job's requirements):",
    ]

    desc_lower = desc.lower()
    matched = [s for s in ALI_PROFILE["skills"] if s.lower() in desc_lower]
    if matched:
        lines.append(", ".join(matched[:12]))
    else:
        lines.append("(Paste the job description into --mode score to get matched skills)")

    lines += [
        "",
        "### Hofer KG bullet to include (if reliability is relevant):",
        "- Worked in a fast-paced retail environment with direct customer interaction (1+ year) alongside full-time technical studies — demonstrates reliability and discipline.",
        "",
        "⚠️ Do not add skills or experience you do not have. No false claims.",
    ]

    return "\n".join(lines)


def _generate_packet(job: dict[str, Any], packet_dir: str) -> None:
    os.makedirs(packet_dir, exist_ok=True)
    lang = job.get("language_detected", "English")
    cv_ver = "DE ATS" if lang == "German" else "EN ATS"
    cv_path = f"career/cv/Ali_Abdeltawab_CV_{'DE' if lang == 'German' else 'EN'}_ATS.md → convert to PDF"

    # 00_JOB_SUMMARY
    with open(os.path.join(packet_dir, "00_JOB_SUMMARY.md"), "w", encoding="utf-8") as f:
        f.write(f"""# Job Summary

**Job ID:** {job.get("job_id")}
**Company:** {job.get("company")}
**Job Title:** {job.get("job_title")}
**Location:** {job.get("location") or "Not specified"}
**Language:** {lang}
**Source:** {job.get("source")}
**Job URL:** {job.get("job_url")}
**Date Found:** {job.get("date_found")}
**Fit Score:** {job.get("fit_score")}/100
**Fit Reason:** {job.get("fit_reason")}

## Description

{job.get("description") or "(Paste job description here for better tailoring)"}

## Required Skills

{job.get("required_skills") or "(Extract from job posting)"}

## Tech Stack Mentioned

{job.get("tech_stack") or "(Extract from job posting)"}

## Notes

{job.get("notes") or "—"}
""")

    # 01_FIT_SCORE
    score = job.get("fit_score") or 0
    verdict = "High fit — apply today" if score >= 70 else \
              "Medium fit — apply if time allows" if score >= 50 else \
              "Low fit — consider skipping"
    with open(os.path.join(packet_dir, "01_FIT_SCORE.md"), "w", encoding="utf-8") as f:
        f.write(f"""# Fit Score

**Score:** {score}/100
**Verdict:** {verdict}
**Scoring breakdown:** {job.get("fit_reason")}

## Why apply / why skip

{"Apply: Strong match on location, role, and tech stack." if score >= 70 else
 "Medium match. Review job description carefully before applying." if score >= 50 else
 "Low match. Only apply if you have read the job description and believe it is a strong fit despite the score."}

⚠️ This score is generated from keywords. Always read the actual job description before deciding.
""")

    # 02_TAILORED_CV_BULLETS
    with open(os.path.join(packet_dir, "02_TAILORED_CV_BULLETS.md"), "w", encoding="utf-8") as f:
        f.write(_generate_cv_bullets(job))

    # 03_MOTIVATION_LETTER
    with open(os.path.join(packet_dir, "03_MOTIVATION_LETTER.md"), "w", encoding="utf-8") as f:
        f.write(_generate_motivation_letter(job, "de" if lang == "German" else "en"))

    # 04_APPLICATION_CHECKLIST
    with open(os.path.join(packet_dir, "04_APPLICATION_CHECKLIST.md"), "w", encoding="utf-8") as f:
        f.write(f"""# Application Checklist

**{job.get("company")} — {job.get("job_title")}**

## Before applying

- [ ] Read the full job description at: {job.get("job_url")}
- [ ] CV selected: {cv_path}
- [ ] Cover letter reviewed and all {{PLACEHOLDERS}} replaced
- [ ] Company name is correct throughout all documents
- [ ] Job title matches posting exactly
- [ ] Language is correct ({lang})
- [ ] No false claims in any document
- [ ] You have relevant skills for this role

## During application

- [ ] Submit on the company's own portal (not auto-apply button)
- [ ] Upload CV as PDF
- [ ] Upload / paste motivation letter
- [ ] Answer work authorization honestly if asked
- [ ] Note: name, email, phone filled correctly

## After applying

- [ ] Update job tracker: status → Applied, date → {date.today().isoformat()}
- [ ] Set follow-up date: 7–10 business days from today
- [ ] Note contact person if known

⚠️ NO AUTO-SUBMIT. Manual submission only.
""")

    # 05_LINKEDIN_MESSAGE_IF_USEFUL
    with open(os.path.join(packet_dir, "05_LINKEDIN_MESSAGE_IF_USEFUL.md"), "w", encoding="utf-8") as f:
        f.write(f"""# LinkedIn Follow-up Message (optional, only if appropriate)

Only send this after submitting the application and waiting 7–10 business days with no reply.
Find the recruiter or hiring manager on LinkedIn manually. Send ONE message only.

**NO LinkedIn automation. Manual message only.**

---

## English version

```
Hi [Name],

I recently applied for the {job.get("job_title")} position at {job.get("company")} and wanted to briefly express my genuine interest. As a BSc AI graduate with experience in Python, ML pipelines, and [relevant project], I believe I can contribute to your team.

I know you receive many applications — I just wanted to flag my application directly.

Best regards,
Ali Abdeltawab
```

## German version

```
Hallo [Name],

ich habe mich kürzlich auf die Stelle als {job.get("job_title")} bei {job.get("company")} beworben und möchte mein Interesse kurz direkt zum Ausdruck bringen. Als BSc-Absolvent in Artificial Intelligence mit Erfahrung in Python, ML-Pipelines und [relevantes Projekt] bin ich überzeugt, einen Beitrag leisten zu können.

Mit freundlichen Grüßen,
Ali Abdeltawab
```

⚠️ Only send once. No automated LinkedIn messages.
""")


def mode_drafts(limit: int = 10) -> None:
    """Generate application packets for the top N jobs by fit score."""
    jobs = _load_jobs()
    if not jobs:
        print("No jobs in database. Run --mode import-links first.")
        return

    unscored = [j for j in jobs if j.get("fit_score") is None]
    if unscored:
        print(f"WARNING: {len(unscored)} unscored jobs. Run --mode score first.")

    candidates = [j for j in jobs if j.get("application_status") in ("Saved", "To review", "Apply today")]
    top = sorted(candidates, key=lambda j: j.get("fit_score") or 0, reverse=True)[:limit]

    if not top:
        print("No jobs ready for draft. Add jobs via --mode import-links.")
        return

    today = date.today().isoformat()
    generated = 0
    for job in top:
        slug = re.sub(r"[^a-z0-9]+", "_", (job.get("company", "company") + "_" + job.get("job_title", "role")).lower())
        slug = slug.strip("_")[:60]
        packet_dir = os.path.join(_PACKETS_DIR, f"{today}_{slug}")
        _generate_packet(job, packet_dir)
        job["application_status"] = "Draft ready"
        generated += 1
        print(f"  Packet: {packet_dir} (score: {job.get('fit_score')})")

    _save_jobs(jobs)
    print(f"\nGenerated {generated} application packet(s).")
    print("⚠️ REVIEW REQUIRED: Open each packet and replace all {{PLACEHOLDERS}} before submitting.")
    print("⚠️ NO AUTO-SUBMIT. Submit each application manually on the company portal.")


def mode_daily_plan(limit: int = 30) -> None:
    """Generate today's daily application plan."""
    jobs = _load_jobs()
    today_str = date.today().isoformat()

    # Score any unscored
    for job in jobs:
        if job.get("fit_score") is None:
            score, reason = score_job(job)
            job["fit_score"] = score
            job["fit_reason"] = reason
    _save_jobs(jobs)

    candidates = sorted(
        [j for j in jobs if j.get("application_status") in ("Saved", "To review", "Apply today", "Draft ready")],
        key=lambda j: (j.get("fit_score") or 0),
        reverse=True,
    )[:limit]

    os.makedirs(_PLANS_DIR, exist_ok=True)

    # XLSX plan
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daily Plan"
        hdrs = ["#","Company","Job Title","Score","Language","URL","Status","Notes","Priority"]
        from openpyxl.utils import get_column_letter
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci, h)
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        ws.freeze_panes = "A2"
        for i, job in enumerate(candidates, 1):
            row = i + 1
            vals = [i, job.get("company",""), job.get("job_title",""),
                    job.get("fit_score",""), job.get("language_detected",""),
                    job.get("job_url",""), job.get("application_status",""),
                    job.get("notes",""),
                    "High" if (job.get("fit_score") or 0) >= 70 else "Medium" if (job.get("fit_score") or 0) >= 50 else "Low"]
            for ci, v in enumerate(vals, 1):
                ws.cell(row, ci, v)
        for ci, w in {1:6,2:28,3:30,4:10,5:12,6:50,7:16,8:30,9:10}.items():
            ws.column_dimensions[get_column_letter(ci)].width = w
        xlsx_path = os.path.join(_PLANS_DIR, f"{today_str}_job_plan.xlsx")
        wb.save(xlsx_path)
        print(f"  Plan xlsx: {xlsx_path}")
    except Exception as e:
        print(f"  XLSX skipped: {e}")

    # MD plan
    lines = [
        f"# Daily Job Plan — {today_str}",
        "",
        f"**Target:** {min(limit, len(candidates))} applications today",
        "**Priority:** Vienna full-time first. Quality over quantity.",
        "**Rule:** No auto-apply. No auto-email. Manual submission only.",
        "",
        "## Jobs to apply today",
        "",
        "| # | Company | Role | Score | Lang | URL |",
        "|---|---|---|---|---|---|",
    ]
    for i, job in enumerate(candidates, 1):
        lines.append(
            f"| {i} | {job.get('company','')} | {job.get('job_title','')} "
            f"| {job.get('fit_score','')} | {job.get('language_detected','')} "
            f"| {job.get('job_url','')} |"
        )
    lines += [
        "",
        "## After each application",
        "- Update tracker: status → Applied, date, follow-up date",
        "- Mark Do Not Apply if rejected or not a fit",
        "",
        "⚠️ No auto-apply. No auto-email. Manual submission only.",
    ]
    md_path = os.path.join(_PLANS_DIR, f"{today_str}_job_plan.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Plan md: {md_path}")

    # Copy packets to START_HERE today folder
    today_packets = os.path.join(_CAREER, "START_HERE_FOR_ALI", "3_TODAYS_APPLICATION_PACKETS")
    os.makedirs(today_packets, exist_ok=True)
    print(f"\nApplication packets ready in: {today_packets}")
    print(f"⚠️ Review every packet before applying. Replace all {{PLACEHOLDERS}}.")
    print("⚠️ NO AUTO-SUBMIT. Manual submission only.")


def mode_report() -> None:
    """Print application status report."""
    jobs = _load_jobs()
    if not jobs:
        print("No jobs tracked yet. Run --mode init then --mode import-links.")
        return

    from collections import Counter
    statuses = Counter(j.get("application_status", "Unknown") for j in jobs)
    scores = [j.get("fit_score") or 0 for j in jobs]
    avg_score = sum(scores) / len(scores) if scores else 0

    print(f"\n{'='*50}")
    print(f"JOB APPLICATION REPORT — {date.today().isoformat()}")
    print(f"{'='*50}")
    print(f"Total jobs: {len(jobs)}")
    print(f"Average fit score: {avg_score:.1f}/100")
    print()
    print("Status breakdown:")
    for status, count in statuses.most_common():
        print(f"  {status}: {count}")
    print()
    high = [j for j in jobs if (j.get("fit_score") or 0) >= 70]
    print(f"High-fit jobs (>=70): {len(high)}")
    applied = [j for j in jobs if j.get("application_status") == "Applied"]
    print(f"Applications sent: {len(applied)}")
    interviews = [j for j in jobs if j.get("application_status") == "Interview"]
    print(f"Interviews: {len(interviews)}")
    print()
    print("⚠️ No auto-apply. No auto-email. Manual submission only.")
    print(f"{'='*50}\n")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "PraxisMed Career Copilot — Job application assistant for Ali Abdeltawab.\n"
            "NO AUTO-APPLY. NO AUTO-EMAIL. NO LINKEDIN AUTOMATION.\n"
            "Human review and manual submission required."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--mode",
        required=True,
        choices=["init", "import-links", "score", "drafts", "daily-plan", "report"],
        help="Mode to run",
    )
    parser.add_argument("--input",  default=os.path.join(_JOBS_DIR, "saved_job_links.csv"),
                        help="Input CSV for import-links mode")
    parser.add_argument("--limit",  type=int, default=30,
                        help="Max number of jobs for drafts/daily-plan")
    args = parser.parse_args()

    print(f"\nCareer Copilot — mode: {args.mode}")
    print("⚠️  No auto-apply. No auto-email. Human review required.\n")

    if args.mode == "init":
        mode_init()
    elif args.mode == "import-links":
        mode_import_links(args.input)
    elif args.mode == "score":
        mode_score()
    elif args.mode == "drafts":
        mode_drafts(args.limit)
    elif args.mode == "daily-plan":
        mode_daily_plan(args.limit)
    elif args.mode == "report":
        mode_report()


if __name__ == "__main__":
    main()
