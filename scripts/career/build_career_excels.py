"""
build_career_excels.py

Builds career Excel files:
  career/job_search/job_tracker.xlsx            — 9-sheet master job tracker
  career/START_HERE_FOR_ALI/1_DAILY_JOB_PLAN.xlsx
  career/START_HERE_FOR_ALI/2_APPLICATION_TRACKER.xlsx

No auto-apply. No auto-email. No LinkedIn automation. Manual submission only.
"""

from __future__ import annotations

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    sys.exit(1)

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.dirname(os.path.dirname(_HERE))
_CAREER = os.path.join(_REPO, "career")
_JOB_SEARCH = os.path.join(_CAREER, "job_search")
_START_HERE = os.path.join(_CAREER, "START_HERE_FOR_ALI")

NAVY   = "1F4E79"
TEAL   = "117A65"
ORANGE = "CA6F1E"
GREEN  = "1E8449"
GRAY   = "F2F3F4"
WHITE  = "FFFFFF"
YELLOW = "FFF3CD"
RED_LT = "FADBD8"


def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(wrap=False, h="left", v="center"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)
def _dv(formula, sqref):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = sqref
    return dv


def _header(ws, headers, color=NAVY, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(color)
        c.font = _font(bold=True, color=WHITE)
        c.alignment = _align(h="center")


def _widths(ws, w_dict):
    for col, w in w_dict.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Job Tracker ──────────────────────────────────────────────────────────────

JOBS_HEADERS = [
    "Job ID", "Company", "Job Title", "Role Category", "Location",
    "Remote/Hybrid/On-site", "Language", "Source", "Job URL", "Date Found",
    "Deadline", "Work Type", "Seniority", "Visa/Work Auth Mention",
    "Tech Stack", "Required Skills", "Nice-to-have Skills", "Salary",
    "Fit Score", "Fit Reason", "Priority", "Application Status",
    "CV Version", "Cover Letter Version", "Applied Date", "Follow-up Date",
    "Contact Person", "Notes", "Next Action",
]

STATUS_DV    = '"Saved,To review,Apply today,Draft ready,Applied,Follow-up,Interview,Rejected,Offer,Not a fit"'
PRIORITY_DV  = '"High,Medium,Low"'
ROLE_DV      = '"Data Analyst,Data Scientist,ML Engineer,MLOps,AI Engineer,BI Analyst,Data Engineer,Software Engineer,Risk Modelling,Internship,Working Student"'
WORKTYPE_DV  = '"Full-time,Internship,Working Student,Part-time,Contract"'
LANG_DV      = '"German,English,German+English"'
REMOTE_DV    = '"On-site,Hybrid,Remote,Flexible"'
CV_DV        = '"EN ATS,DE ATS"'


def _build_jobs_sheet(ws):
    _header(ws, JOBS_HEADERS, NAVY)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(JOBS_HEADERS))}1"
    ws.add_data_validation(_dv(STATUS_DV,   f"V2:V1000"))
    ws.add_data_validation(_dv(PRIORITY_DV, f"U2:U1000"))
    ws.add_data_validation(_dv(ROLE_DV,     f"D2:D1000"))
    ws.add_data_validation(_dv(WORKTYPE_DV, f"L2:L1000"))
    ws.add_data_validation(_dv(LANG_DV,     f"G2:G1000"))
    ws.add_data_validation(_dv(REMOTE_DV,   f"F2:F1000"))
    ws.add_data_validation(_dv(CV_DV,       f"W2:W1000"))
    _widths(ws, {1:10,2:28,3:30,4:20,5:20,6:16,7:14,8:14,9:45,10:14,
                  11:14,12:16,13:14,14:20,15:35,16:35,17:30,18:14,
                  19:12,20:35,21:12,22:18,23:14,24:20,25:14,26:16,
                  27:22,28:40,29:30})


def _build_dashboard_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    ws.cell(1,1,"JOB APPLICATION DASHBOARD — Ali Abdeltawab").fill = _fill(NAVY)
    ws.cell(1,1).font = _font(bold=True, color=WHITE, size=14)
    ws.cell(1,1).alignment = _align(h="center")
    ws.merge_cells("A1:C1")

    ws.cell(3,1,"Total Jobs Tracked").font = _font(bold=True)
    ws.cell(3,2,'=COUNTA(Jobs!B2:B10000)').alignment = _align(h="center")

    ws.cell(4,1,"Applied").font = _font(bold=True)
    ws.cell(4,2,'=COUNTIF(Jobs!V2:V10000,"Applied")').alignment = _align(h="center")

    ws.cell(5,1,"Follow-up Pending").font = _font(bold=True)
    ws.cell(5,2,'=COUNTIF(Jobs!V2:V10000,"Follow-up")').alignment = _align(h="center")

    ws.cell(6,1,"Interviews").font = _font(bold=True)
    ws.cell(6,2,'=COUNTIF(Jobs!V2:V10000,"Interview")').alignment = _align(h="center")

    ws.cell(7,1,"Offers").font = _font(bold=True)
    ws.cell(7,2,'=COUNTIF(Jobs!V2:V10000,"Offer")').alignment = _align(h="center")

    ws.cell(8,1,"High Priority — Not Applied").font = _font(bold=True)
    ws.cell(8,2,'=COUNTIFS(Jobs!U2:U10000,"High",Jobs!V2:V10000,"Saved")').alignment = _align(h="center")

    ws.cell(10,1,"TARGET — Applications Today").font = _font(bold=True, color=ORANGE)
    ws.cell(10,2,"20–50 (quality first)").font = _font(bold=True, color=ORANGE)

    for r in range(3, 11):
        ws.row_dimensions[r].height = 22
        ws.cell(r, 1).fill = _fill(GRAY)


def _build_applications_sheet(ws):
    hdrs = ["Job ID","Company","Job Title","Applied Date","CV Used",
            "Letter Used","Status","Follow-up Date","Response","Notes"]
    _header(ws, hdrs, TEAL)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"
    _widths(ws, {1:10,2:28,3:30,4:14,5:14,6:18,7:16,8:16,9:20,10:40})


def _build_interviews_sheet(ws):
    hdrs = ["Job ID","Company","Job Title","Interview Date","Interview Type",
            "Interviewer","Preparation Notes","Outcome","Follow-up"]
    _header(ws, hdrs, GREEN)
    ws.freeze_panes = "A2"
    _widths(ws, {1:10,2:28,3:30,4:16,5:20,6:22,7:50,8:18,9:30})


def _build_followups_sheet(ws):
    hdrs = ["Job ID","Company","Job Title","Applied Date","Follow-up Date",
            "Contact Person","LinkedIn URL","Message Sent","Response","Next Action"]
    _header(ws, hdrs, ORANGE)
    ws.freeze_panes = "A2"
    _widths(ws, {1:10,2:28,3:30,4:14,5:16,6:22,7:35,8:16,9:20,10:30})


def _build_saved_links_sheet(ws):
    hdrs = ["URL","Company","Job Title","Date Saved","Source","Notes","Import Status"]
    _header(ws, hdrs, TEAL)
    ws.freeze_panes = "A2"

    banner = ws.cell(1, 1)  # overwrite header for row 1
    # Actually put banner in row 1, headers in row 2
    ws.insert_rows(1)
    ws.cell(1,1,"Paste saved job URLs here. Import with: python scripts/career/job_application_copilot.py --mode import-links")
    ws.cell(1,1).fill = _fill(YELLOW)
    ws.cell(1,1).font = Font(italic=True, size=10, name="Calibri", color="7D6608")
    ws.merge_cells(f"A1:{get_column_letter(len(hdrs))}1")
    _header(ws, hdrs, TEAL, row=2)
    ws.freeze_panes = "A3"
    _widths(ws, {1:60,2:28,3:30,4:14,5:14,6:30,7:16})


def _build_companies_sheet(ws):
    hdrs = ["Company","Website","HQ Location","Industry","Size","Open Roles URL",
            "Recruiter Name","Recruiter LinkedIn","Notes","Priority"]
    _header(ws, hdrs, NAVY)
    ws.freeze_panes = "A2"
    _widths(ws, {1:28,2:35,3:20,4:22,5:14,6:40,7:22,8:40,9:40,10:12})


def _build_keywords_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30

    ws.cell(1,1,"JOB SEARCH KEYWORDS — Ali Abdeltawab").fill = _fill(NAVY)
    ws.cell(1,1).font = _font(bold=True, color=WHITE, size=12)
    ws.cell(1,1).alignment = _align(h="center")
    ws.merge_cells("A1:B1")

    en_kws = [
        "Junior Data Scientist", "Data Analyst", "Machine Learning Engineer Junior",
        "ML Intern", "AI Engineer Intern", "MLOps Intern", "Python Data Engineer",
        "BI Analyst", "Analytics Engineer Junior", "Risk Modelling Intern",
        "Junior Quantitative Analyst", "Software Engineering Intern Python",
        "Working Student Data Science", "Working Student Machine Learning",
    ]
    de_kws = [
        "Junior Data Scientist", "Data Analyst", "Machine Learning Engineer Junior",
        "Praktikum Data Science", "Werkstudent Data Analytics", "Junior BI Analyst",
        "Python Entwickler Data", "KI Praktikum", "Machine Learning Praktikum",
        "Risk Modelling Praktikum", "Quantitative Analyst Praktikum",
        "Software Engineering Praktikum Python", "Datenanalyse Praktikum",
        "Business Intelligence Junior", "Werkstudent Machine Learning",
    ]

    ws.cell(3,1,"English search terms").fill = _fill(TEAL)
    ws.cell(3,1).font = _font(bold=True, color=WHITE)
    ws.cell(3,2,"German search terms (Stellenbörsen)").fill = _fill(TEAL)
    ws.cell(3,2).font = _font(bold=True, color=WHITE)

    for i, (en, de) in enumerate(zip(en_kws, de_kws), 4):
        ws.cell(i, 1, en).fill = _fill(GRAY if i % 2 == 0 else WHITE)
        ws.cell(i, 2, de).fill = _fill(GRAY if i % 2 == 0 else WHITE)

    ws.cell(20,1,"Location filters").fill = _fill(ORANGE)
    ws.cell(20,1).font = _font(bold=True, color=WHITE)
    for r, loc in enumerate(["1. Wien / Vienna (highest priority)",
                              "2. Austria remote or hybrid",
                              "3. Austria-wide if strong fit",
                              "4. Germany remote/hybrid only (secondary)"], 21):
        ws.cell(r, 1, loc).fill = _fill(GRAY)


def _build_templates_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80

    ws.cell(1,1,"APPLICATION TEMPLATES — Manual review required before every use").fill = _fill(NAVY)
    ws.cell(1,1).font = _font(bold=True, color=WHITE, size=12)
    ws.cell(1,1).alignment = _align(h="center")

    items = [
        ("EN CV (ATS)", "career/cv/Ali_Abdeltawab_CV_EN_ATS.md → convert to PDF"),
        ("DE CV (ATS)", "career/cv/Ali_Abdeltawab_CV_DE_ATS.md → convert to PDF"),
        ("EN Motivation Letter Template", "career/templates/motivation_letter_en.md — replace all {{PLACEHOLDERS}}"),
        ("DE Motivationsschreiben Vorlage", "career/templates/motivation_letter_de.md — alle {{PLATZHALTER}} ersetzen"),
        ("LinkedIn Profile Rewrite", "career/linkedin/LINKEDIN_PROFILE_REWRITE.md"),
        ("LinkedIn Step-by-Step", "career/linkedin/LINKEDIN_UPDATE_STEP_BY_STEP.md"),
        ("GitHub Profile README", "career/github/GITHUB_PROFILE_README.md"),
        ("Project Descriptions", "career/github/PROJECT_DESCRIPTIONS_FOR_RECRUITERS.md"),
    ]

    for i, (name, path) in enumerate(items, 3):
        ws.cell(i, 1, f"{name}: {path}").fill = _fill(GRAY if i % 2 == 0 else WHITE)


def build_job_tracker(path: str) -> None:
    wb = openpyxl.Workbook()
    sheets = ["Dashboard","Jobs","Applications","Interviews",
              "Follow-ups","Saved Job Links","Companies","Keywords","Application Templates"]
    wb.active.title = sheets[0]
    for s in sheets[1:]:
        wb.create_sheet(s)

    _build_dashboard_sheet(wb["Dashboard"])
    _build_jobs_sheet(wb["Jobs"])
    _build_applications_sheet(wb["Applications"])
    _build_interviews_sheet(wb["Interviews"])
    _build_followups_sheet(wb["Follow-ups"])
    _build_saved_links_sheet(wb["Saved Job Links"])
    _build_companies_sheet(wb["Companies"])
    _build_keywords_sheet(wb["Keywords"])
    _build_templates_sheet(wb["Application Templates"])
    wb.save(path)


# ── Daily Job Plan ───────────────────────────────────────────────────────────

def build_daily_job_plan(path: str, today: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Today's Plan"
    ws.sheet_view.showGridLines = False

    ws.cell(1,1,f"DAILY JOB PLAN — {today}").fill = _fill(NAVY)
    ws.cell(1,1).font = _font(bold=True, color=WHITE, size=14)
    ws.cell(1,1).alignment = _align(h="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30

    ws.cell(2,1,"WARNING: Quality first. 20 strong applications > 50 weak ones. Manual submission only.").fill = _fill(YELLOW)
    ws.cell(2,1).font = Font(italic=True, size=10, name="Calibri", color="7D6608")
    ws.merge_cells("A2:F2")

    plan_headers = ["#","Company","Job Title","Role Category","Job URL","Priority","Fit Score","Status","Notes"]
    _header(ws, plan_headers, TEAL, row=3)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(plan_headers))}3"

    targets = [
        ("Vienna Full-time — Data/ML/Analytics", 10, NAVY),
        ("Austria-wide — Full-time / Internship", 10, TEAL),
        ("Working Student / Internship", 10, GREEN),
        ("Extra (fit score > 70)", 20, ORANGE),
    ]
    row = 4
    job_num = 1
    for category, count, color in targets:
        ws.cell(row, 1, category).fill = _fill(color)
        ws.cell(row, 1).font = _font(bold=True, color=WHITE)
        ws.merge_cells(f"A{row}:I{row}")
        row += 1
        for _ in range(count):
            ws.cell(row, 1, job_num).fill = _fill(GRAY if job_num % 2 == 0 else WHITE)
            for ci in range(2, 10):
                ws.cell(row, ci).fill = _fill(GRAY if job_num % 2 == 0 else WHITE)
            row += 1
            job_num += 1

    for ci, w in {1:6,2:28,3:30,4:20,5:50,6:12,7:12,8:18,9:30}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)


# ── Application Tracker ──────────────────────────────────────────────────────

def build_application_tracker(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    banner_text = (
        "APPLICATION TRACKER — Update after EVERY manual submission. "
        "No auto-apply. No auto-email. Manual only."
    )
    ws.cell(1,1,banner_text).fill = _fill(NAVY)
    ws.cell(1,1).font = _font(bold=True, color=WHITE, size=12)
    ws.cell(1,1).alignment = _align(h="center")
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 26

    hdrs = ["#","Company","Job Title","Applied Date","CV Used","Letter Language",
            "Status","Follow-up Date","Reply Received","Contact","Notes"]
    _header(ws, hdrs, TEAL, row=2)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(hdrs))}2"

    ws.add_data_validation(_dv(STATUS_DV, "G3:G1000"))
    ws.add_data_validation(_dv(CV_DV,     "E3:E1000"))

    for ci, w in {1:6,2:28,3:30,4:14,5:14,6:16,7:18,8:16,9:18,10:22,11:40}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    from datetime import date
    today = date.today().strftime("%Y-%m-%d")
    os.makedirs(_JOB_SEARCH, exist_ok=True)
    os.makedirs(_START_HERE, exist_ok=True)

    print("Building job_tracker.xlsx...")
    p = os.path.join(_JOB_SEARCH, "job_tracker.xlsx")
    build_job_tracker(p)
    print(f"  Written: {p} ({os.path.getsize(p):,} bytes)")

    print("Building 1_DAILY_JOB_PLAN.xlsx...")
    p = os.path.join(_START_HERE, "1_DAILY_JOB_PLAN.xlsx")
    build_daily_job_plan(p, today)
    print(f"  Written: {p} ({os.path.getsize(p):,} bytes)")

    print("Building 2_APPLICATION_TRACKER.xlsx...")
    p = os.path.join(_START_HERE, "2_APPLICATION_TRACKER.xlsx")
    build_application_tracker(p)
    print(f"  Written: {p} ({os.path.getsize(p):,} bytes)")


if __name__ == "__main__":
    main()
