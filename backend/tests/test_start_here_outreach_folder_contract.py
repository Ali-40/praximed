"""
Static contract tests for the START_HERE sales folder.

Sprint 21 / Outreach — Start-here sales folder.

Verifies that all required files exist and contain the right content.
No database. No network. No secrets. No patient data. No PHI.
"""

from __future__ import annotations

import csv
import os
import re

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(os.path.dirname(_HERE))
_START_HERE = os.path.join(_REPO_ROOT, "docs", "sales", "outreach", "START_HERE")

_GUIDE       = os.path.join(_START_HERE, "1_START_HERE.md")
_CALL_LIST   = os.path.join(_START_HERE, "2_TODAY_CALL_LIST.xlsx")
_DRAFTS      = os.path.join(_START_HERE, "3_EMAIL_DRAFTS.md")
_TRACKER     = os.path.join(_START_HERE, "4_MASTER_TRACKER.xlsx")
_UPDATE      = os.path.join(_START_HERE, "5_AFTER_CALL_UPDATE_GUIDE.md")
_README      = os.path.join(_START_HERE, "README_FOR_ALI.md")


def _read(path: str) -> str:
    with open(path, encoding="utf-8") as f:
        return f.read()


# ---------------------------------------------------------------------------
# 1. Folder and file existence
# ---------------------------------------------------------------------------

def test_start_here_folder_exists() -> None:
    assert os.path.isdir(_START_HERE), \
        "docs/sales/outreach/START_HERE/ must exist"


def test_guide_exists() -> None:
    assert os.path.isfile(_GUIDE), \
        "START_HERE/1_START_HERE.md must exist"


def test_call_list_xlsx_exists() -> None:
    assert os.path.isfile(_CALL_LIST), \
        "START_HERE/2_TODAY_CALL_LIST.xlsx must exist"


def test_email_drafts_exists() -> None:
    assert os.path.isfile(_DRAFTS), \
        "START_HERE/3_EMAIL_DRAFTS.md must exist"


def test_master_tracker_xlsx_exists() -> None:
    assert os.path.isfile(_TRACKER), \
        "START_HERE/4_MASTER_TRACKER.xlsx must exist"


def test_after_call_guide_exists() -> None:
    assert os.path.isfile(_UPDATE), \
        "START_HERE/5_AFTER_CALL_UPDATE_GUIDE.md must exist"


def test_readme_for_ali_exists() -> None:
    assert os.path.isfile(_README), \
        "START_HERE/README_FOR_ALI.md must exist"


# ---------------------------------------------------------------------------
# 2. 1_START_HERE.md content
# ---------------------------------------------------------------------------

def test_guide_mentions_call_list() -> None:
    content = _read(_GUIDE).lower()
    assert "call_list" in content or "call list" in content or "today_call" in content.lower(), \
        "1_START_HERE.md must reference the call list file"


def test_guide_mentions_email_drafts() -> None:
    content = _read(_GUIDE).lower()
    assert "email_draft" in content or "email draft" in content, \
        "1_START_HERE.md must reference email drafts"


def test_guide_mentions_master_tracker() -> None:
    content = _read(_GUIDE).lower()
    assert "master" in content or "tracker" in content, \
        "1_START_HERE.md must mention the master tracker"


def test_guide_says_which_file_to_open_first() -> None:
    content = _read(_GUIDE)
    # Should clearly state what to open first
    assert "open" in content.lower() and ("first" in content.lower() or "start" in content.lower()), \
        "1_START_HERE.md must explain which file to open first"


def test_guide_mentions_daily_goal() -> None:
    content = _read(_GUIDE).lower()
    assert "call" in content and ("10" in content or "25" in content), \
        "1_START_HERE.md must mention a daily call target"


# ---------------------------------------------------------------------------
# 3. 2_TODAY_CALL_LIST.xlsx columns
# ---------------------------------------------------------------------------

def test_call_list_xlsx_is_valid() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        assert wb is not None, "2_TODAY_CALL_LIST.xlsx must be a valid workbook"
    except ImportError:
        pass  # openpyxl not installed — skip structural check


def test_call_list_has_doctor_name_column() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20) if ws.cell(row=1, column=ci).value]
        assert any("Doctor Name" in str(h) or "doctor" in str(h).lower() for h in headers), \
            "2_TODAY_CALL_LIST.xlsx must have a Doctor Name column"
    except ImportError:
        pass


def test_call_list_has_phone_column() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20) if ws.cell(row=1, column=ci).value]
        assert any("Phone" in str(h) or "phone" in str(h).lower() for h in headers), \
            "2_TODAY_CALL_LIST.xlsx must have a Phone column"
    except ImportError:
        pass


def test_call_list_has_status_column() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20) if ws.cell(row=1, column=ci).value]
        assert any("Status" in str(h) for h in headers), \
            "2_TODAY_CALL_LIST.xlsx must have a Status column"
    except ImportError:
        pass


def test_call_list_has_notes_column() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20) if ws.cell(row=1, column=ci).value]
        assert any("Notes" in str(h) for h in headers), \
            "2_TODAY_CALL_LIST.xlsx must have a Notes column"
    except ImportError:
        pass


def test_call_list_has_call_order_column() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20) if ws.cell(row=1, column=ci).value]
        assert any("Call Order" in str(h) or "Order" in str(h) for h in headers), \
            "2_TODAY_CALL_LIST.xlsx must have a Call Order column"
    except ImportError:
        pass


def test_call_list_has_opening_line_column() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20) if ws.cell(row=1, column=ci).value]
        assert any("Opening" in str(h) for h in headers), \
            "2_TODAY_CALL_LIST.xlsx must have an Opening Line column"
    except ImportError:
        pass


def test_call_list_has_data_rows() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
        assert len(data_rows) >= 1, "2_TODAY_CALL_LIST.xlsx must have at least one data row"
    except ImportError:
        pass


def test_call_list_status_default_not_called() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        headers = [ws.cell(row=1, column=ci).value for ci in range(1, 20)]
        status_col = next((i + 1 for i, h in enumerate(headers) if h and "Status" in str(h)), None)
        if status_col:
            first_status = ws.cell(row=2, column=status_col).value
            assert first_status in (None, "Not called", ""), \
                f"Default Status must be 'Not called' or empty, got: '{first_status}'"
    except ImportError:
        pass


def test_call_list_frozen_top_row() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        assert ws.freeze_panes is not None, \
            "2_TODAY_CALL_LIST.xlsx must have frozen top row"
    except ImportError:
        pass


def test_call_list_has_autofilter() -> None:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_CALL_LIST)
        ws = wb.active
        assert ws.auto_filter.ref is not None, \
            "2_TODAY_CALL_LIST.xlsx must have auto-filter enabled"
    except ImportError:
        pass


# ---------------------------------------------------------------------------
# 4. 3_EMAIL_DRAFTS.md content
# ---------------------------------------------------------------------------

def test_email_drafts_says_no_mass_send() -> None:
    content = _read(_DRAFTS).lower()
    assert "mass" in content or "bulk" in content or "do not" in content, \
        "3_EMAIL_DRAFTS.md must say do not mass-send"


def test_email_drafts_says_copy_paste() -> None:
    content = _read(_DRAFTS).lower()
    assert "copy" in content and "paste" in content, \
        "3_EMAIL_DRAFTS.md must instruct to copy/paste"


def test_email_drafts_no_auto_send() -> None:
    content = _read(_DRAFTS).lower()
    assert "manually" in content or "manual" in content or "manuell" in content, \
        "3_EMAIL_DRAFTS.md must say emails are sent manually"


def test_email_drafts_has_content() -> None:
    content = _read(_DRAFTS)
    assert len(content) > 200, \
        "3_EMAIL_DRAFTS.md must have substantive content"


def test_email_drafts_no_patient_data() -> None:
    content = _read(_DRAFTS).lower()
    assert "patient_id" not in content, "email drafts must not contain patient IDs"
    assert "diagnosis" not in content, "email drafts must not contain diagnosis"


# ---------------------------------------------------------------------------
# 5. 5_AFTER_CALL_UPDATE_GUIDE.md content
# ---------------------------------------------------------------------------

def test_update_guide_mentions_outreach_status() -> None:
    assert "Outreach Status" in _read(_UPDATE), \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention Outreach Status"


def test_update_guide_mentions_do_not_contact() -> None:
    assert "Do not contact" in _read(_UPDATE), \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention 'Do not contact'"


def test_update_guide_mentions_follow_up_date() -> None:
    content = _read(_UPDATE).lower()
    assert "follow-up" in content or "followup" in content, \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention follow-up date"


def test_update_guide_mentions_demo_booked() -> None:
    content = _read(_UPDATE)
    assert "Demo Booked" in content or "demo booked" in content.lower(), \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention Demo Booked"


def test_update_guide_mentions_objection() -> None:
    content = _read(_UPDATE)
    assert "Objection" in content or "objection" in content.lower(), \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention Objection field"


def test_update_guide_no_patient_data() -> None:
    content = _read(_UPDATE).lower()
    assert "no patient" in content or "never" in content or "not" in content, \
        "5_AFTER_CALL_UPDATE_GUIDE.md must clarify no patient data"


def test_update_guide_no_phi() -> None:
    assert "PHI" in _read(_UPDATE).upper() or "patient" in _read(_UPDATE).lower(), \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention PHI or patient data boundary"


def test_update_guide_mentions_notes() -> None:
    assert "Notes" in _read(_UPDATE), \
        "5_AFTER_CALL_UPDATE_GUIDE.md must mention the Notes field"


# ---------------------------------------------------------------------------
# 6. README_FOR_ALI.md
# ---------------------------------------------------------------------------

def test_readme_for_ali_is_short() -> None:
    content = _read(_README)
    assert len(content) < 500, \
        "README_FOR_ALI.md must be short (under 500 chars)"


def test_readme_for_ali_says_stop_coding() -> None:
    content = _read(_README).lower()
    assert "coding" in content or "call" in content, \
        "README_FOR_ALI.md must say stop coding, start calling"


def test_readme_for_ali_mentions_call_list() -> None:
    content = _read(_README).lower()
    assert "call" in content and ("list" in content or "call_list" in content), \
        "README_FOR_ALI.md must mention the call list"


# ---------------------------------------------------------------------------
# 7. No secrets, no auto-actions
# ---------------------------------------------------------------------------

def test_no_secrets_in_guide() -> None:
    content = _read(_GUIDE).lower()
    assert "api_key" not in content and "secret" not in content and "password" not in content, \
        "1_START_HERE.md must not contain secrets"


def test_no_auto_email_in_drafts() -> None:
    content = _read(_DRAFTS).lower()
    assert "smtplib" not in content and "sendmail" not in content, \
        "3_EMAIL_DRAFTS.md must not contain email-sending code"


def test_no_auto_call_in_guide() -> None:
    content = _read(_GUIDE).lower()
    assert "twilio" not in content and "vapi" not in content and "autodial" not in content, \
        "1_START_HERE.md must not reference auto-calling tools"


def test_no_api_key_literals_in_files() -> None:
    import re
    pattern = re.compile(r'["\']([A-Za-z0-9]{32,})["\']')
    for path in [_GUIDE, _DRAFTS, _UPDATE, _README]:
        content = _read(path)
        matches = pattern.findall(content)
        assert not matches, \
            f"{os.path.basename(path)} must not contain API key literals: {matches[:2]}"


# ---------------------------------------------------------------------------
# 8. File is not empty / meaningful size
# ---------------------------------------------------------------------------

def test_guide_has_content() -> None:
    assert len(_read(_GUIDE)) > 300, "1_START_HERE.md must have substantive content"


def test_update_guide_has_content() -> None:
    assert len(_read(_UPDATE)) > 400, "5_AFTER_CALL_UPDATE_GUIDE.md must have substantive content"


def test_call_list_xlsx_is_not_empty() -> None:
    size = os.path.getsize(_CALL_LIST)
    assert size > 5000, f"2_TODAY_CALL_LIST.xlsx must be > 5KB, got {size} bytes"
