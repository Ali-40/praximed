"""
Static contract tests for the Ali-ready Outreach Command Center.

Sprint 21 / Outreach — 50-call command center and professional sender pack.

Verifies file existence, workbook structure (10 sheets), 50-row counts,
Sender Identity content, Weekly Plan content, Alex/PraxisMed branding,
column presence, and safety rules.
No database. No network. No secrets. No patient data. No PHI.
"""

from __future__ import annotations

import os
import re

_HERE      = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(os.path.dirname(_HERE))
_FOLDER    = os.path.join(_REPO_ROOT, "docs", "sales", "outreach", "START_HERE_FOR_ALI")

_README    = os.path.join(_FOLDER, "0_READ_ME_FIRST.md")
_CMD_CTR   = os.path.join(_FOLDER, "OUTREACH_COMMAND_CENTER.xlsx")
_MON_CALLS = os.path.join(_FOLDER, "1_MONDAY_CALLS.xlsx")
_SUN_EMAIL = os.path.join(_FOLDER, "2_SUNDAY_EMAILS.xlsx")
_MASTER    = os.path.join(_FOLDER, "3_MASTER_TRACKER.xlsx")
_CALL_SCPT = os.path.join(_FOLDER, "4_CALL_SCRIPT_PRINTABLE.md")
_EMAIL_TPL = os.path.join(_FOLDER, "5_EMAIL_TEMPLATES_PRINTABLE.md")


def _read(path: str) -> str:
    with open(path, encoding="utf-8") as f:
        return f.read()


def _wb(path: str):
    import openpyxl
    return openpyxl.load_workbook(path)


def _sheet_headers(wb, sheet_name: str, header_row: int = 1) -> list[str]:
    ws = wb[sheet_name]
    return [
        str(ws.cell(row=header_row, column=ci).value or "")
        for ci in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=ci).value
    ]


# ---------------------------------------------------------------------------
# 1. Folder and file existence
# ---------------------------------------------------------------------------

def test_start_here_for_ali_folder_exists() -> None:
    assert os.path.isdir(_FOLDER), "docs/sales/outreach/START_HERE_FOR_ALI/ must exist"


def test_readme_first_exists() -> None:
    assert os.path.isfile(_README), "0_READ_ME_FIRST.md must exist"


def test_command_center_xlsx_exists() -> None:
    assert os.path.isfile(_CMD_CTR), "OUTREACH_COMMAND_CENTER.xlsx must exist"


def test_monday_calls_xlsx_exists() -> None:
    assert os.path.isfile(_MON_CALLS), "1_MONDAY_CALLS.xlsx must exist"


def test_sunday_emails_xlsx_exists() -> None:
    assert os.path.isfile(_SUN_EMAIL), "2_SUNDAY_EMAILS.xlsx must exist"


def test_master_tracker_xlsx_exists() -> None:
    assert os.path.isfile(_MASTER), "3_MASTER_TRACKER.xlsx must exist"


def test_call_script_printable_exists() -> None:
    assert os.path.isfile(_CALL_SCPT), "4_CALL_SCRIPT_PRINTABLE.md must exist"


def test_email_templates_printable_exists() -> None:
    assert os.path.isfile(_EMAIL_TPL), "5_EMAIL_TEMPLATES_PRINTABLE.md must exist"


# ---------------------------------------------------------------------------
# 2. OUTREACH_COMMAND_CENTER.xlsx — sheet names
# ---------------------------------------------------------------------------

def test_command_center_has_start_here_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "START HERE" in wb.sheetnames, "Workbook must have 'START HERE' sheet"


def test_command_center_has_monday_calls_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Monday Calls" in wb.sheetnames, "Workbook must have 'Monday Calls' sheet"


def test_command_center_has_sunday_emails_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Sunday Emails" in wb.sheetnames, "Workbook must have 'Sunday Emails' sheet"


def test_command_center_has_follow_ups_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Follow Ups" in wb.sheetnames, "Workbook must have 'Follow Ups' sheet"


def test_command_center_has_master_leads_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Master Leads" in wb.sheetnames, "Workbook must have 'Master Leads' sheet"


def test_command_center_has_call_script_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Call Script" in wb.sheetnames, "Workbook must have 'Call Script' sheet"


def test_command_center_has_email_templates_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Email Templates" in wb.sheetnames, "Workbook must have 'Email Templates' sheet"


def test_command_center_has_stats_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Stats" in wb.sheetnames, "Workbook must have 'Stats' sheet"


def test_command_center_has_sender_identity_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Sender Identity" in wb.sheetnames, "Workbook must have 'Sender Identity' sheet"


def test_command_center_has_weekly_plan_sheet() -> None:
    wb = _wb(_CMD_CTR)
    assert "Weekly Plan" in wb.sheetnames, "Workbook must have 'Weekly Plan' sheet"


def test_command_center_has_ten_sheets() -> None:
    wb = _wb(_CMD_CTR)
    assert len(wb.sheetnames) >= 10, \
        f"Workbook must have at least 10 sheets, got {len(wb.sheetnames)}"


# ---------------------------------------------------------------------------
# 3. Monday Calls columns
# ---------------------------------------------------------------------------

def _call_headers(from_file=_CMD_CTR):
    wb = _wb(from_file)
    sheet_name = "Monday Calls" if "Monday Calls" in wb.sheetnames else wb.sheetnames[0]
    # row 1 = banner, headers in row 2
    return _sheet_headers(wb, sheet_name, header_row=2)


def test_monday_calls_has_doctor_name() -> None:
    assert any("Doctor Name" in h for h in _call_headers()), \
        "Monday Calls must have 'Doctor Name' column"


def test_monday_calls_has_specialty() -> None:
    hdrs = _call_headers()
    assert any("Specialty" in h for h in hdrs), \
        "Monday Calls must have 'Specialty' column"


def test_monday_calls_has_phone() -> None:
    assert any("Phone" in h for h in _call_headers()), \
        "Monday Calls must have 'Phone' column"


def test_monday_calls_has_outreach_status() -> None:
    assert any("Outreach Status" in h or "Status" in h for h in _call_headers()), \
        "Monday Calls must have 'Outreach Status' column"


def test_monday_calls_has_call_result() -> None:
    assert any("Call Result" in h or "Result" in h for h in _call_headers()), \
        "Monday Calls must have 'Call Result' column"


def test_monday_calls_has_follow_up_date() -> None:
    hdrs = _call_headers()
    assert any("Follow" in h for h in hdrs), \
        "Monday Calls must have 'Follow-up Date' column"


def test_monday_calls_has_demo_offered() -> None:
    assert any("Demo Offered" in h for h in _call_headers()), \
        "Monday Calls must have 'Demo Offered' column"


def test_monday_calls_has_demo_booked() -> None:
    assert any("Demo Booked" in h for h in _call_headers()), \
        "Monday Calls must have 'Demo Booked' column"


def test_monday_calls_has_notes() -> None:
    assert any("Notes" in h for h in _call_headers()), \
        "Monday Calls must have 'Notes' column"


def test_monday_calls_has_opening_line() -> None:
    assert any("Opening" in h for h in _call_headers()), \
        "Monday Calls must have 'Opening Line' column"


def test_monday_calls_has_call_order() -> None:
    assert any("Order" in h for h in _call_headers()), \
        "Monday Calls must have 'Call Order' column"


def test_monday_calls_has_data_rows() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Monday Calls"]
    data = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
    assert len(data) >= 1, "Monday Calls must have at least one data row"


def test_monday_calls_has_50_rows() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Monday Calls"]
    # row 1 = banner, row 2 = header, data starts row 3
    data = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c for c in r)]
    assert len(data) >= 50, \
        f"Monday Calls must have at least 50 data rows, got {len(data)}"


def test_monday_calls_call_order_goes_to_50() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Monday Calls"]
    orders = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[0] is not None:
            try:
                orders.append(int(r[0]))
            except (TypeError, ValueError):
                pass
    assert 50 in orders, "Monday Calls must include Call Order 50"


def test_monday_calls_frozen_header() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Monday Calls"]
    assert ws.freeze_panes is not None, "Monday Calls must have frozen header row"


def test_monday_calls_has_autofilter() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Monday Calls"]
    assert ws.auto_filter.ref is not None, "Monday Calls must have auto-filter"


# ---------------------------------------------------------------------------
# 4. Sunday Emails columns
# ---------------------------------------------------------------------------

def _email_headers(from_file=_CMD_CTR):
    wb = _wb(from_file)
    sheet_name = "Sunday Emails" if "Sunday Emails" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    # Header is on row 2 (row 1 is warning banner)
    return [
        str(ws.cell(row=2, column=ci).value or "")
        for ci in range(1, ws.max_column + 1)
        if ws.cell(row=2, column=ci).value
    ]


def test_sunday_emails_has_doctor_name() -> None:
    assert any("Doctor Name" in h for h in _email_headers()), \
        "Sunday Emails must have 'Doctor Name' column"


def test_sunday_emails_has_subject() -> None:
    assert any("Subject" in h for h in _email_headers()), \
        "Sunday Emails must have 'Subject' column"


def test_sunday_emails_has_email_draft() -> None:
    assert any("Draft" in h or "Email Draft" in h for h in _email_headers()), \
        "Sunday Emails must have 'Email Draft' column"


def test_sunday_emails_has_manual_send_status() -> None:
    hdrs = _email_headers()
    assert any("Manual" in h or "Send" in h or "Status" in h for h in hdrs), \
        "Sunday Emails must have 'Manual Send Status' column"


def test_sunday_emails_has_reply_status() -> None:
    assert any("Reply" in h for h in _email_headers()), \
        "Sunday Emails must have 'Reply Status' column"


def test_sunday_emails_has_follow_up_date() -> None:
    assert any("Follow" in h for h in _email_headers()), \
        "Sunday Emails must have 'Follow-up Date' column"


def test_sunday_emails_has_data_rows() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Sunday Emails"]
    data = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c for c in r)]
    assert len(data) >= 1, "Sunday Emails must have at least one data row"


def test_sunday_emails_has_50_rows() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Sunday Emails"]
    # row 1 = banner, row 2 = header, data starts row 3
    data = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c for c in r)]
    assert len(data) >= 50, \
        f"Sunday Emails must have at least 50 data rows, got {len(data)}"


def test_sunday_emails_email_order_goes_to_50() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Sunday Emails"]
    orders = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r[0] is not None:
            try:
                orders.append(int(r[0]))
            except (TypeError, ValueError):
                pass
    assert 50 in orders, "Sunday Emails must include Email Order 50"


def test_sunday_emails_warning_banner_present() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Sunday Emails"]
    first_row_value = ws.cell(row=1, column=1).value or ""
    assert "MANUAL" in str(first_row_value).upper() or "mass" in str(first_row_value).lower(), \
        "Sunday Emails must have a warning banner in row 1"


# ---------------------------------------------------------------------------
# 5. Standalone files — basic validity
# ---------------------------------------------------------------------------

def test_monday_calls_standalone_has_data() -> None:
    wb = _wb(_MON_CALLS)
    ws = wb.active
    data = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
    assert len(data) >= 1, "1_MONDAY_CALLS.xlsx must have at least one data row"


def test_sunday_emails_standalone_has_data() -> None:
    wb = _wb(_SUN_EMAIL)
    ws = wb.active
    data = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c for c in r)]
    assert len(data) >= 1, "2_SUNDAY_EMAILS.xlsx must have at least one data row"


def test_master_tracker_standalone_has_many_rows() -> None:
    wb = _wb(_MASTER)
    ws = wb.active
    data = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r)]
    assert len(data) >= 100, \
        f"3_MASTER_TRACKER.xlsx must have 100+ rows, got {len(data)}"


def test_master_tracker_has_doctor_name() -> None:
    wb = _wb(_MASTER)
    hdrs = _sheet_headers(wb, wb.active.title)
    assert any("Doctor Name" in h for h in hdrs), \
        "3_MASTER_TRACKER.xlsx must have 'Doctor Name' column"


def test_command_center_is_reasonable_size() -> None:
    size = os.path.getsize(_CMD_CTR)
    assert size > 50_000, f"OUTREACH_COMMAND_CENTER.xlsx must be >50KB, got {size} bytes"


# ---------------------------------------------------------------------------
# 6. README content
# ---------------------------------------------------------------------------

def test_readme_mentions_saturday_prepare() -> None:
    content = _read(_README).lower()
    assert "saturday" in content or "samstag" in content, \
        "0_READ_ME_FIRST.md must mention Saturday preparation"


def test_readme_mentions_sunday_emails() -> None:
    content = _read(_README).lower()
    assert "sunday" in content or "sonntag" in content, \
        "0_READ_ME_FIRST.md must mention Sunday emails"


def test_readme_mentions_monday_calls() -> None:
    content = _read(_README).lower()
    assert "monday" in content or "montag" in content, \
        "0_READ_ME_FIRST.md must mention Monday calls"


def test_readme_mentions_no_raw_csv() -> None:
    content = _read(_README).lower()
    assert "csv" in content and ("not" in content or "do not" in content or "unless" in content), \
        "0_READ_ME_FIRST.md must say do not use raw CSV files unless needed"


def test_readme_mentions_do_not_mass_send() -> None:
    content = _read(_README).lower()
    assert "mass" in content or "bulk" in content or "auto" in content, \
        "0_READ_ME_FIRST.md must mention no mass-sending"


def test_readme_mentions_manual_only() -> None:
    content = _read(_README).lower()
    assert "manual" in content or "manuell" in content, \
        "0_READ_ME_FIRST.md must say emails are sent manually"


# ---------------------------------------------------------------------------
# 7. Call script content
# ---------------------------------------------------------------------------

def test_call_script_has_latido_positioning() -> None:
    content = _read(_CALL_SCPT)
    assert "LATIDO" in content, \
        "4_CALL_SCRIPT_PRINTABLE.md must include LATIDO-compatible positioning"


def test_call_script_has_no_diagnosis() -> None:
    content = _read(_CALL_SCPT).lower()
    assert "keine diagnosen" in content or "no diagnos" in content, \
        "4_CALL_SCRIPT_PRINTABLE.md must explicitly state no diagnosis claims"


def test_call_script_has_no_medical_advice() -> None:
    content = _read(_CALL_SCPT).lower()
    assert "medizinische beratung" in content or "medical advice" in content, \
        "4_CALL_SCRIPT_PRINTABLE.md must explicitly state no medical advice"


def test_call_script_has_no_auto_confirm() -> None:
    content = _read(_CALL_SCPT).lower()
    assert "bestätigt keine termine" in content or "auto-confirm" in content, \
        "4_CALL_SCRIPT_PRINTABLE.md must state no automatic appointment confirmation"


def test_call_script_has_staff_in_control() -> None:
    content = _read(_CALL_SCPT).lower()
    assert "kontrolle" in content or "in control" in content, \
        "4_CALL_SCRIPT_PRINTABLE.md must state staff stays in control"


def test_call_script_has_do_not_contact_instruction() -> None:
    content = _read(_CALL_SCPT)
    assert "Do not contact" in content or "do not contact" in content.lower(), \
        "4_CALL_SCRIPT_PRINTABLE.md must include 'Do not contact' instruction"


# ---------------------------------------------------------------------------
# 8. Email templates content
# ---------------------------------------------------------------------------

def test_email_templates_says_no_mass_send() -> None:
    content = _read(_EMAIL_TPL).lower()
    assert "mass" in content or "bulk" in content or "no mass" in content, \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must say no mass-send"


def test_email_templates_says_manual_review() -> None:
    content = _read(_EMAIL_TPL).lower()
    assert "manual" in content and "review" in content, \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must say manual review only"


def test_email_templates_has_latido_positioning() -> None:
    assert "LATIDO" in _read(_EMAIL_TPL), \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must include LATIDO note"


def test_email_templates_has_no_diagnosis_claim() -> None:
    content = _read(_EMAIL_TPL).lower()
    assert "diagnos" in content, \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must address diagnosis boundary"


def test_email_templates_has_no_medical_advice_claim() -> None:
    content = _read(_EMAIL_TPL).lower()
    assert "medizinische beratung" in content or "medical advice" in content, \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must address medical advice boundary"


def test_email_templates_has_do_not_contact() -> None:
    content = _read(_EMAIL_TPL)
    assert "Do not contact" in content or "do not contact" in content.lower(), \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must include 'Do not contact' rule"


def test_email_templates_has_no_patient_data() -> None:
    content = _read(_EMAIL_TPL).lower()
    assert "patient" in content or "phi" in content, \
        "5_EMAIL_TEMPLATES_PRINTABLE.md must mention no patient data / PHI"


# ---------------------------------------------------------------------------
# 9. Safety: no secrets, no auto-actions
# ---------------------------------------------------------------------------

def _all_md() -> str:
    return "\n".join(_read(p) for p in [_README, _CALL_SCPT, _EMAIL_TPL])


def test_no_database_url() -> None:
    assert "DATABASE_URL" not in _all_md(), "Markdown files must not contain DATABASE_URL"


def test_no_jwt_secret() -> None:
    content = _all_md().lower()
    assert "jwt_secret" not in content and "secret_key" not in content, \
        "Markdown files must not contain JWT secret"


def test_no_vapi_api_key() -> None:
    content = _all_md().lower()
    assert "vapi_api_key" not in content and "vapi_secret" not in content, \
        "Markdown files must not contain Vapi API key"


def test_no_webhook_secret() -> None:
    assert "webhook_secret" not in _all_md().lower(), \
        "Markdown files must not contain webhook secret"


def test_no_api_key_literals() -> None:
    pattern = re.compile(r'["\']([A-Za-z0-9]{32,})["\']')
    for path in [_README, _CALL_SCPT, _EMAIL_TPL]:
        matches = pattern.findall(_read(path))
        assert not matches, \
            f"{os.path.basename(path)} must not contain hardcoded API keys: {matches[:2]}"


def test_no_auto_email_code() -> None:
    content = _all_md().lower()
    assert "smtplib" not in content and "sendmail" not in content, \
        "Files must not contain email-sending code"


def test_no_auto_call_code() -> None:
    content = _all_md().lower()
    assert "twilio" not in content and "vapi" not in content, \
        "Files must not reference auto-calling tools"


def test_no_patient_data_in_md() -> None:
    content = _all_md().lower()
    assert "patient_id" not in content, "Markdown files must not contain patient_id"
    assert "diagnosis" not in content or "no diagnos" in content or "keine diagnosen" in content, \
        "Markdown files must not make diagnosis claims"


# ---------------------------------------------------------------------------
# 10. File sizes reasonable
# ---------------------------------------------------------------------------

def test_command_center_not_empty() -> None:
    assert os.path.getsize(_CMD_CTR) > 50_000, "OUTREACH_COMMAND_CENTER.xlsx must be >50KB"


def test_monday_calls_not_empty() -> None:
    assert os.path.getsize(_MON_CALLS) > 5_000, "1_MONDAY_CALLS.xlsx must be >5KB"


def test_sunday_emails_not_empty() -> None:
    assert os.path.getsize(_SUN_EMAIL) > 5_000, "2_SUNDAY_EMAILS.xlsx must be >5KB"


def test_master_tracker_not_empty() -> None:
    assert os.path.getsize(_MASTER) > 50_000, "3_MASTER_TRACKER.xlsx must be >50KB"


def test_call_script_has_content() -> None:
    assert len(_read(_CALL_SCPT)) > 500, "4_CALL_SCRIPT_PRINTABLE.md must have substantive content"


def test_email_templates_has_content() -> None:
    assert len(_read(_EMAIL_TPL)) > 400, "5_EMAIL_TEMPLATES_PRINTABLE.md must have substantive content"


# ---------------------------------------------------------------------------
# 11. Sender Identity sheet
# ---------------------------------------------------------------------------

def _sender_identity_text() -> str:
    wb = _wb(_CMD_CTR)
    ws = wb["Sender Identity"]
    parts = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell:
                parts.append(str(cell))
    return " ".join(parts)


def test_sender_identity_has_alex_name() -> None:
    text = _sender_identity_text().lower()
    assert "alex" in text, "Sender Identity must mention 'Alex'"


def test_sender_identity_has_praxismed_team() -> None:
    text = _sender_identity_text()
    assert "PraxisMed" in text, "Sender Identity must mention 'PraxisMed'"


def test_sender_identity_has_ali_abdeltawab() -> None:
    text = _sender_identity_text()
    assert "Ali Abdeltawab" in text or "ali abdeltawab" in text.lower(), \
        "Sender Identity must include Ali Abdeltawab's real name"


def test_sender_identity_not_fake_person() -> None:
    text = _sender_identity_text().lower()
    assert "fake" not in text or "not a fake" in text or "not fake" in text, \
        "Sender Identity must not claim a fake person — it must clarify this is an alias"


def test_sender_identity_has_transparency_note() -> None:
    text = _sender_identity_text().lower()
    assert "explain" in text or "asked" in text or "vollständig" in text or "transparent" in text, \
        "Sender Identity must include a transparency note about the alias"


def test_sender_identity_has_gmail_setup() -> None:
    text = _sender_identity_text().lower()
    assert "gmail" in text or "settings" in text or "einstellungen" in text, \
        "Sender Identity must include Gmail setup instructions"


# ---------------------------------------------------------------------------
# 12. Weekly Plan sheet
# ---------------------------------------------------------------------------

def _weekly_plan_text() -> str:
    wb = _wb(_CMD_CTR)
    ws = wb["Weekly Plan"]
    parts = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell:
                parts.append(str(cell))
    return " ".join(parts)


def test_weekly_plan_has_saturday() -> None:
    text = _weekly_plan_text().lower()
    assert "samstag" in text or "saturday" in text, \
        "Weekly Plan must include Saturday"


def test_weekly_plan_has_sunday() -> None:
    text = _weekly_plan_text().lower()
    assert "sonntag" in text or "sunday" in text, \
        "Weekly Plan must include Sunday"


def test_weekly_plan_has_monday() -> None:
    text = _weekly_plan_text().lower()
    assert "montag" in text or "monday" in text, \
        "Weekly Plan must include Monday"


def test_weekly_plan_has_50_call_target() -> None:
    text = _weekly_plan_text()
    assert "50" in text, "Weekly Plan must reference 50 (call target)"


def test_weekly_plan_no_auto_send() -> None:
    text = _weekly_plan_text().lower()
    assert "auto-send" not in text or "kein auto-send" in text or "no auto" in text, \
        "Weekly Plan must not promote auto-send"


# ---------------------------------------------------------------------------
# 13. Alex / PraxisMed Team branding in templates
# ---------------------------------------------------------------------------

def test_email_templates_has_alex_version() -> None:
    content = _read(_EMAIL_TPL)
    assert "Alex" in content, "Email templates must include Alex sender version"


def test_email_templates_has_ali_abdeltawab_version() -> None:
    content = _read(_EMAIL_TPL)
    assert "Ali Abdeltawab" in content, \
        "Email templates must include transparent Ali Abdeltawab version"


def test_email_templates_two_versions_labeled() -> None:
    content = _read(_EMAIL_TPL)
    assert ("Version A" in content or "VERSION A" in content) and \
           ("Version B" in content or "VERSION B" in content), \
        "Email templates must have Version A and Version B clearly labeled"


def test_email_templates_alex_not_fake() -> None:
    content = _read(_EMAIL_TPL).lower()
    assert "alias" in content or "communication" in content or "vollständig" in content or "explain" in content, \
        "Email templates must clarify Alex is an alias, not a fake identity"


def test_call_script_has_alex_wording() -> None:
    content = _read(_CALL_SCPT)
    assert "Alex" in content, "Call script must include Alex wording/explanation"


# ---------------------------------------------------------------------------
# 14. Stats target = 50
# ---------------------------------------------------------------------------

def test_stats_sheet_references_50() -> None:
    wb = _wb(_CMD_CTR)
    ws = wb["Stats"]
    vals = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                vals.append(str(cell))
    text = " ".join(vals)
    assert "50" in text, "Stats sheet must reference target of 50"


def test_standalone_monday_has_50_rows() -> None:
    wb = _wb(_MON_CALLS)
    ws = wb.active
    # row 1 = banner, row 2 = header, data from row 3
    data = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c for c in r)]
    assert len(data) >= 50, \
        f"1_MONDAY_CALLS.xlsx must have at least 50 data rows, got {len(data)}"


def test_standalone_sunday_has_50_rows() -> None:
    wb = _wb(_SUN_EMAIL)
    ws = wb.active
    data = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c for c in r)]
    assert len(data) >= 50, \
        f"2_SUNDAY_EMAILS.xlsx must have at least 50 data rows, got {len(data)}"
