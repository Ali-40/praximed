"""
Static contract tests for the Career Job Application Copilot.

Sprint 21 / Career — Job application copilot and profile refresh.

Verifies:
  - career folder structure exists
  - CV files exist (English + German)
  - LinkedIn, GitHub, template files exist
  - Job tracker exists
  - Copilot script exists and supports all CLI modes
  - Copilot does NOT auto-apply, auto-email, or automate LinkedIn
  - Old JKU MSc Statistics/Data Science is removed
  - TU Wien MSc Software Engineering (accepted, Oct 2026) is present
  - Vienna priority, full-time priority, project positioning
  - No secrets, no DATABASE_URL, no JWT, no API key literals

No database. No network. No real job applications. No PHI. No secrets.
"""

from __future__ import annotations

import importlib.util
import os
import re

_HERE      = os.path.dirname(os.path.abspath(__file__))
_REPO      = os.path.dirname(os.path.dirname(_HERE))
_CAREER    = os.path.join(_REPO, "career")
_SCRIPTS   = os.path.join(_REPO, "scripts", "career")
_TESTS     = _HERE

_CV_EN     = os.path.join(_CAREER, "cv", "Ali_Abdeltawab_CV_EN_ATS.md")
_CV_DE     = os.path.join(_CAREER, "cv", "Ali_Abdeltawab_CV_DE_ATS.md")
_LI_RWRITE = os.path.join(_CAREER, "linkedin", "LINKEDIN_PROFILE_REWRITE.md")
_LI_GUIDE  = os.path.join(_CAREER, "linkedin", "LINKEDIN_UPDATE_STEP_BY_STEP.md")
_GH_README = os.path.join(_CAREER, "github", "GITHUB_PROFILE_README.md")
_GH_PROJS  = os.path.join(_CAREER, "github", "PROJECT_DESCRIPTIONS_FOR_RECRUITERS.md")
_GH_REPOS  = os.path.join(_CAREER, "github", "REPO_SELECTION_RECOMMENDATIONS.md")
_TRACKER   = os.path.join(_CAREER, "job_search", "job_tracker.xlsx")
_COPILOT   = os.path.join(_SCRIPTS, "job_application_copilot.py")
_ML_EN     = os.path.join(_CAREER, "templates", "motivation_letter_en.md")
_ML_DE     = os.path.join(_CAREER, "templates", "motivation_letter_de.md")
_START     = os.path.join(_CAREER, "START_HERE_FOR_ALI")
_README    = os.path.join(_START, "0_READ_ME_FIRST.md")
_SUBGUIDE  = os.path.join(_START, "5_MANUAL_SUBMISSION_GUIDE.md")


def _read(path: str) -> str:
    with open(path, encoding="utf-8") as f:
        return f.read()


def _copilot_src() -> str:
    return _read(_COPILOT)


def _all_career_md() -> str:
    paths = [_CV_EN, _CV_DE, _LI_RWRITE, _LI_GUIDE, _GH_README, _GH_PROJS,
             _GH_REPOS, _ML_EN, _ML_DE, _README, _SUBGUIDE]
    return "\n".join(_read(p) for p in paths)


# ---------------------------------------------------------------------------
# 1. Folder and file existence
# ---------------------------------------------------------------------------

def test_career_folder_exists() -> None:
    assert os.path.isdir(_CAREER), "career/ must exist"


def test_career_cv_folder_exists() -> None:
    assert os.path.isdir(os.path.join(_CAREER, "cv")), "career/cv/ must exist"


def test_career_linkedin_folder_exists() -> None:
    assert os.path.isdir(os.path.join(_CAREER, "linkedin")), "career/linkedin/ must exist"


def test_career_github_folder_exists() -> None:
    assert os.path.isdir(os.path.join(_CAREER, "github")), "career/github/ must exist"


def test_career_job_search_folder_exists() -> None:
    assert os.path.isdir(os.path.join(_CAREER, "job_search")), "career/job_search/ must exist"


def test_career_templates_folder_exists() -> None:
    assert os.path.isdir(os.path.join(_CAREER, "templates")), "career/templates/ must exist"


def test_career_start_here_folder_exists() -> None:
    assert os.path.isdir(_START), "career/START_HERE_FOR_ALI/ must exist"


def test_english_cv_exists() -> None:
    assert os.path.isfile(_CV_EN), "career/cv/Ali_Abdeltawab_CV_EN_ATS.md must exist"


def test_german_cv_exists() -> None:
    assert os.path.isfile(_CV_DE), "career/cv/Ali_Abdeltawab_CV_DE_ATS.md must exist"


def test_linkedin_rewrite_exists() -> None:
    assert os.path.isfile(_LI_RWRITE), "career/linkedin/LINKEDIN_PROFILE_REWRITE.md must exist"


def test_linkedin_guide_exists() -> None:
    assert os.path.isfile(_LI_GUIDE), "career/linkedin/LINKEDIN_UPDATE_STEP_BY_STEP.md must exist"


def test_github_profile_readme_exists() -> None:
    assert os.path.isfile(_GH_README), "career/github/GITHUB_PROFILE_README.md must exist"


def test_github_project_descriptions_exist() -> None:
    assert os.path.isfile(_GH_PROJS), "career/github/PROJECT_DESCRIPTIONS_FOR_RECRUITERS.md must exist"


def test_github_repo_recommendations_exist() -> None:
    assert os.path.isfile(_GH_REPOS), "career/github/REPO_SELECTION_RECOMMENDATIONS.md must exist"


def test_job_tracker_exists() -> None:
    assert os.path.isfile(_TRACKER), "career/job_search/job_tracker.xlsx must exist"


def test_copilot_script_exists() -> None:
    assert os.path.isfile(_COPILOT), "scripts/career/job_application_copilot.py must exist"


def test_motivation_letter_en_exists() -> None:
    assert os.path.isfile(_ML_EN), "career/templates/motivation_letter_en.md must exist"


def test_motivation_letter_de_exists() -> None:
    assert os.path.isfile(_ML_DE), "career/templates/motivation_letter_de.md must exist"


def test_start_here_readme_exists() -> None:
    assert os.path.isfile(_README), "career/START_HERE_FOR_ALI/0_READ_ME_FIRST.md must exist"


def test_manual_submission_guide_exists() -> None:
    assert os.path.isfile(_SUBGUIDE), "career/START_HERE_FOR_ALI/5_MANUAL_SUBMISSION_GUIDE.md must exist"


def test_job_tracker_not_empty() -> None:
    assert os.path.getsize(_TRACKER) > 5_000, "job_tracker.xlsx must be >5KB"


# ---------------------------------------------------------------------------
# 2. Copilot script — CLI mode support
# ---------------------------------------------------------------------------

def test_copilot_supports_mode_init() -> None:
    src = _copilot_src()
    assert '"init"' in src or "'init'" in src, "Copilot must support --mode init"


def test_copilot_supports_mode_import_links() -> None:
    src = _copilot_src()
    assert "import-links" in src, "Copilot must support --mode import-links"


def test_copilot_supports_mode_score() -> None:
    src = _copilot_src()
    assert '"score"' in src or "'score'" in src, "Copilot must support --mode score"


def test_copilot_supports_mode_drafts() -> None:
    src = _copilot_src()
    assert '"drafts"' in src or "'drafts'" in src, "Copilot must support --mode drafts"


def test_copilot_supports_mode_daily_plan() -> None:
    src = _copilot_src()
    assert "daily-plan" in src, "Copilot must support --mode daily-plan"


def test_copilot_supports_mode_report() -> None:
    src = _copilot_src()
    assert '"report"' in src or "'report'" in src, "Copilot must support --mode report"


# ---------------------------------------------------------------------------
# 3. Safety: no auto-apply, no auto-submit, no LinkedIn automation
# ---------------------------------------------------------------------------

def test_no_auto_apply_in_copilot() -> None:
    src = _copilot_src().lower()
    assert "auto_apply" not in src and "autoapply" not in src, \
        "Copilot must not contain auto-apply code"


def test_no_selenium_or_playwright() -> None:
    src = _copilot_src().lower()
    assert "selenium" not in src, "Copilot must not use selenium (browser automation)"
    assert "playwright" not in src, "Copilot must not use playwright (browser automation)"


def test_no_smtp_in_copilot() -> None:
    src = _copilot_src().lower()
    assert "smtplib" not in src, "Copilot must not use smtplib (email sending)"
    assert "sendmail" not in src, "Copilot must not use sendmail"


def test_no_linkedin_api_scraping() -> None:
    src = _copilot_src().lower()
    assert "linkedin_api" not in src and "linkedin.com/api" not in src, \
        "Copilot must not scrape LinkedIn API"


def test_no_twilio_or_vapi_in_copilot() -> None:
    src = _copilot_src().lower()
    assert "twilio" not in src, "Copilot must not call Twilio (auto-calling)"


def test_no_requests_post_apply() -> None:
    src = _copilot_src()
    assert "requests.post" not in src or "apply" not in src.lower(), \
        "Copilot must not POST application requests automatically"


def test_no_auto_apply_in_submission_guide() -> None:
    content = _read(_SUBGUIDE).lower()
    assert "no auto-apply" in content or "no automation" in content or "auto-apply" in content, \
        "Manual submission guide must explicitly say no auto-apply / no automation"


def test_no_fake_identity_in_copilot() -> None:
    src = _copilot_src().lower()
    assert "fake_name" not in src and "fake_identity" not in src, \
        "Copilot must not create fake identities"


# ---------------------------------------------------------------------------
# 4. CV — old MSc removed, TU Wien added
# ---------------------------------------------------------------------------

def test_en_cv_no_jku_msc_statistics() -> None:
    content = _read(_CV_EN)
    assert "MSc Statistics" not in content and "MSc Statitics" not in content, \
        "EN CV must not contain old JKU MSc Statistics/Data Science"


def test_de_cv_no_jku_msc_statistics() -> None:
    content = _read(_CV_DE)
    assert "MSc Statitics" not in content, \
        "DE CV must not contain misspelled old JKU MSc (Statitics)"


def test_en_cv_has_tu_wien() -> None:
    content = _read(_CV_EN)
    assert "TU Wien" in content, "EN CV must include TU Wien MSc Software Engineering"


def test_de_cv_has_tu_wien() -> None:
    content = _read(_CV_DE)
    assert "TU Wien" in content, "DE CV must include TU Wien MSc Software Engineering"


def test_en_cv_tu_wien_not_blocking_work() -> None:
    content = _read(_CV_EN).lower()
    assert "alongside" in content or "parallel" in content or "neben" in content, \
        "EN CV must note TU Wien is pursued alongside work (not blocking employment)"


def test_en_cv_has_jku_bsc_ai() -> None:
    content = _read(_CV_EN)
    assert "BSc Artificial Intelligence" in content and "Johannes Kepler" in content, \
        "EN CV must include BSc Artificial Intelligence from JKU"


def test_de_cv_has_jku_bsc_ai() -> None:
    content = _read(_CV_DE)
    assert "BSc Artificial Intelligence" in content and "Johannes Kepler" in content, \
        "DE CV must include BSc Artificial Intelligence from JKU"


def test_en_cv_has_vienna_location() -> None:
    content = _read(_CV_EN)
    assert "Vienna" in content or "Wien" in content, \
        "EN CV must mention Vienna as location"


def test_de_cv_has_vienna_location() -> None:
    content = _read(_CV_DE)
    assert "Wien" in content, "DE CV must mention Wien as location"


def test_en_cv_has_github_link() -> None:
    content = _read(_CV_EN)
    assert "github.com/Ali-40" in content, "EN CV must include GitHub link"


def test_de_cv_has_github_link() -> None:
    content = _read(_CV_DE)
    assert "github.com/Ali-40" in content, "DE CV must include GitHub link"


def test_en_cv_has_linkedin_link() -> None:
    content = _read(_CV_EN)
    assert "linkedin.com" in content, "EN CV must include LinkedIn link"


def test_de_cv_has_linkedin_link() -> None:
    content = _read(_CV_DE)
    assert "linkedin.com" in content, "DE CV must include LinkedIn link"


def test_en_cv_has_hofer() -> None:
    content = _read(_CV_EN)
    assert "Hofer" in content, "EN CV must include Hofer KG experience"


def test_de_cv_has_hofer() -> None:
    content = _read(_CV_DE)
    assert "Hofer" in content, "DE CV must include Hofer KG experience"


def test_en_cv_has_praxismed() -> None:
    content = _read(_CV_EN)
    assert "PraxisMed" in content, "EN CV must include PraxisMed project"


def test_de_cv_has_praxismed() -> None:
    content = _read(_CV_DE)
    assert "PraxisMed" in content, "DE CV must include PraxisMed project"


def test_en_cv_has_forecasting_project() -> None:
    content = _read(_CV_EN)
    assert "Forecasting" in content or "forecasting" in content, \
        "EN CV must include Cloud Demand Forecasting project"


def test_en_cv_has_python_skill() -> None:
    content = _read(_CV_EN)
    assert "Python" in content, "EN CV must list Python as a skill"


def test_en_cv_has_sql_skill() -> None:
    content = _read(_CV_EN)
    assert "SQL" in content, "EN CV must list SQL as a skill"


def test_en_cv_german_b1() -> None:
    content = _read(_CV_EN)
    assert "B1" in content, "EN CV must show German B1 level"


def test_de_cv_german_b1() -> None:
    content = _read(_CV_DE)
    assert "B1" in content, "DE CV must show German B1 level"


# ---------------------------------------------------------------------------
# 5. Profile — full-time priority, Vienna priority, openness to internships
# ---------------------------------------------------------------------------

def test_copilot_profile_has_full_time_priority() -> None:
    src = _copilot_src()
    assert "Full-time" in src, "Copilot profile must mark Full-time as priority"


def test_copilot_profile_has_vienna_target() -> None:
    src = _copilot_src()
    assert "Vienna" in src, "Copilot profile must target Vienna as primary location"


def test_copilot_profile_open_to_internship() -> None:
    src = _copilot_src()
    assert "Internship" in src, "Copilot profile must include openness to internships"


def test_copilot_profile_has_praxismed() -> None:
    src = _copilot_src()
    assert "PraxisMed" in src, "Copilot profile must include PraxisMed project"


def test_copilot_profile_has_forecasting() -> None:
    src = _copilot_src()
    assert "Forecasting" in src or "forecasting" in src, \
        "Copilot profile must include Cloud Demand Forecasting project"


def test_copilot_profile_has_github() -> None:
    src = _copilot_src()
    assert "github.com/Ali-40" in src, "Copilot profile must include GitHub link"


def test_copilot_profile_has_linkedin() -> None:
    src = _copilot_src()
    assert "linkedin.com" in src, "Copilot profile must include LinkedIn link"


def test_copilot_no_jku_msc_statistics() -> None:
    src = _copilot_src()
    assert "MSc Statistics" not in src and "MSc Statitics" not in src, \
        "Copilot must not reference old JKU MSc Statistics/Data Science"


def test_copilot_has_tu_wien() -> None:
    src = _copilot_src()
    assert "TU Wien" in src, "Copilot profile must include TU Wien MSc Software Engineering"


# ---------------------------------------------------------------------------
# 6. LinkedIn rewrite content
# ---------------------------------------------------------------------------

def test_linkedin_removes_old_msc() -> None:
    content = _read(_LI_RWRITE)
    assert "REMOVE" in content or "remove" in content.lower(), \
        "LinkedIn rewrite must instruct removal of old MSc Statistics"


def test_linkedin_adds_tu_wien() -> None:
    content = _read(_LI_RWRITE)
    assert "TU Wien" in content, "LinkedIn rewrite must add TU Wien MSc Software Engineering"


def test_linkedin_has_praxismed_project() -> None:
    content = _read(_LI_RWRITE)
    assert "PraxisMed" in content, "LinkedIn rewrite must include PraxisMed project"


def test_linkedin_has_open_to_work() -> None:
    content = _read(_LI_RWRITE).lower()
    assert "open to work" in content, "LinkedIn rewrite must mention Open to Work settings"


def test_linkedin_has_github_link() -> None:
    content = _read(_LI_RWRITE)
    assert "github.com/Ali-40" in content, "LinkedIn rewrite must include GitHub link"


# ---------------------------------------------------------------------------
# 7. Job tracker structure
# ---------------------------------------------------------------------------

def test_job_tracker_has_jobs_sheet() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(_TRACKER)
    assert "Jobs" in wb.sheetnames, "job_tracker.xlsx must have 'Jobs' sheet"


def test_job_tracker_has_dashboard_sheet() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(_TRACKER)
    assert "Dashboard" in wb.sheetnames, "job_tracker.xlsx must have 'Dashboard' sheet"


def test_job_tracker_has_applications_sheet() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(_TRACKER)
    assert "Applications" in wb.sheetnames, "job_tracker.xlsx must have 'Applications' sheet"


def test_job_tracker_has_keywords_sheet() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(_TRACKER)
    assert "Keywords" in wb.sheetnames, "job_tracker.xlsx must have 'Keywords' sheet"


def test_job_tracker_jobs_has_fit_score_column() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(_TRACKER)
    ws = wb["Jobs"]
    hdrs = [ws.cell(1, ci).value for ci in range(1, ws.max_column + 1)]
    assert any("Fit" in str(h) for h in hdrs), \
        "Jobs sheet must have a 'Fit Score' column"


def test_job_tracker_jobs_has_status_column() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(_TRACKER)
    ws = wb["Jobs"]
    hdrs = [ws.cell(1, ci).value for ci in range(1, ws.max_column + 1)]
    assert any("Status" in str(h) for h in hdrs), \
        "Jobs sheet must have an 'Application Status' column"


# ---------------------------------------------------------------------------
# 8. No secrets, no DATABASE_URL, no JWT, no API keys
# ---------------------------------------------------------------------------

def test_no_database_url_in_copilot() -> None:
    assert "DATABASE_URL" not in _copilot_src(), \
        "Copilot must not contain DATABASE_URL"


def test_no_jwt_in_copilot() -> None:
    src = _copilot_src().lower()
    assert "jwt_secret" not in src and "secret_key" not in src, \
        "Copilot must not contain JWT secret"


def test_no_vapi_key_in_copilot() -> None:
    src = _copilot_src().lower()
    assert "vapi_api_key" not in src and "vapi_secret" not in src, \
        "Copilot must not contain Vapi API key"


def test_no_hardcoded_api_keys_in_copilot() -> None:
    pattern = re.compile(r'["\']([A-Za-z0-9_\-]{32,})["\']')
    matches = pattern.findall(_copilot_src())
    assert not matches, f"Copilot must not contain hardcoded API key-like strings: {matches[:2]}"


def test_no_database_url_in_md_files() -> None:
    assert "DATABASE_URL" not in _all_career_md(), \
        "Career markdown files must not contain DATABASE_URL"


def test_no_jwt_in_md_files() -> None:
    content = _all_career_md().lower()
    assert "jwt_secret" not in content, "Career markdown files must not contain jwt_secret"


# ---------------------------------------------------------------------------
# 9. Motivation letter templates
# ---------------------------------------------------------------------------

def test_motivation_letter_en_has_placeholders() -> None:
    content = _read(_ML_EN)
    assert "{{" in content, "EN motivation letter template must use {{PLACEHOLDER}} format"


def test_motivation_letter_de_has_placeholders() -> None:
    content = _read(_ML_DE)
    assert "{{" in content, "DE motivation letter template must use {{PLATZHALTER}} format"


def test_motivation_letter_en_has_praxismed_reference() -> None:
    content = _read(_ML_EN)
    assert "PraxisMed" in content or "relevant project" in content.lower(), \
        "EN motivation letter template must reference PraxisMed or relevant project"


def test_motivation_letter_en_says_no_false_claims() -> None:
    content = _read(_ML_EN).lower()
    assert "false" in content or "no false" in content, \
        "EN motivation letter template must warn against false claims"


def test_motivation_letter_de_says_no_false_claims() -> None:
    content = _read(_ML_DE).lower()
    assert "false" in content or "keine falschen" in content or "falsch" in content, \
        "DE motivation letter template must warn against false claims"


def test_motivation_letter_en_has_checklist() -> None:
    content = _read(_ML_EN)
    assert "checklist" in content.lower() or "[ ]" in content, \
        "EN motivation letter template must have a pre-send checklist"


# ---------------------------------------------------------------------------
# 10. Daily guide / submission guide safety
# ---------------------------------------------------------------------------

def test_readme_says_no_auto_apply() -> None:
    content = _read(_README).lower()
    assert "no auto-apply" in content or "no automation" in content or "manual" in content, \
        "0_READ_ME_FIRST.md must explicitly state no auto-apply"


def test_readme_mentions_vienna_priority() -> None:
    content = _read(_README)
    assert "Vienna" in content or "Wien" in content, \
        "0_READ_ME_FIRST.md must mention Vienna as priority location"


def test_readme_mentions_full_time() -> None:
    content = _read(_README).lower()
    assert "full-time" in content or "fulltime" in content or "vollzeit" in content, \
        "0_READ_ME_FIRST.md must mention full-time as priority"


def test_submission_guide_says_no_auto_submit() -> None:
    content = _read(_SUBGUIDE).lower()
    assert "no auto-submit" in content or "manual submission" in content, \
        "5_MANUAL_SUBMISSION_GUIDE.md must state no auto-submit"


def test_submission_guide_has_work_auth_note() -> None:
    content = _read(_SUBGUIDE)
    assert "student residence permit" in content or "work authorization" in content.lower() or \
           "Rot-Weiß-Rot" in content or "RWR" in content, \
        "Manual submission guide must include work authorization guidance"


def test_submission_guide_says_no_fake_citizenship() -> None:
    content = _read(_SUBGUIDE).lower()
    assert "fabricate" in content or "do not fake" in content or "honestly" in content or \
           "citizenship" in content, \
        "Submission guide must warn against faking citizenship/work status"
